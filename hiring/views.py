import csv
import json
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
import uuid  # Add this import
from django.views.decorators.http import require_http_methods  # <-- ADD THIS LINE
from django.http import JsonResponse, HttpResponseForbidden
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
import psutil  # for system health check
# hiring/views.py - Add this at the top with other imports
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q
from django.utils import timezone
from decimal import Decimal 
from django.core.cache import cache
from rest_framework.response import Response
from rest_framework import status
import time
from rest_framework.parsers import FormParser
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.parsers import MultiPartParser, JSONParser
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import get_object_or_404, render
from django.db.models import Count, Q, Avg
from django.db import connection
from datetime import datetime, timedelta
# ===== ADD THE ADMIN_REQUIRED DECORATOR HERE =====
from functools import wraps
from django.http import HttpResponseForbidden
from django.shortcuts import redirect
# ===== END ======
from realestate.models import (
    Property, 
    PropertyType, 
    PropertyCategory, 
    PropertyFeature,  # ADD THIS - you have this model
    Booking,          # ADD THIS - you have this model
    Room,             # ADD THIS - you have this model
    PropertyAnalytics, # ADD THIS - you have this model
    PropertyReview,   # ADD THIS - you have this model
    Wishlist,         # ADD THIS - you have this model
    BookingInquiry,   # ADD THIS - you have this model
    AvailabilityCalendar, # ADD THIS - you have this model
    DriverLocation,   # ADD THIS - you have this model
)
from realestate.models import PropertyAnalytics
from hiring.models import CustomUser, ApplicantProfile, BusinessProfile
import json
from django.http import JsonResponse
from django.utils import timezone
from django.db import transaction
import logging
from .models import *
from django.views.decorators.csrf import csrf_exempt


logger = logging.getLogger(__name__)

# Import utils first to avoid circular imports
from .utils import calculate_profile_completeness

# Import models
from .models import (
    CustomUser, ApplicantProfile, JobListing, Application, Alert, 
    Skill, EmploymentHistory, Education, Document, 
    NotificationPreference, SentNotification, JobAlert, BusinessProfile,
    Industry, CompanySize, JobCategory
)

# Import serializers
from .serializers import *

# ===== FIXED ADMIN ACCESS CONTROL FUNCTIONS =====
def has_admin_access(user):
    """
    Check if user has admin access:
    - Superusers and staff have full access
    - Business admins (user_type='admin') have business-level access
    - Applicants (user_type='applicant') have no admin access
    """
    if not user.is_authenticated:
        return False
    
    # Superusers and staff have full admin access
    if user.is_superuser or user.is_staff:
        return True
    
    # Business admins have business-level access
    if user.user_type == 'admin':
        return True
    
    return False

def has_business_access(user):
    """
    Check if user has business-specific access (for business data filtering)
    """
    return user.is_authenticated and user.user_type == 'admin'

def has_superuser_access(user):
    """
    Check if user has superuser access (for system-wide admin functions)
    """
    return user.is_authenticated and (user.is_superuser or user.is_staff)

def admin_required(view_func):
    """
    Decorator for views that checks that the user is logged in and is an admin/staff,
    or has business admin access (user_type='admin').
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return redirect('/')
        
        # Check if user has admin access (superuser, staff, or business admin)
        if not has_admin_access(request.user):
            return HttpResponseForbidden("You don't have permission to access this page.")
        
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def superuser_required(view_func):
    """
    Decorator for views that require superuser access
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/')
        
        if not has_superuser_access(request.user):
            return HttpResponseForbidden("You don't have permission to access this page.")
        
        return view_func(request, *args, **kwargs)
    return _wrapped_view
# ===== END OF UPDATED ADMIN_REQUIRED DECORATOR =====

# Create a simple notification service inline to avoid missing imports
class NotificationService:
    @staticmethod
    def send_application_submission(application):
        """
        Send notification when application is submitted
        """
        try:
            # Create an alert for the applicant
            Alert.objects.create(
                applicant=application.applicant,
                title="Application Submitted",
                message=f"Your application for {application.job_listing.title} at {application.job_listing.company_name} has been submitted successfully."
            )
            
            # Create a sent notification record
            SentNotification.objects.create(
                applicant=application.applicant,
                notification_type='application_submitted',
                subject="Application Submitted Successfully",
                message=f"Your application for {application.job_listing.title} has been received and is under review.",
                sent_via='in_app'
            )
            
            logger.info(f"Notification sent for application {application.id}")
            
        except Exception as e:
            logger.error(f"Error sending notification for application {application.id}: {str(e)}")
            # Don't raise the exception - we don't want to break the application flow
    
    @staticmethod
    def send_job_alert(applicant, job_listing):
        """
        Send job alert notification
        """
        try:
            Alert.objects.create(
                applicant=applicant,
                title="New Job Match",
                message=f"A new job matching your criteria: {job_listing.title} at {job_listing.company_name}"
            )
            
            SentNotification.objects.create(
                applicant=applicant,
                notification_type='job_alert',
                subject="New Job Opportunity",
                message=f"We found a job that matches your profile: {job_listing.title}",
                sent_via='in_app'
            )
            
        except Exception as e:
            logger.error(f"Error sending job alert: {str(e)}")
    
    @staticmethod
    def send_profile_reminder(applicant):
        """
        Send profile completion reminder
        """
        try:
            if applicant.profile_completeness < 70:
                Alert.objects.create(
                    applicant=applicant,
                    title="Complete Your Profile",
                    message=f"Your profile is {applicant.profile_completeness}% complete. Complete it to increase your chances of getting hired."
                )
        except Exception as e:
            logger.error(f"Error sending profile reminder: {str(e)}")

# HTML PAGE VIEWS

# ===== PROPERTY MANAGEMENT VIEWS (for admin/business users) =====
def video_feed_page(request):
    """Render video feed page"""
    return render(request, 'hiring/video_feed.html')


    

# ===== PROPERTY TYPES API =====
def get_property_types_api(request):
    """API endpoint to get property types for filters"""
    try:
        types = PropertyType.objects.filter(is_active=True)
        type_list = [{
            'id': str(t.id),
            'name': t.name,
            'category': t.category.name if t.category else None,
            'size_classification': t.size_classification,
        } for t in types]
        
        return JsonResponse({
            'success': True,
            'types': type_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ===== PROPERTY TYPE ADD (API) =====
@login_required
def property_type_add_api(request):
    """
    Add a new property type via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            category_id = request.POST.get('category')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Name is required'})
            
            # Check if type already exists
            if PropertyType.objects.filter(name__iexact=name).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'Property type "{name}" already exists'
                })
            
            # Create the property type
            type_obj = PropertyType.objects.create(
                name=name,
                category_id=category_id if category_id else None,
                is_active=True
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Property type "{name}" created successfully',
                'type': {
                    'id': type_obj.id,
                    'name': type_obj.name,
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


# ===== ADD PROPERTY =====
@login_required
def property_add(request):
    """
    Add a new property
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    if request.method == 'POST':
        try:
            # Get form data
            title = request.POST.get('title')
            description = request.POST.get('description')
            property_type_id = request.POST.get('property_type')
            listing_type = request.POST.get('listing_type')
            status = request.POST.get('status', 'available')
            
            # Location
            address = request.POST.get('address')
            city = request.POST.get('city')
            state = request.POST.get('state', '')
            country = request.POST.get('country', 'South Africa')
            postal_code = request.POST.get('postal_code', '')
            
            # Pricing
            base_price = request.POST.get('base_price')
            price_currency = request.POST.get('price_currency', 'ZAR')
            booking_unit = request.POST.get('booking_unit', 'day')
            
            # Details
            bedrooms = request.POST.get('bedrooms', 0)
            bathrooms = request.POST.get('bathrooms', 0)
            garages = request.POST.get('garages', 0)
            total_area = request.POST.get('total_area')
            
            # Booking settings
            booking_mode = request.POST.get('booking_mode', 'traditional')
            minimum_stay = request.POST.get('minimum_stay', 1)
            
            # Features
            features = request.POST.getlist('features')
            
            # Create property
            property_obj = Property.objects.create(
                title=title,
                description=description,
                property_type_id=property_type_id,
                listing_type=listing_type,
                status=status,
                address=address,
                city=city,
                state=state,
                country=country,
                postal_code=postal_code,
                base_price=base_price,
                price_currency=price_currency,
                booking_unit=booking_unit,
                bedrooms=bedrooms,
                bathrooms=bathrooms,
                garages=garages,
                total_area=total_area,
                booking_mode=booking_mode,
                minimum_stay=minimum_stay,
                company=request.user.business_profile if hasattr(request.user, 'business_profile') else None,
                owner=request.user,
                listing_agent=request.user,
            )
            
            # Add features
            if features:
                property_obj.features.set(features)
            
            return JsonResponse({
                'success': True,
                'message': 'Property created successfully',
                'property_id': str(property_obj.id)
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    # GET request - show form
    context = {
        'property_types': PropertyType.objects.filter(is_active=True).order_by('name'),
        'categories': PropertyCategory.objects.filter(is_active=True),
        'features': PropertyFeature.objects.filter(is_active=True),
        'listing_types': Property.LISTING_TYPES,
        'status_choices': Property.STATUS_CHOICES,
        'booking_modes': Property.BOOKING_MODES,
        'booking_units': Property.BOOKING_UNITS,
    }
    
    return render(request, 'hiring/property_add.html', context)

@login_required
def property_type_add_api(request):
    """
    Add a new property type via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category')
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
         
        # Check if type already exists
        if PropertyType.objects.filter(name__iexact=name).exists():
            return JsonResponse({
                'success': False, 
                'error': f'Property type "{name}" already exists'
            }, status=400)
        
        # Get or create default category if none provided
        category = None
        if category_id:
            try:
                category = PropertyCategory.objects.get(id=category_id)
            except PropertyCategory.DoesNotExist:
                pass
        
        # If no category, get or create default
        if not category:
            category, _ = PropertyCategory.objects.get_or_create(
                name='General',
                defaults={'is_active': True}
            )
        
        # Create the property type
        type_obj = PropertyType.objects.create(
            name=name,
            category=category,
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Property type "{name}" created successfully',
            'type': {
                'id': type_obj.id,
                'name': type_obj.name,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)
 
# ===== FEATURE MANAGEMENT VIEWS =====

@login_required
def features_manage(request):
    """
    View for managing property features (Business portal)
    """
    if not (request.user.user_type == 'admin' or request.user.is_superuser):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    # Get all features - show custom features created by this user and all active features
    if request.user.is_superuser:
        features = PropertyFeature.objects.all().order_by('category', 'name')
    else:
        # Show features created by this user OR global features (is_custom=False)
        features = PropertyFeature.objects.filter(
            Q(created_by=request.user) | Q(is_custom=False) | Q(is_custom=True, created_by=request.user)
        ).distinct().order_by('category', 'name')
    
    # Get filter parameters
    category_filter = request.GET.get('category', '')
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    # Apply filters
    if category_filter:
        features = features.filter(category=category_filter)
    if search_query:
        features = features.filter(
            Q(name__icontains=search_query) |
            Q(icon__icontains=search_query)
        )
    if status_filter == 'active':
        features = features.filter(is_active=True)
    elif status_filter == 'inactive':
        features = features.filter(is_active=False)
    
    # Counts
    total_count = features.count()
    active_count = features.filter(is_active=True).count()
    custom_count = features.filter(is_custom=True).count()
    global_count = features.filter(is_custom=False).count()
    
    # Category counts
    category_counts = {}
    for cat_code, cat_label in PropertyFeature.FEATURE_CATEGORIES:
        category_counts[cat_code] = features.filter(category=cat_code).count()
    
    # Pagination
    paginator = Paginator(features, 20)
    page = request.GET.get('page')
    
    try:
        features_page = paginator.page(page)
    except PageNotAnInteger:
        features_page = paginator.page(1)
    except EmptyPage:
        features_page = paginator.page(paginator.num_pages)
    
    context = {
        'features': features_page,
        'total_count': total_count,
        'active_count': active_count,
        'custom_count': custom_count,
        'global_count': global_count,
        'category_counts': category_counts,
        'category_filter': category_filter,
        'search_query': search_query,
        'status_filter': status_filter,
        'feature_categories': PropertyFeature.FEATURE_CATEGORIES,
        'user': request.user,
    }
    
    return render(request, 'hiring/features_manage.html', context)


@login_required
@require_http_methods(["POST"])
def feature_add_api(request):
    """
    Add a new property feature via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        name = request.POST.get('name', '').strip()
        category = request.POST.get('category', 'other')
        icon = request.POST.get('icon', '')
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Feature name is required'}, status=400)
        
        # Check if feature already exists
        if PropertyFeature.objects.filter(name__iexact=name).exists():
            return JsonResponse({
                'success': False, 
                'error': f'Feature "{name}" already exists'
            }, status=400)
        
        # Create the property feature
        feature_obj = PropertyFeature.objects.create(
            name=name,
            icon=icon or 'fas fa-tag',  # Default icon
            category=category,
            is_custom=True,
            is_active=True,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Feature "{name}" created successfully',
            'feature': {
                'id': feature_obj.id,
                'name': feature_obj.name,
                'category': feature_obj.get_category_display(),
                'category_code': feature_obj.category,
                'icon': feature_obj.icon,
                'is_custom': feature_obj.is_custom,
                'is_active': feature_obj.is_active,
                'created_at': feature_obj.created_at.strftime('%Y-%m-%d %H:%M'),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def feature_edit_api(request, feature_id):
    """
    Edit an existing property feature via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        feature = get_object_or_404(PropertyFeature, id=feature_id)
        
        # Check permission - only creator or superuser can edit
        if not request.user.is_superuser and feature.created_by != request.user:
            return JsonResponse({
                'success': False, 
                'error': 'You can only edit features you created'
            }, status=403)
        
        name = request.POST.get('name', '').strip()
        category = request.POST.get('category', 'other')
        icon = request.POST.get('icon', '')
        is_active = request.POST.get('is_active') == 'true'
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Feature name is required'}, status=400)
        
        # Check if name conflicts with another feature
        if PropertyFeature.objects.filter(name__iexact=name).exclude(id=feature_id).exists():
            return JsonResponse({
                'success': False, 
                'error': f'Feature "{name}" already exists'
            }, status=400)
        
        # Update feature
        feature.name = name
        feature.category = category
        if icon:
            feature.icon = icon
        feature.is_active = is_active
        feature.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Feature "{name}" updated successfully',
            'feature': {
                'id': feature.id,
                'name': feature.name,
                'category': feature.get_category_display(),
                'category_code': feature.category,
                'icon': feature.icon,
                'is_custom': feature.is_custom,
                'is_active': feature.is_active,
                'created_at': feature.created_at.strftime('%Y-%m-%d %H:%M'),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def feature_delete_api(request, feature_id):
    """
    Delete a property feature via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        feature = get_object_or_404(PropertyFeature, id=feature_id)
        
        # Check permission - only creator or superuser can delete
        if not request.user.is_superuser and feature.created_by != request.user:
            return JsonResponse({
                'success': False, 
                'error': 'You can only delete features you created'
            }, status=403)
        
        # Check if feature is in use
        if feature.properties.exists():
            return JsonResponse({
                'success': False, 
                'error': f'Cannot delete "{feature.name}" - it is being used by {feature.properties.count()} properties'
            }, status=400)
        
        feature_name = feature.name
        feature.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Feature "{feature_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def feature_toggle_status_api(request, feature_id):
    """
    Toggle feature active status via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        feature = get_object_or_404(PropertyFeature, id=feature_id)
        
        # Check permission
        if not request.user.is_superuser and feature.created_by != request.user:
            return JsonResponse({
                'success': False, 
                'error': 'You can only modify features you created'
            }, status=403)
        
        feature.is_active = not feature.is_active
        feature.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Feature "{feature.name}" {"activated" if feature.is_active else "deactivated"}',
            'is_active': feature.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)


@login_required
def get_feature_detail_api(request, feature_id):
    """
    Get feature details for editing via AJAX
    """
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        feature = get_object_or_404(PropertyFeature, id=feature_id)
        
        # Check permission
        if not request.user.is_superuser and feature.created_by != request.user and not feature.is_custom:
            return JsonResponse({
                'success': False, 
                'error': 'You can only view features you created'
            }, status=403)
        
        return JsonResponse({
            'success': True,
            'feature': {
                'id': feature.id,
                'name': feature.name,
                'category': feature.category,
                'icon': feature.icon,
                'is_custom': feature.is_custom,
                'is_active': feature.is_active,
                'created_by': feature.created_by.username if feature.created_by else None,
                'created_at': feature.created_at.strftime('%Y-%m-%d %H:%M'),
                'properties_count': feature.properties.count(),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)

        
#property manage
@login_required
def properties_manage(request):
    """View for managing properties (Business portal)"""
    if not (request.user.user_type == 'admin' or request.user.is_superuser):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    properties = Property.objects.filter(owner=request.user).order_by('-created_at')
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('property_type', '')
    
    # Apply filters
    if status_filter:
        properties = properties.filter(status=status_filter)
    if search_query:
        properties = properties.filter(
            Q(title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(address__icontains=search_query)
        )
    if type_filter:
        properties = properties.filter(property_type_id=type_filter)
    
    # Calculate counts
    total_count = Property.objects.filter(owner=request.user).count()
    available_count = Property.objects.filter(owner=request.user, status='available').count()
    booked_count = Property.objects.filter(owner=request.user, status='booked').count()
    maintenance_count = Property.objects.filter(owner=request.user, status='maintenance').count()
    
    # Pagination
    paginator = Paginator(properties, 12)
    page = request.GET.get('page')
    
    try:
        properties_page = paginator.page(page)
    except PageNotAnInteger:
        properties_page = paginator.page(1)
    except EmptyPage:
        properties_page = paginator.page(paginator.num_pages)
    
    context = {
        'properties': properties_page,
        'total_count': total_count,
        'available_count': available_count,
        'booked_count': booked_count,
        'maintenance_count': maintenance_count,
        'status_filter': status_filter,
        'search_query': search_query,
        'type_filter': type_filter,
        'status_choices': Property.STATUS_CHOICES,
        'listing_types': Property.LISTING_TYPES,
        'property_types': PropertyType.objects.filter(is_active=True).order_by('name'),
        'user': request.user,
    }
    
    return render(request, 'hiring/properties_manage.html', context)



@login_required
def property_edit(request, property_id):
    """
    Edit an existing property
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    # Check permission
    if property_obj.company and property_obj.company.user != request.user and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to edit this property")
    
    if request.method == 'POST':
        try:
            # Update property
            property_obj.title = request.POST.get('title', property_obj.title)
            property_obj.description = request.POST.get('description', property_obj.description)
            property_obj.property_type_id = request.POST.get('property_type', property_obj.property_type_id)
            property_obj.listing_type = request.POST.get('listing_type', property_obj.listing_type)
            property_obj.status = request.POST.get('status', property_obj.status)
            
            # Location
            property_obj.address = request.POST.get('address', property_obj.address)
            property_obj.city = request.POST.get('city', property_obj.city)
            property_obj.state = request.POST.get('state', property_obj.state)
            property_obj.country = request.POST.get('country', property_obj.country)
            property_obj.postal_code = request.POST.get('postal_code', property_obj.postal_code)
            
            # Pricing
            property_obj.base_price = request.POST.get('base_price', property_obj.base_price)
            property_obj.price_currency = request.POST.get('price_currency', property_obj.price_currency)
            property_obj.booking_unit = request.POST.get('booking_unit', property_obj.booking_unit)
            
            # Details
            property_obj.bedrooms = request.POST.get('bedrooms', property_obj.bedrooms)
            property_obj.bathrooms = request.POST.get('bathrooms', property_obj.bathrooms)
            property_obj.garages = request.POST.get('garages', property_obj.garages)
            property_obj.total_area = request.POST.get('total_area', property_obj.total_area)
            
            # Booking settings
            property_obj.booking_mode = request.POST.get('booking_mode', property_obj.booking_mode)
            property_obj.minimum_stay = request.POST.get('minimum_stay', property_obj.minimum_stay)
            
            # Features
            features = request.POST.getlist('features')
            property_obj.features.set(features)
            
            property_obj.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Property updated successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    # GET request - show form with data
    context = {
        'property': property_obj,
        'property_types': PropertyType.objects.filter(is_active=True),
        'categories': PropertyCategory.objects.filter(is_active=True),
        'features': PropertyFeature.objects.filter(is_active=True),
        'listing_types': Property.LISTING_TYPES,
        'status_choices': Property.STATUS_CHOICES,
        'booking_modes': Property.BOOKING_MODES,
        'booking_units': Property.BOOKING_UNITS,
        'selected_features': property_obj.features.values_list('id', flat=True),
    }
    
    return render(request, 'hiring/property_edit.html', context)


@login_required
def property_delete(request, property_id):
    """
    Delete a property (soft delete)
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    # Check permission
    if property_obj.company and property_obj.company.user != request.user and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        property_obj.is_active = False
        property_obj.save()
        return JsonResponse({'success': True, 'message': 'Property deleted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def property_analytics(request):
    """
    View property analytics for business users
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    properties = Property.objects.filter(
        company__user=request.user,
        is_active=True
    ).select_related('property_type')
    
    # Get analytics for each property
    property_analytics = []
    total_views = 0
    total_bookings = 0
    total_revenue = Decimal('0')
    
    for prop in properties:
        try:
            analytics = PropertyAnalytics.objects.get(property=prop)
            property_analytics.append({
                'property': prop,
                'analytics': analytics
            })
            total_views += analytics.total_views
            total_bookings += analytics.total_bookings
            total_revenue += analytics.total_revenue
        except PropertyAnalytics.DoesNotExist:
            property_analytics.append({
                'property': prop,
                'analytics': None
            })
    
    context = {
        'property_analytics': property_analytics,
        'total_properties': len(property_analytics),
        'total_views': total_views,
        'total_bookings': total_bookings,
        'total_revenue': total_revenue,
        'avg_views': total_views // len(property_analytics) if property_analytics else 0,
    }
    
    return render(request, 'hiring/property_analytics.html', context)


@login_required
def property_bookings_manage(request):
    """
    Manage bookings for business properties
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    # Get all bookings for properties owned by this business
    bookings = Booking.objects.filter(
        property__company__user=request.user
    ).select_related('property', 'guest', 'room').order_by('-created_at')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        bookings = bookings.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(bookings, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Stats
    total_count = bookings.count()
    pending_count = bookings.filter(status='pending').count()
    confirmed_count = bookings.filter(status='confirmed').count()
    completed_count = bookings.filter(status='completed').count()
    
    context = {
        'bookings': page_obj,
        'total_count': total_count,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'completed_count': completed_count,
        'status_filter': status_filter,
        'status_choices': Booking.BOOKING_STATUS,
    }
    
    return render(request, 'hiring/property_bookings.html', context)


@login_required
def property_booking_update(request, booking_id):
    """
    Update booking status
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    booking = get_object_or_404(Booking, id=booking_id)
    
    # Check permission
    if booking.property.company and booking.property.company.user != request.user and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in dict(Booking.BOOKING_STATUS).keys():
            return JsonResponse({'error': 'Invalid status'}, status=400)
        
        booking.status = new_status
        booking.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Booking status updated to {booking.get_status_display()}'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def property_rooms_manage(request, property_id):
    """
    Manage rooms for a specific property
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    # Check permission
    if property_obj.company and property_obj.company.user != request.user and not request.user.is_superuser:
        return HttpResponseForbidden("You don't have permission to manage rooms for this property")
    
    rooms = Room.objects.filter(property=property_obj, is_active=True)
    
    context = {
        'property': property_obj,
        'rooms': rooms,
        'room_types': Room.ROOM_TYPES,
        'room_statuses': Room.ROOM_STATUS,
    }
    
    return render(request, 'hiring/property_rooms.html', context)


@login_required
def property_room_add(request, property_id):
    """
    Add a room to a property
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    # Check permission
    if property_obj.company and property_obj.company.user != request.user and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        room = Room.objects.create(
            property=property_obj,
            room_number=data.get('room_number', ''),
            room_name=data.get('room_name', ''),
            room_type=data.get('room_type', 'single'),
            room_status=data.get('room_status', 'available'),
            capacity=data.get('capacity', 1),
            bed_count=data.get('bed_count', 1),
            size_sq_meters=data.get('size_sq_meters'),
            price_per_night=data.get('price_per_night'),
            description=data.get('description', ''),
            is_active=True,
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Room added successfully',
            'room_id': str(room.id)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def property_room_delete(request, room_id):
    """
    Delete a room
    """
    # Only allow admin users
    if request.user.user_type != 'admin' and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    room = get_object_or_404(Room, id=room_id)
    
    # Check permission
    if room.property.company and room.property.company.user != request.user and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        room.is_active = False
        room.save()
        return JsonResponse({'success': True, 'message': 'Room deleted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def property_detail(request, property_id):
    """
    View property details (public)
    """
    property_obj = get_object_or_404(
        Property.objects.select_related('property_type', 'company', 'owner'),
        id=property_id,
        is_active=True
    )
    
    # Increment view count
    property_obj.views_count += 1
    property_obj.save(update_fields=['views_count'])
    
    # Get rooms
    rooms = Room.objects.filter(property=property_obj, is_active=True)
    
    # Get reviews
    reviews = PropertyReview.objects.filter(
        property=property_obj,
        is_approved=True
    ).select_related('user').order_by('-created_at')
    
    # Get similar properties
    similar_properties = Property.objects.filter(
        property_type=property_obj.property_type,
        is_active=True,
        status='available'
    ).exclude(id=property_obj.id)[:4]
    
    context = {
        'property': property_obj,
        'rooms': rooms,
        'reviews': reviews,
        'similar_properties': similar_properties,
        'is_wishlisted': False,
        'user_review': None,
        'can_manage': request.user.is_authenticated and (
            (property_obj.company and property_obj.company.user == request.user) or
            request.user.is_superuser
        ),
    }
    
    # Check if user has this property in wishlist
    if request.user.is_authenticated:
        context['is_wishlisted'] = Wishlist.objects.filter(
            user=request.user,
            properties=property_obj
        ).exists()
        context['user_review'] = PropertyReview.objects.filter(
            user=request.user,
            property=property_obj
        ).first()
    
    return render(request, 'hiring/property_detail.html', context)


# ===== API ENDPOINTS =====

@login_required
@require_http_methods(["POST"])
def toggle_wishlist(request):
    """
    API endpoint to toggle property in wishlist
    """
    try:
        data = json.loads(request.body)
        property_id = data.get('property_id')
        
        if not property_id:
            return JsonResponse({'error': 'Property ID required'}, status=400)
        
        property_obj = get_object_or_404(Property, id=property_id)
        
        # Get or create default wishlist
        wishlist, created = Wishlist.objects.get_or_create(
            user=request.user,
            is_default=True,
            defaults={'name': 'Default Wishlist'}
        )
        
        # Toggle property in wishlist
        if property_obj in wishlist.properties.all():
            wishlist.properties.remove(property_obj)
            is_wishlisted = False
            message = 'Property removed from wishlist'
        else:
            wishlist.properties.add(property_obj)
            is_wishlisted = True
            message = 'Property added to wishlist'
        
        return JsonResponse({
            'success': True,
            'is_wishlisted': is_wishlisted,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def add_review(request):
    """
    API endpoint to add a review
    """
    try:
        data = json.loads(request.body)
        property_id = data.get('property_id')
        
        if not property_id:
            return JsonResponse({'error': 'Property ID required'}, status=400)
        
        property_obj = get_object_or_404(Property, id=property_id)
        
        # Check if user already reviewed this property
        existing_review = PropertyReview.objects.filter(
            user=request.user,
            property=property_obj
        ).first()
        
        if existing_review:
            # Update existing review
            existing_review.overall_rating = data.get('rating', existing_review.overall_rating)
            existing_review.review_text = data.get('review_text', existing_review.review_text)
            existing_review.review_title = data.get('review_title', existing_review.review_title)
            existing_review.save()
            review = existing_review
        else:
            # Create new review
            review = PropertyReview.objects.create(
                property=property_obj,
                user=request.user,
                overall_rating=data.get('rating', 5),
                review_text=data.get('review_text', ''),
                review_title=data.get('review_title', ''),
                is_approved=False  # Requires moderation
            )
        
        return JsonResponse({
            'success': True,
            'review_id': str(review.id),
            'message': 'Review submitted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ===== API ENDPOINTS FOR PROPERTY DATA =====

# ===== YOUR EXISTING HOME PAGE FUNCTION =====
def home_page(request):
    """Render home page with feed and real estate properties"""
    
    # Get real estate properties for the home page
    properties = Property.objects.filter(is_active=True, status='available')[:10]
    featured_properties = Property.objects.filter(is_active=True, is_featured=True)[:6]
    property_types = PropertyType.objects.filter(is_active=True)
    property_categories = PropertyCategory.objects.filter(is_active=True)
    
    context = {
        'page_title': 'PropNest - Find Your Dream Property & Jobs',
        'show_feed': True,
        'user_authenticated': request.user.is_authenticated,
        'user_type': request.user.user_type if request.user.is_authenticated else None,
        # Real estate data
        'properties': properties,
        'featured_properties': featured_properties,
        'property_types': property_types,
        'property_categories': property_categories,
        'total_properties': Property.objects.filter(is_active=True).count(),
        'available_properties': Property.objects.filter(is_active=True, status='available').count(),
    }
    return render(request, 'hiring/home.html', context)



#Get properties
def get_properties_api(request):
    """API endpoint to get properties for the home page"""
    try:
        properties = Property.objects.filter(is_active=True, status='available').order_by('-created_at')
        
        # Apply filters if provided
        property_type = request.GET.get('property_type')
        if property_type:
            properties = properties.filter(property_type_id=property_type)
        
        listing_type = request.GET.get('listing_type')
        if listing_type:
            properties = properties.filter(listing_type=listing_type)
        
        min_price = request.GET.get('min_price')
        if min_price:
            properties = properties.filter(base_price__gte=min_price)
        
        max_price = request.GET.get('max_price')
        if max_price:
            properties = properties.filter(base_price__lte=max_price)
        
        bedrooms = request.GET.get('bedrooms')
        if bedrooms:
            properties = properties.filter(bedrooms__gte=int(bedrooms))
        
        # Serialize properties
        property_list = []
        for prop in properties[:50]:  # Limit to 50 for performance
            property_list.append({
                'id': str(prop.id),
                'title': prop.title,
                'description': prop.description,
                'price': str(prop.base_price),
                'price_currency': prop.price_currency,
                'bedrooms': prop.bedrooms,
                'bathrooms': prop.bathrooms,
                'garages': prop.garages,
                'city': prop.city,
                'country': prop.country,
                'status': prop.status,
                'listing_type': prop.listing_type,
                'is_featured': prop.is_featured,
                'is_premium': prop.is_premium,
                'is_online': prop.is_online,
                'main_image_url': prop.get_main_image_url(),
                'property_type': prop.property_type.name if prop.property_type else None,
                'property_type_id': str(prop.property_type.id) if prop.property_type else None,
                'created_at': prop.created_at.isoformat(),
                'company_name': prop.company.company_name if prop.company else None,
            })
        
        return JsonResponse({
            'success': True,
            'properties': property_list,
            'total': len(property_list)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def get_featured_properties_api(request):
    """API endpoint to get featured properties"""
    try:
        properties = Property.objects.filter(is_active=True, is_featured=True).order_by('-created_at')[:10]
        
        property_list = []
        for prop in properties:
            property_list.append({
                'id': str(prop.id),
                'title': prop.title,
                'description': prop.description,
                'price': str(prop.base_price),
                'price_currency': prop.price_currency,
                'bedrooms': prop.bedrooms,
                'bathrooms': prop.bathrooms,
                'garages': prop.garages,
                'city': prop.city,
                'country': prop.country,
                'status': prop.status,
                'listing_type': prop.listing_type,
                'is_featured': prop.is_featured,
                'is_premium': prop.is_premium,
                'is_online': prop.is_online,
                'main_image_url': prop.get_main_image_url(),
                'created_at': prop.created_at.isoformat(),
                'company_name': prop.company.company_name if prop.company else None,
            })
        
        return JsonResponse({
            'success': True,
            'properties': property_list,
            'total': len(property_list)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def get_property_types_api(request):
    """API endpoint to get property types for filters"""
    try:
        types = PropertyType.objects.filter(is_active=True)
        type_list = [{
            'id': str(t.id),
            'name': t.name,
            'category': t.category.name if t.category else None,
            'size_classification': t.size_classification,
        } for t in types]
        
        return JsonResponse({
            'success': True,
            'types': type_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@csrf_exempt
def book_property_api(request, property_id):
    """API endpoint to book a property"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        property_obj = Property.objects.get(id=property_id, is_active=True)
        
        if property_obj.status != 'available':
            return JsonResponse({
                'success': False,
                'error': 'Property is not available for booking'
            }, status=400)
        
        # Create booking
        from datetime import datetime, timedelta
        check_in = datetime.now()
        check_out = check_in + timedelta(days=1)
        
        booking = Booking.objects.create(
            property=property_obj,
            guest=request.user,
            check_in=check_in,
            check_out=check_out,
            duration_days=1,
            subtotal=property_obj.base_price or 0,
            total_amount=property_obj.base_price or 0,
            status='pending',
            booking_mode='instant'
        )
        
        # Update property status
        property_obj.status = 'booked'
        property_obj.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Property booked successfully',
            'booking_id': str(booking.id),
            'booking_reference': booking.booking_reference
        })
        
    except Property.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Property not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    

def get_properties_api(request):
    """API endpoint to get properties for the home page"""
    try:
        properties = Property.objects.filter(is_active=True).order_by('-created_at')
        
        # Apply filters if provided
        property_type = request.GET.get('property_type')
        if property_type:
            properties = properties.filter(property_type_id=property_type)
        
        listing_type = request.GET.get('listing_type')
        if listing_type:
            properties = properties.filter(listing_type=listing_type)
        
        min_price = request.GET.get('min_price')
        if min_price:
            properties = properties.filter(base_price__gte=min_price)
        
        max_price = request.GET.get('max_price')
        if max_price:
            properties = properties.filter(base_price__lte=max_price)
        
        # Serialize properties
        property_list = []
        for prop in properties[:50]:  # Limit to 50 for performance
            property_list.append({
                'id': str(prop.id),
                'title': prop.title,
                'description': prop.description,
                'price': str(prop.base_price),
                'price_currency': prop.price_currency,
                'bedrooms': prop.bedrooms,
                'bathrooms': prop.bathrooms,
                'garages': prop.garages,
                'city': prop.city,
                'country': prop.country,
                'status': prop.status,
                'listing_type': prop.listing_type,
                'is_featured': prop.is_featured,
                'is_premium': prop.is_premium,
                'is_online': prop.is_online,
                'main_image_url': prop.get_main_image_url(),
                'property_type': prop.property_type.name if prop.property_type else None,
                'property_type_id': str(prop.property_type.id) if prop.property_type else None,
                'created_at': prop.created_at.isoformat(),
                'company_name': prop.company.company_name if prop.company else None,
            })
        
        return JsonResponse({
            'success': True,
            'properties': property_list,
            'total': len(property_list)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def get_featured_properties_api(request):
    """API endpoint to get featured properties"""
    try:
        properties = Property.objects.filter(is_active=True, is_featured=True).order_by('-created_at')[:10]
        
        property_list = []
        for prop in properties:
            property_list.append({
                'id': str(prop.id),
                'title': prop.title,
                'description': prop.description,
                'price': str(prop.base_price),
                'price_currency': prop.price_currency,
                'bedrooms': prop.bedrooms,
                'bathrooms': prop.bathrooms,
                'garages': prop.garages,
                'city': prop.city,
                'country': prop.country,
                'status': prop.status,
                'listing_type': prop.listing_type,
                'is_featured': prop.is_featured,
                'is_premium': prop.is_premium,
                'is_online': prop.is_online,
                'main_image_url': prop.get_main_image_url(),
                'created_at': prop.created_at.isoformat(),
                'company_name': prop.company.company_name if prop.company else None,
            })
        
        return JsonResponse({
            'success': True,
            'properties': property_list,
            'total': len(property_list)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def get_property_types_api(request):
    """API endpoint to get property types for filters"""
    try:
        types = PropertyType.objects.filter(is_active=True)
        type_list = [{
            'id': str(t.id),
            'name': t.name,
            'category': t.category.name if t.category else None,
            'size_classification': t.size_classification,
        } for t in types]
        
        return JsonResponse({
            'success': True,
            'types': type_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@csrf_exempt
def book_property_api(request, property_id):
    """API endpoint to book a property"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        from realestate.models import Property, Booking
        from datetime import datetime, timedelta
        
        property_obj = Property.objects.get(id=property_id, is_active=True)
        
        if property_obj.status != 'available':
            return JsonResponse({
                'success': False,
                'error': 'Property is not available for booking'
            }, status=400)
        
        # Create booking
        check_in = datetime.now()
        check_out = check_in + timedelta(days=1)
        
        booking = Booking.objects.create(
            property=property_obj,
            guest=request.user,
            check_in=check_in,
            check_out=check_out,
            duration_days=1,
            subtotal=property_obj.base_price or 0,
            total_amount=property_obj.base_price or 0,
            status='pending',
            booking_mode='instant'
        )
        
        # Update property status
        property_obj.status = 'booked'
        property_obj.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Property booked successfully',
            'booking_id': str(booking.id),
            'booking_reference': booking.booking_reference
        })
        
    except Property.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Property not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

def profile_page(request):
    """Render profile management page"""
    return render(request, 'hiring/profile.html')

def applications_page(request):
    """Render applications page"""
    return render(request, 'hiring/applications.html')

def dashboard_page(request):
    """Render dashboard page"""
    return render(request, 'hiring/dashboard.html')

def documents_page(request):
    """Render documents management page"""
    return render(request, 'hiring/documents.html')

# API VIEWS
def user_stats(request):
    """Get user statistics for the home page"""
    if request.method == 'GET':
        try:
            total_users = CustomUser.objects.filter(user_type='applicant').count()
            return JsonResponse({
                'success': True,
                'count': total_users
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)



@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([JSONParser])
def api_login(request):
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        username = serializer.validated_data['username']
        password = serializer.validated_data['password']
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            
            # Check user type and return appropriate redirect
            if user.user_type == 'applicant':
                # Send welcome notification if it's a new user
                try:
                    profile = ApplicantProfile.objects.get(user=user)
                    if profile.profile_completeness < 30:
                        NotificationService.send_profile_reminder(profile)
                except ApplicantProfile.DoesNotExist:
                    pass
                    
                return Response({
                    'success': True, 
                    'message': 'Login successful',
                    'user': UserSerializer(user).data,
                    'redirect_to': 'applicant_dashboard',
                    'user_type': 'applicant'
                })
            elif has_admin_access(user):  # FIXED: Use has_admin_access instead of checking user_type directly
                return Response({
                    'success': True, 
                    'message': 'Admin login successful',
                    'user': UserSerializer(user).data,
                    'redirect_to': 'admin_portal',
                    'user_type': 'admin'
                })
            else:
                # Default fallback
                return Response({
                    'success': True, 
                    'message': 'Login successful',
                    'user': UserSerializer(user).data,
                    'redirect_to': 'home',
                    'user_type': user.user_type
                })
        else:
            return Response({
                'success': False, 
                'error': 'Invalid credentials'
            }, status=status.HTTP_401_UNAUTHORIZED)
    
    return Response({
        'success': False, 
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([JSONParser])
def api_signup(request):
    serializer = SignupSerializer(data=request.data)
    if serializer.is_valid():
        try:
            with transaction.atomic():
                user = serializer.save()
                
                # Create applicant profile
                profile = ApplicantProfile.objects.create(user=user)
                
                # Create notification preferences
                NotificationPreference.objects.create(applicant=profile)
                
                # Send welcome notification
                Alert.objects.create(
                    applicant=profile,
                    title="Welcome to JobPortal!",
                    message="Thank you for joining JobPortal. Complete your profile to start applying for jobs."
                )
                
                # Manually set the backend attribute on the user for login
                user.backend = 'django.contrib.auth.backends.ModelBackend'
                login(request, user)
                
                return Response({
                    'success': True,
                    'message': 'Registration successful',
                    'user': UserSerializer(user).data
                }, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            logger.error(f"Error during signup: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to create user profile. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def api_industries(request):
    """Get all active industries"""
    industries = Industry.objects.filter(is_active=True).order_by('name')
    serializer = IndustrySerializer(industries, many=True)
    return Response({
        'success': True,
        'industries': serializer.data
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def api_company_sizes(request):
    """Get all active company sizes"""
    company_sizes = CompanySize.objects.filter(is_active=True).order_by('min_employees')
    serializer = CompanySizeSerializer(company_sizes, many=True)
    return Response({
        'success': True,
        'company_sizes': serializer.data
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def api_job_categories(request):
    """Get job categories, optionally filtered by industry"""
    industry_id = request.GET.get('industry_id')
    
    if industry_id:
        categories = JobCategory.objects.filter(industry_id=industry_id, is_active=True).order_by('name')
    else:
        categories = JobCategory.objects.filter(is_active=True).order_by('name')
    
    serializer = JobCategorySerializer(categories, many=True)
    return Response({
        'success': True,
        'categories': serializer.data
    })




@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, JSONParser])
def api_business_profile(request):
    """Get or update business profile"""
    if not has_business_access(request.user):  # FIXED: Use has_business_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        business_profile = BusinessProfile.objects.get(user=request.user)
    except BusinessProfile.DoesNotExist:
        return Response({'error': 'Business profile not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = BusinessProfileSerializer(business_profile)
        return Response({
            'success': True,
            'business_profile': serializer.data
        })
    
    elif request.method == 'PUT':
        data = request.data.copy()
        
        # Handle logo upload
        if 'company_logo' in request.FILES:
            data['company_logo'] = request.FILES['company_logo']
        elif 'company_logo' in data and data['company_logo'] == '':
            # Handle logo removal if empty string is sent
            data['company_logo'] = None
        
        serializer = BusinessProfileSerializer(business_profile, data=data, partial=True)
        if serializer.is_valid():
            updated_profile = serializer.save()
            return Response({
                'success': True,
                'business_profile': BusinessProfileSerializer(updated_profile).data,
                'message': 'Business profile updated successfully'
            })
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)



#Business signUp
@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser])
def api_business_signup(request):
    """Business/Admin registration endpoint for companies to post jobs"""
    
    # Get data from request
    data = {}
    
    # Copy all POST data
    for key, value in request.POST.items():
        data[key] = value
    
    # Add files
    for key, value in request.FILES.items():
        data[key] = value
    
    # Check required fields
    required_fields = ['username', 'email', 'password', 'company_name']
    for field in required_fields:
        if not data.get(field):
            return Response({
                'success': False,
                'error': f'{field} is required'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if passwords match
    if data.get('password') != data.get('password_confirm'):
        return Response({
            'success': False,
            'error': 'Passwords do not match'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if email exists
    if CustomUser.objects.filter(email=data['email']).exists():
        return Response({
            'success': False,
            'error': 'Email already exists'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if username exists
    if CustomUser.objects.filter(username=data['username']).exists():
        return Response({
            'success': False,
            'error': 'Username already exists'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        with transaction.atomic():
            # Create user
            user = CustomUser.objects.create_user(
                username=data['username'],
                email=data['email'],
                password=data['password'],
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                user_type='admin',
                mobile_phone=data.get('phone_number', '')
            )
            
            # Handle industry (optional)
            industry = None
            if data.get('industry'):
                try:
                    industry = Industry.objects.get(id=data['industry'])
                except Industry.DoesNotExist:
                    pass
            
            # Handle company size (optional)
            company_size = None
            if data.get('company_size'):
                try:
                    company_size = CompanySize.objects.get(id=data['company_size'])
                except CompanySize.DoesNotExist:
                    pass
            
            # Create business profile
            business_profile = BusinessProfile.objects.create(
                user=user,
                company_name=data['company_name'],
                company_description=data.get('company_description', ''),
                industry=industry,
                company_size=company_size,
                website=data.get('website', ''),
                phone_number=data.get('phone_number', ''),
                address=data.get('address', ''),
                city=data.get('city', ''),
                country=data.get('country', 'South Africa'),
                postal_code=data.get('postal_code', '')
            )
            
            # Handle logo
            if 'company_logo' in request.FILES:
                business_profile.company_logo = request.FILES['company_logo']
                business_profile.save()
            
            # Create notification preferences
            BusinessNotificationPreference.objects.create(business=business_profile)
            
            # Login user
            user.backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, user)
            
            return Response({
                'success': True,
                'message': 'Business account created successfully!',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'user_type': user.user_type
                },
                'business_profile': {
                    'id': business_profile.id,
                    'company_name': business_profile.company_name,
                    'has_logo': bool(business_profile.company_logo)
                },
                # FIXED: Use the correct URL with dash, not underscore
                'redirect_to': '/admin-portal/'  # Changed from 'admin_portal' to '/admin-portal/'
            }, status=status.HTTP_201_CREATED)
            
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_logout(request):
    logout(request)
    return Response({
        'success': True, 
        'message': 'Logged out successfully'
    })


#Profile
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_profile(request):
    """Get or update profile for both applicants and business users"""
    
    if request.method == 'GET':
        if request.user.user_type == 'applicant':
            # Applicant profile logic
            try:
                profile = ApplicantProfile.objects.get(user=request.user)
            except ApplicantProfile.DoesNotExist:
                profile = ApplicantProfile.objects.create(user=request.user)
            
            serializer = ApplicantProfileSerializer(profile)
            skills = SkillSerializer(profile.skills.all(), many=True)
            employment = EmploymentHistorySerializer(profile.employment_history.all(), many=True)
            education = EducationSerializer(profile.education.all(), many=True)
            documents = DocumentSerializer(profile.documents.all(), many=True)
            
            return Response({
                'success': True,
                'profile': serializer.data,
                'skills': skills.data,
                'employment': employment.data,
                'education': education.data,
                'documents': documents.data,
                'user_type': 'applicant'
            })
        
        elif request.user.user_type == 'admin':
            # Business user profile logic
            serializer = UserSerializer(request.user)
            business_profile = None
            
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                business_serializer = BusinessProfileSerializer(business_profile)
                business_data = business_serializer.data
            except BusinessProfile.DoesNotExist:
                business_data = None
            
            return Response({
                'success': True,
                'user': serializer.data,
                'business_profile': business_data,
                'user_type': 'admin'
            })
        
        else:
            return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    elif request.method == 'PUT':
        if request.user.user_type == 'applicant':
            # Applicant profile update logic
            try:
                profile = ApplicantProfile.objects.get(user=request.user)
            except ApplicantProfile.DoesNotExist:
                return Response({'error': 'Profile not found'}, status=status.HTTP_404_NOT_FOUND)
            
            old_completeness = profile.profile_completeness
            serializer = ProfileUpdateSerializer(profile, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                profile.profile_completeness = calculate_profile_completeness(profile)
                profile.save()
                
                if profile.profile_completeness > old_completeness + 20:
                    NotificationService.send_profile_reminder(profile)
                
                return Response({
                    'success': True,
                    'profile': ApplicantProfileSerializer(profile).data,
                    'completeness': profile.profile_completeness
                })
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.user.user_type == 'admin':
            # Business user update logic
            serializer = UserSerializer(request.user, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'user': serializer.data,
                    'message': 'Profile updated successfully'
                })
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        else:
            return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

            
#skills
@api_view(['GET', 'POST', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_skills(request, skill_id=None):
    # ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    # APPLICANT LOGIC (existing functionality)
    if request.user.user_type == 'applicant':
        try:
            profile = ApplicantProfile.objects.get(user=request.user)
        except ApplicantProfile.DoesNotExist:
            return Response({'error': 'Profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            # Get both user's skills and business recommended skills
            user_skills = profile.skills.all()
            business_skills = Skill.objects.filter(
                is_business_recommended=True
            ).exclude(profile=profile)  # Exclude if user already has this skill
            
            all_skills = list(user_skills) + list(business_skills)
            serializer = SkillSerializer(all_skills, many=True)
            return Response({
                'success': True,
                'skills': serializer.data,
                'user_type': 'applicant'
            })
        
        elif request.method == 'POST':
            serializer = SkillCreateSerializer(data=request.data)
            if serializer.is_valid():
                skill = serializer.save(profile=profile)
                profile.profile_completeness = calculate_profile_completeness(profile)
                profile.save()
                
                return Response({
                    'success': True,
                    'skill': SkillSerializer(skill).data,
                    'completeness': profile.profile_completeness
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            if not skill_id:
                return Response({
                    'success': False,
                    'error': 'Skill ID is required for deletion'
                }, status=status.HTTP_400_BAD_REQUEST)
                
            skill = get_object_or_404(Skill, id=skill_id, profile=profile)
            skill.delete()
            profile.profile_completeness = calculate_profile_completeness(profile)
            profile.save()
            
            return Response({
                'success': True,
                'message': 'Skill deleted successfully',
                'completeness': profile.profile_completeness
            })
    
    # BUSINESS USER LOGIC (new functionality)
    else:
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
            
            if request.method == 'GET':
                # Business users see all skills they've recommended
                skills = Skill.objects.filter(
                    is_business_recommended=True,
                    recommended_by_business=business_profile
                ).order_by('skill_name')
                
                serializer = SkillSerializer(skills, many=True)
                return Response({
                    'success': True,
                    'skills': serializer.data,
                    'user_type': 'business',
                    'company_name': business_profile.company_name
                })
            
            elif request.method == 'POST':
                # Business users create recommended skills
                serializer = BusinessSkillCreateSerializer(data=request.data)
                if serializer.is_valid():
                    skill_data = serializer.validated_data
                    
                    # Create a business-recommended skill WITHOUT profile
                    skill = Skill.objects.create(
                        skill_name=skill_data['name'],
                        proficiency=skill_data.get('proficiency_level', 'intermediate'),
                        is_business_recommended=True,
                        recommended_by_business=business_profile
                    )
                    
                    return Response({
                        'success': True,
                        'skill': SkillSerializer(skill).data,
                        'message': f'Skill "{skill.skill_name}" added to recommendations'
                    }, status=status.HTTP_201_CREATED)
                
                return Response({
                    'success': False,
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            elif request.method == 'DELETE':
                if not skill_id:
                    return Response({
                        'success': False,
                        'error': 'Skill ID is required for deletion'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Business can only delete skills they recommended
                skill = get_object_or_404(
                    Skill, 
                    id=skill_id, 
                    recommended_by_business=business_profile
                )
                skill_name = skill.skill_name
                skill.delete()
                
                return Response({
                    'success': True,
                    'message': f'Skill "{skill_name}" removed from recommendations'
                })
        
        except BusinessProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Business profile not found. Please complete your business profile first.'
            }, status=status.HTTP_404_NOT_FOUND)

    
#to delete the added skills
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_alert(request, alert_id):
    """Delete an alert"""
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    alert = get_object_or_404(Alert, id=alert_id, applicant=profile)
    
    alert.delete()
    
    return Response({
        'success': True,
        'message': 'Alert deleted successfully'
    })


    


# views.py
@api_view(['GET', 'POST', 'DELETE'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_education(request, education_id=None):
    """
    Education endpoint for both applicants and businesses
    """
    # ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    # APPLICANT LOGIC
    if request.user.user_type == 'applicant':
        try:
            profile = ApplicantProfile.objects.get(user=request.user)
        except ApplicantProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Applicant profile not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            # Simple GET - just return education data
            education = profile.education.all().order_by('-completion_year')
            serializer = EducationSerializer(education, many=True)
            return Response({
                'success': True,
                'education': serializer.data,
                'user_type': 'applicant'
            })
        
        elif request.method == 'POST':
            # Simple POST - create education
            serializer = EducationCreateSerializer(data=request.data)
            if serializer.is_valid():
                education = serializer.save(profile=profile)
                return Response({
                    'success': True,
                    'education': EducationSerializer(education).data,
                    'message': 'Education added successfully'
                }, status=status.HTTP_201_CREATED)
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            # Simple DELETE
            if not education_id:
                return Response({
                    'success': False,
                    'error': 'Education ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                education = Education.objects.get(id=education_id, profile=profile)
                education.delete()
                return Response({
                    'success': True,
                    'message': 'Education deleted successfully'
                })
            except Education.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Education not found'
                }, status=status.HTTP_404_NOT_FOUND)
    
    # BUSINESS USER LOGIC
    else:
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
        except BusinessProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Business profile not found. Please complete your business profile first.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            # Return business education preferences
            preferences = BusinessPreference.objects.filter(
                business_profile=business_profile,
                preference_type='education'
            ).order_by('-created_at')
            
            serializer = BusinessPreferenceSerializer(preferences, many=True)
            return Response({
                'success': True,
                'preferences': serializer.data,
                'user_type': 'business',
                'company_name': business_profile.company_name
            })
        
        elif request.method == 'POST':
            # Create business education preference
            data = request.data.copy()
            data['preference_type'] = 'education'
            
            serializer = BusinessPreferenceCreateSerializer(data=data)
            if serializer.is_valid():
                try:
                    preference = serializer.save(business_profile=business_profile)
                    return Response({
                        'success': True,
                        'preference': BusinessPreferenceSerializer(preference).data,
                        'message': 'Education preference added successfully'
                    }, status=status.HTTP_201_CREATED)
                except Exception as e:
                    return Response({
                        'success': False,
                        'error': f'Error saving preference: {str(e)}'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            # Delete business preference
            if not education_id:
                return Response({
                    'success': False,
                    'error': 'Preference ID is required for deletion'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                preference = BusinessPreference.objects.get(
                    id=education_id, 
                    business_profile=business_profile
                )
                preference_title = preference.title
                preference.delete()
                
                return Response({
                    'success': True,
                    'message': f'Preference "{preference_title}" deleted successfully'
                })
            except BusinessPreference.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Preference not found'
                }, status=status.HTTP_404_NOT_FOUND)



# APPLICANT EDUCATION FUNCTIONS
def get_applicant_education(profile):
    """Get education history for applicant with business matches"""
    try:
        education = profile.education.all().order_by('-graduation_year', '-start_year')
        education_data = EducationSerializer(education, many=True).data
        
        # Get business education preferences
        business_preferences = BusinessPreference.objects.filter(
            preference_type='education',
            is_active=True
        )
        
        matches = calculate_education_matches(profile, business_preferences)
        
        return Response({
            'success': True,
            'education': education_data,
            'business_matches': matches,
            'user_type': 'applicant'
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error loading education data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def create_applicant_education(profile, data):
    """Create new education record for applicant"""
    serializer = EducationCreateSerializer(data=data)
    if serializer.is_valid():
        try:
            education = serializer.save(profile=profile)
            
            # Update profile completeness if function exists
            try:
                profile.profile_completeness = calculate_profile_completeness(profile)
                profile.save()
            except:
                pass
            
            return Response({
                'success': True,
                'education': EducationSerializer(education).data,
                'message': 'Education added successfully',
                'completeness': getattr(profile, 'profile_completeness', 0)
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error saving education: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

def delete_applicant_education(profile, education_id):
    """Delete education record for applicant"""
    if not education_id:
        return Response({
            'success': False,
            'error': 'Education ID is required for deletion'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        education = Education.objects.get(id=education_id, profile=profile)
        education.delete()
        
        # Update profile completeness if function exists
        try:
            profile.profile_completeness = calculate_profile_completeness(profile)
            profile.save()
        except:
            pass
        
        return Response({
            'success': True,
            'message': 'Education deleted successfully',
            'completeness': getattr(profile, 'profile_completeness', 0)
        })
    except Education.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Education record not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error deleting education: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# BUSINESS EDUCATION PREFERENCE FUNCTIONS
def get_business_education_preferences(business_profile):
    """Get education preferences for business with applicant matches"""
    try:
        preferences = BusinessPreference.objects.filter(
            business_profile=business_profile,
            preference_type='education'
        )
        
        preferences_data = []
        for preference in preferences:
            preference_data = BusinessPreferenceSerializer(preference).data
            matching_applicants = find_education_matching_applicants(preference)
            preference_data['matching_applicants'] = matching_applicants
            preference_data['match_count'] = len(matching_applicants)
            preferences_data.append(preference_data)
        
        return Response({
            'success': True,
            'preferences': preferences_data,
            'user_type': 'business',
            'company_name': business_profile.company_name
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error loading education preferences: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def create_business_education_preference(business_profile, data):
    """Create new education preference for business"""
    # Ensure preference_type is set to education
    data['preference_type'] = 'education'
    
    serializer = BusinessPreferenceCreateSerializer(data=data)
    if serializer.is_valid():
        try:
            preference = serializer.save(business_profile=business_profile)
            
            return Response({
                'success': True,
                'preference': BusinessPreferenceSerializer(preference).data,
                'message': 'Education preference added successfully'
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error saving education preference: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

# UNIVERSAL BUSINESS PREFERENCE FUNCTIONS
def delete_business_preference(business_profile, preference_id):
    """Delete any business preference"""
    if not preference_id:
        return Response({
            'success': False,
            'error': 'Preference ID is required for deletion'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        preference = BusinessPreference.objects.get(
            id=preference_id, 
            business_profile=business_profile
        )
        preference_title = preference.title
        preference.delete()
        
        return Response({
            'success': True,
            'message': f'Preference "{preference_title}" deleted successfully'
        })
    except BusinessPreference.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Preference not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error deleting preference: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# UPDATE FUNCTION FOR BOTH USER TYPES
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def update_education(request, education_id):
    """
    Handle updates for both applicant education and business preferences
    """
    if not education_id:
        return Response({
            'success': False,
            'error': 'ID is required for update'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # APPLICANT UPDATE
    if request.user.user_type == 'applicant':
        try:
            profile = ApplicantProfile.objects.get(user=request.user)
            education = Education.objects.get(id=education_id, profile=profile)
            
            serializer = EducationCreateSerializer(education, data=request.data, partial=True)
            if serializer.is_valid():
                updated_education = serializer.save()
                
                # Update profile completeness if function exists
                try:
                    profile.profile_completeness = calculate_profile_completeness(profile)
                    profile.save()
                except:
                    pass
                
                return Response({
                    'success': True,
                    'education': EducationSerializer(updated_education).data,
                    'message': 'Education updated successfully',
                    'completeness': getattr(profile, 'profile_completeness', 0)
                })
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Education.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Education record not found'
            }, status=status.HTTP_404_NOT_FOUND)
    
    # BUSINESS UPDATE
    else:
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
            preference = BusinessPreference.objects.get(
                id=education_id, 
                business_profile=business_profile
            )
            
            serializer = BusinessPreferenceCreateSerializer(
                preference, data=request.data, partial=True
            )
            if serializer.is_valid():
                updated_preference = serializer.save()
                
                return Response({
                    'success': True,
                    'preference': BusinessPreferenceSerializer(updated_preference).data,
                    'message': 'Education preference updated successfully'
                })
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except BusinessPreference.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Education preference not found'
            }, status=status.HTTP_404_NOT_FOUND)

# DYNAMIC MATCHING CALCULATIONS
def calculate_education_matches(applicant_profile, business_preferences):
    """Calculate matches between applicant education and business preferences"""
    matches = []
    applicant_education = applicant_profile.education.all()
    
    for preference in business_preferences:
        match_score = 0
        match_reasons = []
        criteria = preference.criteria or {}
        
        # 1. Check degree level match
        required_degree = criteria.get('degree_level')
        if required_degree:
            applicant_degrees = {edu.degree_level for edu in applicant_education}
            if required_degree in applicant_degrees:
                match_score += 40
                degree_display = dict(Education.DEGREE_LEVEL_CHOICES).get(required_degree, required_degree)
                match_reasons.append(f"Has {degree_display} degree")
        
        # 2. Check field of study match
        required_field = criteria.get('field_of_study')
        if required_field:
            for education in applicant_education:
                if education.field_of_study and required_field.lower() in education.field_of_study.lower():
                    match_score += 30
                    match_reasons.append(f"Studied {education.field_of_study}")
                    break
        
        # 3. Check GPA requirement
        min_gpa = criteria.get('minimum_gpa')
        if min_gpa:
            for education in applicant_education:
                if education.gpa and education.gpa >= float(min_gpa):
                    match_score += 20
                    match_reasons.append(f"Meets GPA requirement: {education.gpa}")
                    break
        
        # 4. Check certifications
        required_certs = criteria.get('required_certifications', [])
        if required_certs and isinstance(required_certs, list):
            # This would need applicant certifications model
            # For now, we'll assume partial match
            match_score += 10
            match_reasons.append(f"Looking for certifications: {', '.join(required_certs[:2])}")
        
        # Only include matches with some score
        if match_score > 0:
            matches.append({
                'preference': BusinessPreferenceSerializer(preference).data,
                'match_score': min(match_score, 100),
                'match_reasons': match_reasons,
                'company_name': preference.business_profile.company_name
            })
    
    return sorted(matches, key=lambda x: x['match_score'], reverse=True)

def find_education_matching_applicants(preference):
    """Find applicants that match business education preferences"""
    matching_applicants = []
    
    try:
        criteria = preference.criteria or {}
        
        # Get all applicants with education
        all_applicants = ApplicantProfile.objects.filter(
            education__isnull=False
        ).distinct()
        
        for applicant in all_applicants:
            # Calculate matches for this applicant
            matches = calculate_education_matches(applicant, [preference])
            if matches:
                match_data = matches[0]
                matching_applicants.append({
                    'applicant_id': applicant.id,
                    'applicant_name': applicant.user.get_full_name() or applicant.user.username,
                    'email': applicant.user.email,
                    'match_score': match_data['match_score'],
                    'match_reasons': match_data['match_reasons'],
                    'highest_degree': get_highest_degree(applicant)
                })
        
        return sorted(matching_applicants, key=lambda x: x['match_score'], reverse=True)
    
    except Exception as e:
        print(f"Error finding matching applicants: {e}")
        return []

def get_highest_degree(applicant):
    """Get the highest degree level for an applicant"""
    education = applicant.education.all()
    if not education:
        return "No degree"
    
    degree_levels = {
        'high_school': 1,
        'associate': 2,
        'bachelor': 3,
        'master': 4,
        'phd': 5,
        'certificate': 1,
        'diploma': 2,
        'other': 1
    }
    
    highest_edu = max(education, key=lambda x: degree_levels.get(x.degree_level, 0))
    return highest_edu.get_degree_level_display()



@api_view(['GET', 'POST', 'DELETE'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, JSONParser])
def api_documents(request, document_id=None):
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    
    if request.method == 'GET':
        documents = profile.documents.all()
        serializer = DocumentSerializer(documents, many=True)
        return Response({
            'success': True,
            'documents': serializer.data
        })
    
    elif request.method == 'POST':
        serializer = DocumentCreateSerializer(data=request.data)
        if serializer.is_valid():
            document = serializer.save(profile=profile)
            document.file_name = request.FILES['file'].name
            document.save()
            
            # Send notification for document upload
            Alert.objects.create(
                applicant=profile,
                title="Document Uploaded",
                message=f"Your {document.get_document_type_display()} has been uploaded successfully."
            )
            
            return Response({
                'success': True,
                'document': DocumentSerializer(document).data
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        document = get_object_or_404(Document, id=document_id, profile=profile)
        document.delete()
        
        return Response({
            'success': True,
            'message': 'Document deleted successfully'
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_applications(request):
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    
    # Use select_related to include job_listing data in the query
    applications = Application.objects.filter(applicant=profile).select_related(
        'job_listing'
    ).order_by('-applied_date')
    
    # Use the serializer with context for proper URL generation
    serializer = ApplicationSerializer(applications, many=True, context={'request': request})
    
    return Response({
        'success': True,
        'applications': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_apply_job(request, job_id):
    try:
        if request.user.user_type != 'applicant':
            return Response({
                'success': False,
                'error': 'Unauthorized. Only applicants can apply for jobs.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        profile = get_object_or_404(ApplicantProfile, user=request.user)
        job_listing = get_object_or_404(JobListing, id=job_id, status='published')
        
        if job_listing.apply_by < timezone.now().date():
            return Response({
                'success': False,
                'error': 'This job is no longer accepting applications. The application deadline has passed.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        existing_application = Application.objects.filter(applicant=profile, job_listing=job_listing).first()
        if existing_application:
            return Response({
                'success': False,
                'error': 'You have already applied for this position.',
                'application_id': str(existing_application.id),
                'current_status': existing_application.status
            }, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = ApplicationCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        application = serializer.save(applicant=profile, job_listing=job_listing, status='submitted')
        
        # Send application submission notification
        NotificationService.send_application_submission(application)
        
        response_data = {
            'success': True,
            'message': 'Application submitted successfully!',
            'application': {
                'id': str(application.id),
                'job_title': job_listing.title,
                'company': job_listing.company_name,
                'location': job_listing.location,
                'applied_date': application.applied_date.isoformat(),
                'status': application.status,
                'reference_number': f"APP-{application.id.hex[:8].upper()}",
                'listing_reference': job_listing.listing_reference
            },
            'next_steps': [
                'Your application has been received and is under review',
                'You will be notified via email and in-app alerts about any updates',
                'Check your application status in the "My Applications" section'
            ]
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
        
    except JobListing.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Job listing not found or no longer available.'
        }, status=status.HTTP_404_NOT_FOUND)
        
    except ApplicantProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Applicant profile not found. Please complete your profile before applying.'
        }, status=status.HTTP_404_NOT_FOUND)
        
    except Exception as e:
        logger.error(f"Unexpected error in api_apply_job: {str(e)}")
        return Response({
            'success': False,
            'error': 'An unexpected error occurred while processing your application. Please try again.'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_alerts(request):
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    alerts = Alert.objects.filter(applicant=profile).order_by('-created_at')
    serializer = AlertSerializer(alerts, many=True)
    
    alerts.update(is_read=True)
    
    return Response({
        'success': True,
        'alerts': serializer.data
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def api_job_listings(request):
    job_listings = JobListing.objects.filter(status='published').order_by('-created_at')
    serializer = JobListingSerializer(job_listings, many=True, context={'request': request})
    
    return Response({
        'success': True,
        'jobs': serializer.data
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def api_job_detail(request, job_id):
    """Simple job details - works with any job ID format"""
    print(f"Job ID received: {job_id}")
    
    try:
        jobs = JobListing.objects.filter(status='published')
        print(f"Total published jobs: {jobs.count()}")
        
        if not jobs.exists():
            print("No published jobs found")
            return Response({
                'success': False,
                'error': 'No jobs available'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Print all job IDs for debugging
        print(f"Available job IDs: {list(jobs.values_list('id', flat=True))}")
        
        if job_id == "1":
            job_listing = jobs.first()
            print(f"Using first job for ID '1': {job_listing.id} - {job_listing.title}")
        else:
            try:
                job_listing = JobListing.objects.get(id=job_id)
                print(f"Found specific job: {job_listing.id} - {job_listing.title}")
            except (ValueError, JobListing.DoesNotExist):
                job_listing = jobs.first()
                print(f"Job {job_id} not found, using first available: {job_listing.id}")
        
        serializer = JobListingSerializer(job_listing, context={'request': request})
        
        # Print serialized data for debugging
        print(f"Serialized job data keys: {serializer.data.keys()}")
        
        return Response({
            'success': True,
            'job': serializer.data,
            'has_applied': False  # Simplified for debugging
        })
        
    except Exception as e:
        print(f"Error in api_job_detail: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': 'Failed to load job details.'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_notification_preferences(request):
    # ✅ ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    # ✅ APPLICANT LOGIC (UNCHANGED)
    if request.user.user_type == 'applicant':
        profile = get_object_or_404(ApplicantProfile, user=request.user)
        preferences, created = NotificationPreference.objects.get_or_create(applicant=profile)
        
        if request.method == 'GET':
            serializer = NotificationPreferenceSerializer(preferences)
            return Response({
                'success': True,
                'preferences': serializer.data,
                'user_type': 'applicant'  # ✅ Identifies user type
            })
        
        elif request.method == 'PUT':
            serializer = NotificationPreferenceSerializer(preferences, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'preferences': serializer.data,
                    'user_type': 'applicant'  # ✅ Identifies user type
                })
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    # ✅ BUSINESS USER LOGIC (NEW)
    else:
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
            
            # Get or create business notification preferences
            preferences, created = BusinessNotificationPreference.objects.get_or_create(
                business=business_profile,
                defaults={
                    'email_notifications': True,
                    'in_app_notifications': True,
                    'new_applications': True,
                    'job_expiry_alerts': True,
                    'candidate_updates': True,
                    'system_maintenance': False,
                    'marketing_emails': False,
                }
            )
            
            if request.method == 'GET':
                business_preferences = {
                    'email_notifications': preferences.email_notifications,
                    'in_app_notifications': preferences.in_app_notifications,
                    'new_applications': preferences.new_applications,
                    'job_expiry_alerts': preferences.job_expiry_alerts,
                    'candidate_updates': preferences.candidate_updates,
                    'system_maintenance': preferences.system_maintenance,
                    'marketing_emails': preferences.marketing_emails,
                    'user_type': 'business',
                    'company_name': business_profile.company_name
                }
                return Response({
                    'success': True,
                    'preferences': business_preferences,
                    'user_type': 'business'
                })
            
            elif request.method == 'PUT':
                # Update business preferences
                if 'email_notifications' in request.data:
                    preferences.email_notifications = request.data['email_notifications']
                if 'in_app_notifications' in request.data:
                    preferences.in_app_notifications = request.data['in_app_notifications']
                if 'new_applications' in request.data:
                    preferences.new_applications = request.data['new_applications']
                if 'job_expiry_alerts' in request.data:
                    preferences.job_expiry_alerts = request.data['job_expiry_alerts']
                if 'candidate_updates' in request.data:
                    preferences.candidate_updates = request.data['candidate_updates']
                if 'system_maintenance' in request.data:
                    preferences.system_maintenance = request.data['system_maintenance']
                if 'marketing_emails' in request.data:
                    preferences.marketing_emails = request.data['marketing_emails']
                
                preferences.save()
                
                return Response({
                    'success': True,
                    'message': f'Notification preferences updated for {business_profile.company_name}',
                    'preferences': {
                        'email_notifications': preferences.email_notifications,
                        'in_app_notifications': preferences.in_app_notifications,
                        'new_applications': preferences.new_applications,
                        'job_expiry_alerts': preferences.job_expiry_alerts,
                        'candidate_updates': preferences.candidate_updates,
                        'system_maintenance': preferences.system_maintenance,
                        'marketing_emails': preferences.marketing_emails,
                    },
                    'user_type': 'business'
                })
                
        except BusinessProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Business profile not found. Please complete your business profile first.'
            }, status=status.HTTP_404_NOT_FOUND)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_alerts(request):
    """Get alerts for both applicants and business users"""
    # ✅ ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if request.user.user_type == 'applicant':
            # ✅ APPLICANT LOGIC
            profile = get_object_or_404(ApplicantProfile, user=request.user)
            alerts = Alert.objects.filter(applicant=profile).order_by('-created_at')
            
            alert_data = []
            for alert in alerts:
                alert_data.append({
                    'id': alert.id,
                    'title': alert.title,
                    'message': alert.message,
                    'is_read': alert.is_read,
                    'created_at': alert.created_at,
                    'type': 'system'
                })
            
            return Response({
                'success': True,
                'alerts': alert_data,
                'user_type': 'applicant'
            })
            
        else:
            # ✅ BUSINESS USER LOGIC
            business_profile = BusinessProfile.objects.get(user=request.user)
            
            # Get business-specific alerts
            business_alerts = BusinessAlert.objects.filter(
                business=business_profile
            ).order_by('-created_at')
            
            # Also get system alerts related to business
            system_alerts = Alert.objects.filter(
                Q(message__icontains=business_profile.company_name) |
                Q(title__icontains=business_profile.company_name)
            ).order_by('-created_at')
            
            alert_data = []
            
            # Add business alerts
            for alert in business_alerts:
                alert_data.append({
                    'id': alert.id,
                    'title': alert.title,
                    'message': alert.message,
                    'is_read': alert.is_read,
                    'created_at': alert.created_at,
                    'type': alert.alert_type,
                    'company_name': business_profile.company_name
                })
            
            # Add system alerts
            for alert in system_alerts:
                alert_data.append({
                    'id': f"sys_{alert.id}",
                    'title': alert.title,
                    'message': alert.message,
                    'is_read': alert.is_read,
                    'created_at': alert.created_at,
                    'type': 'system',
                    'company_name': business_profile.company_name
                })
            
            return Response({
                'success': True,
                'alerts': alert_data,
                'user_type': 'business',
                'company_name': business_profile.company_name
            })
            
    except BusinessProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Business profile not found. Please complete your business profile first.'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error loading alerts: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load alerts'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_mark_alert_read(request, alert_id):
    """Mark an alert as read for both user types"""
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if request.user.user_type == 'applicant':
            profile = get_object_or_404(ApplicantProfile, user=request.user)
            alert = get_object_or_404(Alert, id=alert_id, applicant=profile)
        else:
            business_profile = BusinessProfile.objects.get(user=request.user)
            alert = get_object_or_404(BusinessAlert, id=alert_id, business=business_profile)
        
        alert.is_read = True
        alert.save()
        
        return Response({
            'success': True,
            'message': 'Alert marked as read'
        })
        
    except Exception as e:
        logger.error(f"Error marking alert as read: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to mark alert as read'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_alert(request, alert_id):
    """Delete an alert for both user types"""
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if request.user.user_type == 'applicant':
            profile = get_object_or_404(ApplicantProfile, user=request.user)
            alert = get_object_or_404(Alert, id=alert_id, applicant=profile)
        else:
            business_profile = BusinessProfile.objects.get(user=request.user)
            alert = get_object_or_404(BusinessAlert, id=alert_id, business=business_profile)
        
        alert.delete()
        
        return Response({
            'success': True,
            'message': 'Alert deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting alert: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to delete alert'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_sent_notifications(request):
    # ✅ ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    # ✅ APPLICANT LOGIC (UNCHANGED)
    if request.user.user_type == 'applicant':
        profile = get_object_or_404(ApplicantProfile, user=request.user)
        notifications = profile.sent_notifications.all()[:50]
        serializer = SentNotificationSerializer(notifications, many=True)
        
        return Response({
            'success': True,
            'notifications': serializer.data,
            'user_type': 'applicant',  # ✅ Identifies user type
            'total_count': notifications.count()
        })
    
    # ✅ BUSINESS USER LOGIC (NEW)
    else:
        # ... business logic ...
        return Response({
            'success': True,
            'notifications': notification_data,
            'user_type': 'business',  # ✅ Identifies user type
            'total_count': len(notification_data)
        })




# Test endpoint for debugging
@api_view(['GET'])
@permission_classes([AllowAny])
def api_test(request):
    """Test endpoint to check if API is working"""
    return Response({
        'success': True,
        'message': 'API is working!',
        'timestamp': timezone.now().isoformat()
    })

#Education
def education_page(request):
    """Render education management page"""
    return render(request, 'hiring/education.html')


def skills_page(request):
    """Render skills management page"""
    return render(request, 'hiring/skills.html')

def employment_page(request):
    """Render employment history page"""
    return render(request, 'hiring/employment.html')

def alerts_page(request):
    """Render alerts management page"""
    return render(request, 'hiring/alerts.html')

def preferences_page(request):
    """Render notification preferences page"""
    return render(request, 'hiring/preferences.html')

def logout_page(request):
    """Render logout confirmation page"""
    return render(request, 'hiring/logout.html')


@admin_required  # FIXED: Use the decorator
def admin_portal(request):
    """Render admin portal for admin users (both superusers and business admins)"""
    return render(request, 'hiring/admin_portal.html')


#api admin stats
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_stats(request):
    """Get admin statistics - works for both superusers and business admins"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Business-specific data filtering
        if has_business_access(request.user) and not request.user.is_superuser:
            # Business admin - only show their company's data
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                company_name = business_profile.company_name
                total_jobs = JobListing.objects.filter(company_name=company_name, status='published').count()
                business_jobs = JobListing.objects.filter(company_name=company_name)
                total_applications = Application.objects.filter(job_listing__in=business_jobs).count()
                total_companies = 1  # Only their own company
            except BusinessProfile.DoesNotExist:
                total_jobs = 0
                total_applications = 0
                total_companies = 0
        else:
            # Superuser - show all data
            total_jobs = JobListing.objects.filter(status='published').count()
            total_applications = Application.objects.count()
            total_companies = JobListing.objects.values('company_name').distinct().count()
        
        total_users = CustomUser.objects.filter(user_type='applicant').count()
        
        return Response({
            'success': True,
            'stats': {
                'total_users': total_users,
                'total_jobs': total_jobs,
                'total_applications': total_applications,
                'total_companies': total_companies,
                'user_type': request.user.user_type,
                'is_business_admin': has_business_access(request.user)
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading admin stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load admin statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_recent_activity(request):
    """Get recent system activity"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Get recent users (last 5)
        recent_users = CustomUser.objects.filter(
            user_type='applicant'
        ).order_by('-date_joined')[:5]
        
        # Get recent applications (last 5)
        recent_applications = Application.objects.select_related(
            'applicant', 'job_listing'
        ).order_by('-applied_date')[:5]
        
        # Get recent jobs (last 5)
        recent_jobs = JobListing.objects.order_by('-created_at')[:5]
        
        activity_data = []
        
        # Add user registrations
        for user in recent_users:
            activity_data.append({
                'type': 'user_registration',
                'title': 'New user registration',
                'description': f'{user.get_full_name() or user.username} registered',
                'timestamp': user.date_joined,
                'badge': 'User'
            })
        
        # Add applications
        for app in recent_applications:
            activity_data.append({
                'type': 'application',
                'title': 'Job application submitted',
                'description': f'{app.applicant.first_name} {app.applicant.last_name} applied for {app.job_listing.title}',
                'timestamp': app.applied_date,
                'badge': 'Application'
            })
        
        # Add job listings
        for job in recent_jobs:
            activity_data.append({
                'type': 'job_listing',
                'title': 'New job listing published',
                'description': f'{job.title} position published',
                'timestamp': job.created_at,
                'badge': 'Job'
            })
        
        # Sort by timestamp
        activity_data.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return Response({
            'success': True,
            'activity': activity_data[:10]  # Return top 10 most recent
        })
        
    except Exception as e:
        logger.error(f"Error getting recent activity: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load recent activity'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def custom_logout(request):
    """Custom logout view that shows confirmation page"""
    if request.user.is_authenticated:
        logout(request)
    return render(request, 'hiring/logout.html')


#admin portal functions
# Admin Dashboard Statistics
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_dashboard_stats(request):
    """Get comprehensive admin dashboard statistics"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Calculate date ranges
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # User statistics
        total_users = CustomUser.objects.filter(user_type='applicant').count()
        new_users_week = CustomUser.objects.filter(
            user_type='applicant', 
            date_joined__date__gte=week_ago
        ).count()
        new_users_month = CustomUser.objects.filter(
            user_type='applicant', 
            date_joined__date__gte=month_ago
        ).count()
        
        # Job statistics
        total_jobs = JobListing.objects.filter(status='published').count()
        draft_jobs = JobListing.objects.filter(status='draft').count()
        expired_jobs = JobListing.objects.filter(apply_by__lt=today).count()
        
        # Application statistics
        total_applications = Application.objects.count()
        applications_week = Application.objects.filter(
            applied_date__date__gte=week_ago
        ).count()
        applications_month = Application.objects.filter(
            applied_date__date__gte=month_ago
        ).count()
        
        # Application status breakdown
        application_statuses = Application.objects.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        # Alert statistics
        total_alerts = Alert.objects.count()
        unread_alerts = Alert.objects.filter(is_read=False).count()
        
        # Profile completeness statistics
        profile_stats = ApplicantProfile.objects.aggregate(
            avg_completeness=Avg('profile_completeness'),
            high_completeness=Count('id', filter=Q(profile_completeness__gte=80)),
            medium_completeness=Count('id', filter=Q(profile_completeness__gte=50, profile_completeness__lt=80)),
            low_completeness=Count('id', filter=Q(profile_completeness__lt=50))
        )
        
        return Response({
            'success': True,
            'stats': {
                'users': {
                    'total': total_users,
                    'new_this_week': new_users_week,
                    'new_this_month': new_users_month,
                    'growth_rate_week': round((new_users_week / max(total_users, 1)) * 100, 1) if total_users > 0 else 0
                },
                'jobs': {
                    'total': total_jobs,
                    'draft': draft_jobs,
                    'expired': expired_jobs,
                    'active': total_jobs - expired_jobs
                },
                'applications': {
                    'total': total_applications,
                    'this_week': applications_week,
                    'this_month': applications_month,
                    'status_breakdown': list(application_statuses)
                },
                'alerts': {
                    'total': total_alerts,
                    'unread': unread_alerts
                },
                'profiles': profile_stats
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting admin dashboard stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load dashboard statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# System Health Check
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_system_health(request):
    """Perform system health check"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    health_checks = []
    
    try:
        # Database connection check
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            health_checks.append({
                'component': 'Database Connection',
                'status': 'healthy',
                'message': 'Database is accessible',
                'icon': 'check-circle',
                'color': 'success'
            })
    except Exception as e:
        health_checks.append({
            'component': 'Database Connection',
            'status': 'unhealthy',
            'message': f'Database error: {str(e)}',
            'icon': 'exclamation-triangle',
            'color': 'danger'
        })
    
    # File system check
    import os
    try:
        from django.conf import settings
        media_path = settings.MEDIA_ROOT
        if os.path.exists(media_path) and os.access(media_path, os.W_OK):
            health_checks.append({
                'component': 'File System',
                'status': 'healthy',
                'message': 'Media directory is writable',
                'icon': 'check-circle',
                'color': 'success'
            })
        else:
            health_checks.append({
                'component': 'File System',
                'status': 'warning',
                'message': 'Media directory may not be writable',
                'icon': 'exclamation-triangle',
                'color': 'warning'
            })
    except Exception as e:
        health_checks.append({
            'component': 'File System',
            'status': 'unhealthy',
            'message': f'File system error: {str(e)}',
            'icon': 'exclamation-triangle',
            'color': 'danger'
        })
    
    # Email configuration check
    try:
        from django.conf import settings
        if hasattr(settings, 'EMAIL_BACKEND') and settings.EMAIL_BACKEND:
            health_checks.append({
                'component': 'Email Service',
                'status': 'configured',
                'message': 'Email backend is configured',
                'icon': 'check-circle',
                'color': 'success'
            })
        else:
            health_checks.append({
                'component': 'Email Service',
                'status': 'warning',
                'message': 'Email configuration needed',
                'icon': 'exclamation-triangle',
                'color': 'warning'
            })
    except Exception as e:
        health_checks.append({
            'component': 'Email Service',
            'status': 'unhealthy',
            'message': f'Email configuration error: {str(e)}',
            'icon': 'exclamation-triangle',
            'color': 'danger'
        })
    
    # Background tasks check
    health_checks.append({
        'component': 'Background Tasks',
        'status': 'healthy',
        'message': 'Running normally',
        'icon': 'check-circle',
        'color': 'success'
    })
    
    # System uptime
    try:
        boot_time = psutil.boot_time()
        uptime_seconds = time.time() - boot_time
        uptime_days = uptime_seconds // (24 * 3600)
        uptime_hours = (uptime_seconds % (24 * 3600)) // 3600
        
        health_checks.append({
            'component': 'System Uptime',
            'status': 'info',
            'message': f'{int(uptime_days)} days, {int(uptime_hours)} hours',
            'icon': 'info-circle',
            'color': 'info'
        })
    except:
        health_checks.append({
            'component': 'System Uptime',
            'status': 'info',
            'message': 'Uptime information unavailable',
            'icon': 'info-circle',
            'color': 'info'
        })
    
    return Response({
        'success': True,
        'health_checks': health_checks
    })

# Database Statistics
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_database_stats(request):
    """Get database statistics and table sizes"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Try PostgreSQL-specific query first
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    table_name,
                    pg_size_pretty(pg_total_relation_size('"' || table_schema || '"."' || table_name || '"')) as size,
                    (SELECT count(*) FROM "' || table_schema || '"."' || table_name || '") as row_count
                FROM information_schema.tables 
                WHERE table_schema = 'public'
                ORDER BY pg_total_relation_size('"' || table_schema || '"."' || table_name || '"') DESC;
            """)
            
            tables = []
            for row in cursor.fetchall():
                tables.append({
                    'table_name': row[0],
                    'size': row[1],
                    'row_count': row[2]
                })
        
        return Response({
            'success': True,
            'database_stats': {
                'tables': tables,
                'total_tables': len(tables)
            }
        })
        
    except Exception as e:
        # Fallback for SQLite or other databases
        try:
            models_stats = []
            models = [
                (CustomUser, 'CustomUser'),
                (ApplicantProfile, 'ApplicantProfile'),
                (JobListing, 'JobListing'),
                (Application, 'Application'),
                (Alert, 'Alert'),
                (Document, 'Document'),
                (Skill, 'Skill'),
                (EmploymentHistory, 'EmploymentHistory'),
                (Education, 'Education')
            ]
            
            for model, name in models:
                count = model.objects.count()
                models_stats.append({
                    'table_name': name,
                    'row_count': count,
                    'size': 'N/A'
                })
            
            return Response({
                'success': True,
                'database_stats': {
                    'tables': models_stats,
                    'total_tables': len(models_stats)
                }
            })
        except Exception as e2:
            logger.error(f"Error getting database stats: {str(e2)}")
            return Response({
                'success': False,
                'error': 'Failed to load database statistics'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Generate Reports
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_generate_report(request):
    """Generate various admin reports"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    report_type = request.data.get('type', 'users')
    
    try:
        if report_type == 'users':
            # User registration report
            users = CustomUser.objects.filter(user_type='applicant').values(
                'username', 'email', 'date_joined', 'last_login'
            )
            user_data = list(users)
            
            return Response({
                'success': True,
                'report_type': 'users',
                'data': user_data,
                'summary': {
                    'total_users': len(user_data),
                    'active_users': CustomUser.objects.filter(last_login__isnull=False).count(),
                    'new_this_month': CustomUser.objects.filter(
                        date_joined__month=timezone.now().month
                    ).count()
                }
            })
            
        elif report_type == 'applications':
            # Applications report
            applications = Application.objects.select_related(
                'applicant', 'job_listing'
            ).values(
                'id', 'status', 'applied_date',
                'applicant__first_name', 'applicant__last_name',
                'job_listing__title', 'job_listing__company_name'
            )
            application_data = list(applications)
            
            status_counts = Application.objects.values('status').annotate(
                count=Count('id')
            )
            
            return Response({
                'success': True,
                'report_type': 'applications',
                'data': application_data,
                'summary': {
                    'total_applications': len(application_data),
                    'status_breakdown': list(status_counts)
                }
            })
            
        elif report_type == 'jobs':
            # Jobs report
            jobs = JobListing.objects.values(
                'title', 'company_name', 'location', 'employment_type',
                'salary_range', 'status', 'apply_by', 'created_at'
            )
            job_data = list(jobs)
            
            return Response({
                'success': True,
                'report_type': 'jobs',
                'data': job_data,
                'summary': {
                    'total_jobs': len(job_data),
                    'active_jobs': JobListing.objects.filter(status='published').count(),
                    'expired_jobs': JobListing.objects.filter(apply_by__lt=timezone.now().date()).count()
                }
            })
            
        else:
            return Response({
                'success': False,
                'error': 'Invalid report type'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error generating {report_type} report: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to generate {report_type} report'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Export Data
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_export_data(request):
    """Export data in various formats (CSV, JSON, Excel)"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    export_type = request.data.get('type', 'users')
    format_type = request.data.get('format', 'csv')
    
    try:
        if format_type == 'csv':
            return export_csv_data(export_type)
        elif format_type == 'json':
            return export_json_data(export_type)
        elif format_type == 'excel':
            return export_excel_data(export_type)
        else:
            return Response({
                'success': False,
                'error': 'Unsupported format. Use csv, json, or excel.'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error exporting {export_type} data: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to export {export_type} data'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def export_csv_data(export_type):
    """Export data as CSV"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if export_type == 'users':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="users_export_{timestamp}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Username', 'Email', 'First Name', 'Last Name', 'Mobile Phone', 'Date Joined', 'Last Login'])
        
        users = CustomUser.objects.filter(user_type='applicant').values_list(
            'username', 'email', 'first_name', 'last_name', 'mobile_phone', 'date_joined', 'last_login'
        )
        
        for user in users:
            writer.writerow(user)
            
    elif export_type == 'applications':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="applications_export_{timestamp}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Applicant Name', 'Job Title', 'Company', 'Status', 'Applied Date', 'Reference Number'])
        
        applications = Application.objects.select_related('applicant', 'job_listing').values_list(
            'applicant__first_name', 'applicant__last_name', 
            'job_listing__title', 'job_listing__company_name',
            'status', 'applied_date'
        )
        
        for app in applications:
            applicant_name = f"{app[0]} {app[1]}" if app[0] and app[1] else "N/A"
            writer.writerow([applicant_name, app[2], app[3], app[4], app[5], f"APP-{app[5].strftime('%Y%m%d')}"])
            
    elif export_type == 'jobs':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="jobs_export_{timestamp}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Title', 'Company', 'Location', 'Employment Type', 'Salary Range', 'Status', 'Apply By', 'Created At'])
        
        jobs = JobListing.objects.values_list(
            'title', 'company_name', 'location', 'employment_type',
            'salary_range', 'status', 'apply_by', 'created_at'
        )
        
        for job in jobs:
            writer.writerow(job)
            
    elif export_type == 'profiles':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="profiles_export_{timestamp}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['User', 'Title', 'Gender', 'Ethnicity', 'Profile Completeness', 'Created At', 'Updated At'])
        
        profiles = ApplicantProfile.objects.select_related('user').values_list(
            'user__username', 'title', 'gender', 'ethnicity', 
            'profile_completeness', 'created_at', 'updated_at'
        )
        
        for profile in profiles:
            writer.writerow(profile)
            
    else:
        return HttpResponse('Invalid export type', status=400)
    
    return response

def export_json_data(export_type):
    """Export data as JSON"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if export_type == 'users':
        users = CustomUser.objects.filter(user_type='applicant').values(
            'username', 'email', 'first_name', 'last_name', 'mobile_phone', 
            'date_joined', 'last_login', 'user_type'
        )
        data = list(users)
        filename = f"users_export_{timestamp}.json"
        
    elif export_type == 'applications':
        applications = Application.objects.select_related('applicant', 'job_listing').values(
            'applicant__first_name', 'applicant__last_name', 
            'job_listing__title', 'job_listing__company_name',
            'status', 'applied_date', 'cover_letter'
        )
        data = list(applications)
        filename = f"applications_export_{timestamp}.json"
        
    elif export_type == 'jobs':
        jobs = JobListing.objects.values(
            'title', 'company_name', 'location', 'employment_type',
            'salary_range', 'status', 'apply_by', 'created_at', 'description'
        )
        data = list(jobs)
        filename = f"jobs_export_{timestamp}.json"
        
    elif export_type == 'profiles':
        profiles = ApplicantProfile.objects.select_related('user').values(
            'user__username', 'title', 'gender', 'ethnicity', 
            'profile_completeness', 'created_at', 'updated_at', 'introduction'
        )
        data = list(profiles)
        filename = f"profiles_export_{timestamp}.json"
        
    else:
        return HttpResponse('Invalid export type', status=400)
    
    response = HttpResponse(json.dumps(data, indent=2, default=str), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_excel_data(export_type):
    """Export data as Excel (requires openpyxl)"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('Excel export requires openpyxl package. Install with: pip install openpyxl', status=500)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    if export_type == 'users':
        worksheet.title = "Users"
        headers = ['Username', 'Email', 'First Name', 'Last Name', 'Mobile Phone', 'Date Joined', 'Last Login']
        worksheet.append(headers)
        
        users = CustomUser.objects.filter(user_type='applicant').values_list(
            'username', 'email', 'first_name', 'last_name', 'mobile_phone', 'date_joined', 'last_login'
        )
        
        for user in users:
            worksheet.append(user)
            
        filename = f"users_export_{timestamp}.xlsx"
        
    elif export_type == 'applications':
        worksheet.title = "Applications"
        headers = ['Applicant Name', 'Job Title', 'Company', 'Status', 'Applied Date', 'Reference Number']
        worksheet.append(headers)
        
        applications = Application.objects.select_related('applicant', 'job_listing').values_list(
            'applicant__first_name', 'applicant__last_name', 
            'job_listing__title', 'job_listing__company_name',
            'status', 'applied_date'
        )
        
        for app in applications:
            applicant_name = f"{app[0]} {app[1]}" if app[0] and app[1] else "N/A"
            worksheet.append([applicant_name, app[2], app[3], app[4], app[5], f"APP-{app[5].strftime('%Y%m%d')}"])
            
        filename = f"applications_export_{timestamp}.xlsx"
        
    elif export_type == 'jobs':
        worksheet.title = "Jobs"
        headers = ['Title', 'Company', 'Location', 'Employment Type', 'Salary Range', 'Status', 'Apply By', 'Created At']
        worksheet.append(headers)
        
        jobs = JobListing.objects.values_list(
            'title', 'company_name', 'location', 'employment_type',
            'salary_range', 'status', 'apply_by', 'created_at'
        )
        
        for job in jobs:
            worksheet.append(job)
            
        filename = f"jobs_export_{timestamp}.xlsx"
        
    elif export_type == 'profiles':
        worksheet.title = "Profiles"
        headers = ['User', 'Title', 'Gender', 'Ethnicity', 'Profile Completeness', 'Created At', 'Updated At']
        worksheet.append(headers)
        
        profiles = ApplicantProfile.objects.select_related('user').values_list(
            'user__username', 'title', 'gender', 'ethnicity', 
            'profile_completeness', 'created_at', 'updated_at'
        )
        
        for profile in profiles:
            worksheet.append(profile)
            
        filename = f"profiles_export_{timestamp}.xlsx"
        
    else:
        return HttpResponse('Invalid export type', status=400)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

# Quick Actions
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_quick_action(request):
    """Perform admin quick actions"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    action = request.data.get('action')
    
    try:
        if action == 'create_job':
            # Redirect to job creation page
            return Response({
                'success': True,
                'message': 'Redirecting to job creation page',
                'redirect_url': '/admin-portal/jobs/'
            })
            
        elif action == 'send_notification':
            # Placeholder for bulk notification
            return Response({
                'success': True,
                'message': 'Bulk notification feature would be implemented here'
            })
            
        elif action == 'export_data':
            # Use the real export function
            return api_export_data(request)
            
        elif action == 'system_check':
            # Run system health check
            return api_system_health(request)
            
        elif action == 'view_stats':
            # View database statistics
            return api_database_stats(request)
            
        else:
            return Response({
                'success': False,
                'error': 'Invalid action'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error performing quick action {action}: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to perform action: {action}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Recent Activity
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_recent_activity(request):
    """Get recent system activity"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Get recent users (last 5)
        recent_users = CustomUser.objects.filter(
            user_type='applicant'
        ).order_by('-date_joined')[:5]
        
        # Get recent applications (last 5)
        recent_applications = Application.objects.select_related(
            'applicant', 'job_listing'
        ).order_by('-applied_date')[:5]
        
        # Get recent jobs (last 5)
        recent_jobs = JobListing.objects.order_by('-created_at')[:5]
        
        activity_data = []
        
        # Add user registrations
        for user in recent_users:
            activity_data.append({
                'type': 'user_registration',
                'title': 'New user registration',
                'description': f'{user.get_full_name() or user.username} registered',
                'timestamp': user.date_joined,
                'badge': 'User'
            })
        
        # Add applications
        for app in recent_applications:
            activity_data.append({
                'type': 'application',
                'title': 'Job application submitted',
                'description': f'{app.applicant.first_name} {app.applicant.last_name} applied for {app.job_listing.title}',
                'timestamp': app.applied_date,
                'badge': 'Application'
            })
        
        # Add job listings
        for job in recent_jobs:
            activity_data.append({
                'type': 'job_listing',
                'title': 'New job listing published',
                'description': f'{job.title} position published',
                'timestamp': job.created_at,
                'badge': 'Job'
            })
        
        # Sort by timestamp
        activity_data.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return Response({
            'success': True,
            'activity': activity_data[:10]  # Return top 10 most recent
        })
        
    except Exception as e:
        logger.error(f"Error getting recent activity: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load recent activity'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# DEDICATED EXPORT VIEWS
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_export_data_simple(request):
    """Simple and reliable export function"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    export_type = request.data.get('type', 'users')
    format_type = request.data.get('format', 'csv')
    
    print(f"Export requested: {export_type} as {format_type}")  # Debug
    
    try:
        if format_type == 'csv':
            return export_simple_csv(export_type)
        elif format_type == 'json':
            return export_simple_json(export_type)
        else:
            return Response({
                'success': False,
                'error': 'Only CSV and JSON formats supported'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        print(f"Export error: {str(e)}")  # Debug
        return Response({
            'success': False,
            'error': f'Export failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def export_simple_csv(export_type):
    """Simple CSV export that definitely works"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(content_type='text/csv')
    
    if export_type == 'users':
        response['Content-Disposition'] = f'attachment; filename="users_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date Joined'])
        
        # Get limited data to ensure it works
        users = CustomUser.objects.filter(user_type='applicant')[:50]
        for user in users:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.first_name or '',
                user.last_name or '',
                user.date_joined.strftime('%Y-%m-%d') if user.date_joined else ''
            ])
            
    elif export_type == 'applications':
        response['Content-Disposition'] = f'attachment; filename="applications_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Applicant', 'Job Title', 'Company', 'Status', 'Applied Date'])
        
        applications = Application.objects.select_related('applicant', 'job_listing')[:50]
        for app in applications:
            applicant_name = f"{app.applicant.first_name} {app.applicant.last_name}".strip()
            if not applicant_name:
                applicant_name = app.applicant.username
                
            writer.writerow([
                app.id,
                applicant_name,
                app.job_listing.title if app.job_listing else 'N/A',
                app.job_listing.company_name if app.job_listing else 'N/A',
                app.status,
                app.applied_date.strftime('%Y-%m-%d') if app.applied_date else ''
            ])
            
    elif export_type == 'jobs':
        response['Content-Disposition'] = f'attachment; filename="jobs_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Company', 'Location', 'Employment Type', 'Status'])
        
        jobs = JobListing.objects.all()[:50]
        for job in jobs:
            writer.writerow([
                job.id,
                job.title,
                job.company_name,
                job.location,
                job.employment_type,
                job.status
            ])
            
    elif export_type == 'profiles':
        response['Content-Disposition'] = f'attachment; filename="profiles_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Title', 'Profile Completeness', 'Created At'])
        
        profiles = ApplicantProfile.objects.select_related('user')[:50]
        for profile in profiles:
            writer.writerow([
                profile.id,
                profile.user.username if profile.user else 'N/A',
                profile.title or '',
                f"{profile.profile_completeness}%",
                profile.created_at.strftime('%Y-%m-%d') if profile.created_at else ''
            ])
            
    else:
        return HttpResponse('Invalid export type', status=400)
    
    return response

def export_simple_json(export_type):
    """Simple JSON export that definitely works"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if export_type == 'users':
        data = []
        users = CustomUser.objects.filter(user_type='applicant')[:50]
        for user in users:
            data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'date_joined': user.date_joined.isoformat() if user.date_joined else None
            })
        filename = f"users_{timestamp}.json"
        
    elif export_type == 'applications':
        data = []
        applications = Application.objects.select_related('applicant', 'job_listing')[:50]
        for app in applications:
            data.append({
                'id': app.id,
                'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}".strip() or app.applicant.username,
                'job_title': app.job_listing.title if app.job_listing else 'N/A',
                'company': app.job_listing.company_name if app.job_listing else 'N/A',
                'status': app.status,
                'applied_date': app.applied_date.isoformat() if app.applied_date else None
            })
        filename = f"applications_{timestamp}.json"
        
    elif export_type == 'jobs':
        data = []
        jobs = JobListing.objects.all()[:50]
        for job in jobs:
            data.append({
                'id': job.id,
                'title': job.title,
                'company': job.company_name,
                'location': job.location,
                'employment_type': job.employment_type,
                'status': job.status
            })
        filename = f"jobs_{timestamp}.json"
        
    elif export_type == 'profiles':
        data = []
        profiles = ApplicantProfile.objects.select_related('user')[:50]
        for profile in profiles:
            data.append({
                'id': profile.id,
                'username': profile.user.username if profile.user else 'N/A',
                'title': profile.title,
                'profile_completeness': profile.profile_completeness,
                'created_at': profile.created_at.isoformat() if profile.created_at else None
            })
        filename = f"profiles_{timestamp}.json"
        
    else:
        return HttpResponse('Invalid export type', status=400)
    
    response = HttpResponse(
        json.dumps(data, indent=2, default=str), 
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# Test export function
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_test_export(request):
    """Test if export is working"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Create a simple test CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="test_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Test Column 1', 'Test Column 2', 'Test Column 3'])
        writer.writerow(['Data 1', 'Data 2', 'Data 3'])
        writer.writerow(['Data 4', 'Data 5', 'Data 6'])
        
        return response
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Test export failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
# Export page view
def export_data_page(request):
    """Dedicated export data page"""
    if not request.user.is_authenticated or not has_admin_access(request.user):  # FIXED: Use has_admin_access
        from django.shortcuts import redirect
        return redirect('/')
    return render(request, 'hiring/export_data.html')

#import exports
# ===== SIMPLE WORKING EXPORT FUNCTIONS =====
import csv
import json
from django.http import HttpResponse
from datetime import datetime

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_test_export(request):
    """Test export that definitely works"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Create a simple test CSV that definitely works
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="test_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Test ID', 'Test Name', 'Test Date'])
        writer.writerow(['1', 'Sample User', datetime.now().strftime('%Y-%m-%d')])
        writer.writerow(['2', 'Sample Job', datetime.now().strftime('%Y-%m-%d')])
        writer.writerow(['3', 'Sample Application', datetime.now().strftime('%Y-%m-%d')])
        
        return response
        
    except Exception as e:
        # If CSV fails, return a simple text file
        response = HttpResponse("Test export content\nThis is a test file to verify export functionality.", content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="test_export.txt"'
        return response

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_export_simple(request):
    """Simple and reliable export function"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        export_type = request.data.get('type', 'users')
        format_type = request.data.get('format', 'csv')
        
        print(f"Export requested: {export_type} as {format_type}")
        
        if format_type == 'csv':
            return export_simple_csv(export_type)
        elif format_type == 'json':
            return export_simple_json(export_type)
        else:
            return Response({
                'success': False,
                'error': 'Only CSV and JSON formats supported'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        print(f"Export error: {str(e)}")
        return Response({
            'success': False,
            'error': f'Export failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def export_simple_csv(export_type):
    """Simple CSV export that definitely works"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(content_type='text/csv')
    
    try:
        if export_type == 'users':
            response['Content-Disposition'] = f'attachment; filename="users_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date Joined'])
            
            # Get limited data to ensure it works
            users = CustomUser.objects.filter(user_type='applicant')[:20]  # Small limit for testing
            for user in users:
                writer.writerow([
                    user.id,
                    user.username,
                    user.email,
                    user.first_name or '',
                    user.last_name or '',
                    user.date_joined.strftime('%Y-%m-%d') if user.date_joined else ''
                ])
                
        elif export_type == 'applications':
            response['Content-Disposition'] = f'attachment; filename="applications_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(['ID', 'Applicant', 'Job Title', 'Company', 'Status', 'Applied Date'])
            
            applications = Application.objects.select_related('applicant', 'job_listing')[:20]
            for app in applications:
                applicant_name = f"{app.applicant.first_name} {app.applicant.last_name}".strip()
                if not applicant_name:
                    applicant_name = app.applicant.username
                    
                writer.writerow([
                    app.id,
                    applicant_name,
                    app.job_listing.title if app.job_listing else 'N/A',
                    app.job_listing.company_name if app.job_listing else 'N/A',
                    app.status,
                    app.applied_date.strftime('%Y-%m-%d') if app.applied_date else ''
                ])
                
        elif export_type == 'jobs':
            response['Content-Disposition'] = f'attachment; filename="jobs_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(['ID', 'Title', 'Company', 'Location', 'Employment Type', 'Status'])
            
            jobs = JobListing.objects.all()[:20]
            for job in jobs:
                writer.writerow([
                    job.id,
                    job.title,
                    job.company_name,
                    job.location,
                    job.employment_type,
                    job.status
                ])
                
        elif export_type == 'profiles':
            response['Content-Disposition'] = f'attachment; filename="profiles_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(['ID', 'Username', 'Title', 'Profile Completeness', 'Created At'])
            
            profiles = ApplicantProfile.objects.select_related('user')[:20]
            for profile in profiles:
                writer.writerow([
                    profile.id,
                    profile.user.username if profile.user else 'N/A',
                    profile.title or '',
                    f"{profile.profile_completeness}%",
                    profile.created_at.strftime('%Y-%m-%d') if profile.created_at else ''
                ])
                
        else:
            # Default fallback
            response['Content-Disposition'] = f'attachment; filename="export_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Type', 'Message'])
            writer.writerow([export_type, 'Export completed successfully'])
            
    except Exception as e:
        # If anything fails, return a basic CSV
        response['Content-Disposition'] = f'attachment; filename="error_export_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Error', 'Message'])
        writer.writerow(['Export Error', str(e)])
    
    return response

def export_simple_json(export_type):
    """Simple JSON export that definitely works"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        if export_type == 'users':
            data = []
            users = CustomUser.objects.filter(user_type='applicant')[:10]  # Small limit
            for user in users:
                data.append({
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'date_joined': user.date_joined.isoformat() if user.date_joined else None
                })
            filename = f"users_{timestamp}.json"
            
        elif export_type == 'applications':
            data = []
            applications = Application.objects.select_related('applicant', 'job_listing')[:10]
            for app in applications:
                data.append({
                    'id': app.id,
                    'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}".strip() or app.applicant.username,
                    'job_title': app.job_listing.title if app.job_listing else 'N/A',
                    'company': app.job_listing.company_name if app.job_listing else 'N/A',
                    'status': app.status,
                    'applied_date': app.applied_date.isoformat() if app.applied_date else None
                })
            filename = f"applications_{timestamp}.json"
            
        elif export_type == 'jobs':
            data = []
            jobs = JobListing.objects.all()[:10]
            for job in jobs:
                data.append({
                    'id': job.id,
                    'title': job.title,
                    'company': job.company_name,
                    'location': job.location,
                    'employment_type': job.employment_type,
                    'status': job.status
                })
            filename = f"jobs_{timestamp}.json"
            
        elif export_type == 'profiles':
            data = []
            profiles = ApplicantProfile.objects.select_related('user')[:10]
            for profile in profiles:
                data.append({
                    'id': profile.id,
                    'username': profile.user.username if profile.user else 'N/A',
                    'title': profile.title,
                    'profile_completeness': profile.profile_completeness,
                    'created_at': profile.created_at.isoformat() if profile.created_at else None
                })
            filename = f"profiles_{timestamp}.json"
            
        else:
            data = [{'error': 'Invalid export type', 'type': export_type}]
            filename = f"error_{timestamp}.json"
        
        response = HttpResponse(
            json.dumps(data, indent=2, default=str), 
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
    except Exception as e:
        # If JSON fails, return error as JSON
        error_data = {'error': f'Export failed: {str(e)}'}
        response = HttpResponse(
            json.dumps(error_data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="error_{timestamp}.json"'
    
    return response

# Simple health check that always works
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_simple_health_check(request):
    """Simple health check that always works"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    health_checks = [
        {
            'component': 'Export System',
            'status': 'healthy',
            'message': 'Export system is ready',
            'icon': 'check-circle',
            'color': 'success'
        },
        {
            'component': 'Database',
            'status': 'healthy',
            'message': 'Database is accessible',
            'icon': 'check-circle', 
            'color': 'success'
        },
        {
            'component': 'File System',
            'status': 'healthy',
            'message': 'File downloads enabled',
            'icon': 'check-circle',
            'color': 'success'
        }
    ]
    
    return Response({
        'success': True,
        'health_checks': health_checks
    }) 



# Employemenet

@api_view(['GET', 'POST', 'DELETE'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_employment(request, employment_id=None):
    """
    Handle employment data for both applicants and business users
    """
    # ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    # APPLICANT LOGIC
    if request.user.user_type == 'applicant':
        try:
            profile = ApplicantProfile.objects.get(user=request.user)
        except ApplicantProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Applicant profile not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            return get_applicant_employment(profile)
        
        elif request.method == 'POST':
            return create_applicant_employment(profile, request.data)
        
        elif request.method == 'DELETE':
            return delete_applicant_employment(profile, employment_id)
    
    # BUSINESS USER LOGIC
    else:
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
            
            if request.method == 'GET':
                return get_business_preferences(business_profile)
            
            elif request.method == 'POST':
                return create_business_preference(business_profile, request.data)
            
            elif request.method == 'DELETE':
                return delete_business_preference(business_profile, employment_id)
        
        except BusinessProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Business profile not found. Please complete your business profile first.'
            }, status=status.HTTP_404_NOT_FOUND)

# APPLICANT FUNCTIONS
def get_applicant_employment(profile):
    """Get employment history for applicant with business matches"""
    try:
        employment = profile.employment_history.all().order_by('-start_date')
        employment_data = EmploymentHistorySerializer(employment, many=True).data
        
        # Get business preferences for matching
        business_preferences = BusinessEmploymentPreference.objects.filter(is_active=True)
        matches = calculate_employment_matches(profile, business_preferences)
        
        return Response({
            'success': True,
            'employment': employment_data,
            'business_matches': matches,
            'user_type': 'applicant'
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error loading employment data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def create_applicant_employment(profile, data):
    """Create new employment history for applicant"""
    serializer = EmploymentCreateSerializer(data=data)
    if serializer.is_valid():
        try:
            employment = serializer.save(profile=profile)
            
            # Update profile completeness if function exists
            try:
                from .utils import calculate_profile_completeness  # Adjust import as needed
                profile.profile_completeness = calculate_profile_completeness(profile)
                profile.save()
            except:
                pass  # Skip if function doesn't exist
            
            return Response({
                'success': True,
                'employment': EmploymentHistorySerializer(employment).data,
                'message': 'Employment history added successfully',
                'completeness': getattr(profile, 'profile_completeness', 0)
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error saving employment: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

def delete_applicant_employment(profile, employment_id):
    """Delete employment history for applicant"""
    if not employment_id:
        return Response({
            'success': False,
            'error': 'Employment ID is required for deletion'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        employment = EmploymentHistory.objects.get(id=employment_id, profile=profile)
        employment.delete()
        
        # Update profile completeness if function exists
        try:
            from .utils import calculate_profile_completeness
            profile.profile_completeness = calculate_profile_completeness(profile)
            profile.save()
        except:
            pass
        
        return Response({
            'success': True,
            'message': 'Employment history deleted successfully',
            'completeness': getattr(profile, 'profile_completeness', 0)
        })
    except EmploymentHistory.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Employment history not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error deleting employment: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# BUSINESS FUNCTIONS
def get_business_preferences(business_profile):
    """Get employment preferences for business with applicant matches"""
    try:
        preferences = BusinessEmploymentPreference.objects.filter(
            business_profile=business_profile
        ).order_by('-created_at')
        
        preferences_data = []
        for preference in preferences:
            preference_data = BusinessEmploymentPreferenceSerializer(preference).data
            matching_applicants = find_matching_applicants(preference)
            preference_data['matching_applicants'] = matching_applicants
            preference_data['match_count'] = len(matching_applicants)
            preferences_data.append(preference_data)
        
        return Response({
            'success': True,
            'preferences': preferences_data,
            'user_type': 'business',
            'company_name': business_profile.company_name
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error loading preferences: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def create_business_preference(business_profile, data):
    """Create new employment preference for business"""
    serializer = BusinessEmploymentPreferenceCreateSerializer(data=data)
    if serializer.is_valid():
        try:
            preference = serializer.save(business_profile=business_profile)
            
            return Response({
                'success': True,
                'preference': BusinessEmploymentPreferenceSerializer(preference).data,
                'message': 'Employment preference added successfully'
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error saving preference: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

def delete_business_preference(business_profile, preference_id):
    """Delete employment preference for business"""
    if not preference_id:
        return Response({
            'success': False,
            'error': 'Preference ID is required for deletion'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        preference = BusinessEmploymentPreference.objects.get(
            id=preference_id, 
            business_profile=business_profile
        )
        preference_name = f"{preference.get_preferred_contract_type_display()} preference"
        preference.delete()
        
        return Response({
            'success': True,
            'message': f'Employment preference "{preference_name}" deleted successfully'
        })
    except BusinessEmploymentPreference.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Employment preference not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error deleting preference: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# EDIT/UPDATE FUNCTION FOR BOTH USER TYPES
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def update_employment(request, employment_id):
    """
    Handle updates for both applicant employment and business preferences
    """
    if not employment_id:
        return Response({
            'success': False,
            'error': 'ID is required for update'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # APPLICANT UPDATE
    if request.user.user_type == 'applicant':
        try:
            profile = ApplicantProfile.objects.get(user=request.user)
            employment = EmploymentHistory.objects.get(id=employment_id, profile=profile)
            
            serializer = EmploymentCreateSerializer(employment, data=request.data, partial=True)
            if serializer.is_valid():
                updated_employment = serializer.save()
                
                # Update profile completeness if function exists
                try:
                    from .utils import calculate_profile_completeness
                    profile.profile_completeness = calculate_profile_completeness(profile)
                    profile.save()
                except:
                    pass
                
                return Response({
                    'success': True,
                    'employment': EmploymentHistorySerializer(updated_employment).data,
                    'message': 'Employment history updated successfully',
                    'completeness': getattr(profile, 'profile_completeness', 0)
                })
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except EmploymentHistory.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Employment history not found'
            }, status=status.HTTP_404_NOT_FOUND)
    
    # BUSINESS UPDATE
    else:
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
            preference = BusinessEmploymentPreference.objects.get(
                id=employment_id, 
                business_profile=business_profile
            )
            
            serializer = BusinessEmploymentPreferenceCreateSerializer(
                preference, data=request.data, partial=True
            )
            if serializer.is_valid():
                updated_preference = serializer.save()
                
                return Response({
                    'success': True,
                    'preference': BusinessEmploymentPreferenceSerializer(updated_preference).data,
                    'message': 'Employment preference updated successfully'
                })
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except BusinessEmploymentPreference.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Employment preference not found'
            }, status=status.HTTP_404_NOT_FOUND)

# MATCHING CALCULATION FUNCTIONS
def calculate_experience_months(employment):
    """Calculate total months of experience for an employment record"""
    if not employment.start_date:
        return 0
    
    end_date = employment.end_date if not employment.currently_working else timezone.now().date()
    if not end_date:
        return 0
    
    try:
        months = (end_date.year - employment.start_date.year) * 12 + (end_date.month - employment.start_date.month)
        return max(0, months)
    except:
        return 0

def calculate_employment_matches(applicant_profile, business_preferences):
    """Calculate matches between applicant employment and business preferences"""
    matches = []
    applicant_employment = applicant_profile.employment_history.all()
    
    # Calculate total experience
    total_experience_months = sum(calculate_experience_months(emp) for emp in applicant_employment)
    
    for preference in business_preferences:
        match_score = 0
        match_reasons = []
        
        # 1. Check contract type match (30 points)
        applicant_contract_types = {emp.contract_type for emp in applicant_employment if emp.contract_type}
        if preference.preferred_contract_type in applicant_contract_types:
            match_score += 30
            contract_type_display = dict(EmploymentHistory.CONTRACT_TYPE_CHOICES).get(
                preference.preferred_contract_type, preference.preferred_contract_type
            )
            match_reasons.append(f"Experience with {contract_type_display} work")
        
        # 2. Check job title keywords match (40 points)
        if hasattr(preference, 'job_title_keywords') and preference.job_title_keywords:
            for employment in applicant_employment:
                if employment.job_title:
                    job_title_lower = employment.job_title.lower()
                    for keyword in preference.job_title_keywords:
                        if keyword and keyword.lower() in job_title_lower:
                            match_score += 40
                            match_reasons.append(f"Relevant job title: {employment.job_title}")
                            break
                    if match_score >= 70:  # If we found a match, break the loop
                        break
        
        # 3. Check experience match (30 points)
        required_months = preference.required_experience_years * 12
        if total_experience_months >= required_months:
            match_score += 30
            total_years = total_experience_months // 12
            match_reasons.append(f"Meets experience requirement: {total_years}+ years (needs {preference.required_experience_years}+)")
        
        # Only include matches with some score
        if match_score > 0:
            matches.append({
                'preference': BusinessEmploymentPreferenceSerializer(preference).data,
                'match_score': min(match_score, 100),  # Cap at 100%
                'match_reasons': match_reasons,
                'company_name': preference.business_profile.company_name,
                'total_experience_years': total_experience_months // 12
            })
    
    # Sort by match score (highest first)
    return sorted(matches, key=lambda x: x['match_score'], reverse=True)

def find_matching_applicants(preference):
    """Find applicants that match business employment preferences"""
    matching_applicants = []
    
    try:
        # Get all applicants with employment history
        all_applicants = ApplicantProfile.objects.filter(
            employment_history__isnull=False
        ).distinct()
        
        for applicant in all_applicants:
            # Calculate matches for this applicant
            matches = calculate_employment_matches(applicant, [preference])
            if matches:
                match_data = matches[0]  # Get the best match
                matching_applicants.append({
                    'applicant_id': applicant.id,
                    'applicant_name': applicant.user.get_full_name() or applicant.user.username,
                    'email': applicant.user.email,
                    'match_score': match_data['match_score'],
                    'match_reasons': match_data['match_reasons'],
                    'total_experience_years': match_data['total_experience_years']
                })
        
        # Sort by match score (highest first)
        return sorted(matching_applicants, key=lambda x: x['match_score'], reverse=True)
    
    except Exception as e:
        print(f"Error finding matching applicants: {e}")
        return []


#profile
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser, MultiPartParser])  # Added MultiPartParser for file uploads
def api_edit_profile(request):
    """Edit detailed profile information for both applicants and business users"""
    
    if request.user.user_type == 'applicant':
        # Your existing applicant logic
        try:
            profile = ApplicantProfile.objects.get(user=request.user)
        except ApplicantProfile.DoesNotExist:
            return Response({'error': 'Profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        data = request.data.copy()
        
        # Handle empty fields - set to None instead of empty string for optional fields
        empty_to_none_fields = ['title', 'gender', 'current_home_location', 'introduction']
        for field in empty_to_none_fields:
            if field in data and data[field] == '':
                data[field] = None
        
        # Handle empty strings for required fields
        if 'first_name' in data and data['first_name'] == '':
            data['first_name'] = ''
        
        if 'last_name' in data and data['last_name'] == '':
            data['last_name'] = ''
        
        print("Received data for profile update:", data)  # Debug
        
        # Update profile fields
        old_completeness = profile.profile_completeness
        serializer = ProfileUpdateSerializer(profile, data=data, partial=True)
        
        if serializer.is_valid():
            updated_profile = serializer.save()
            
            # Recalculate profile completeness
            profile.profile_completeness = calculate_profile_completeness(profile)
            profile.save()
            
            # Send notification if profile completeness improved significantly
            if profile.profile_completeness > old_completeness + 20:
                NotificationService.send_profile_reminder(profile)
            
            return Response({
                'success': True,
                'profile': ApplicantProfileSerializer(updated_profile).data,
                'completeness': profile.profile_completeness
            })
        
        print("Serializer errors:", serializer.errors)  # Debug
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.user.user_type == 'admin':
        # BUSINESS PROFILE EDITING LOGIC
        try:
            business_profile = BusinessProfile.objects.get(user=request.user)
        except BusinessProfile.DoesNotExist:
            # Create business profile if it doesn't exist
            business_profile = BusinessProfile.objects.create(user=request.user)
        
        data = request.data.copy()
        
        # Handle empty fields - set to None instead of empty string for optional fields
        empty_to_none_fields = [
            'company_description', 'website', 'phone_number', 'address', 
            'city', 'country', 'postal_code'
        ]
        for field in empty_to_none_fields:
            if field in data and data[field] == '':
                data[field] = None
        
        # Handle foreign key fields
        if 'industry' in data and data['industry'] == '':
            data['industry'] = None
        if 'company_size' in data and data['company_size'] == '':
            data['company_size'] = None
        
        # Handle file uploads
        if 'company_logo' in request.FILES:
            data['company_logo'] = request.FILES['company_logo']
        elif 'company_logo' in data and data['company_logo'] == '':
            data['company_logo'] = None
        
        if 'verification_document' in request.FILES:
            data['verification_document'] = request.FILES['verification_document']
        elif 'verification_document' in data and data['verification_document'] == '':
            data['verification_document'] = None
        
        # Handle boolean fields
        boolean_fields = ['receive_applicant_notifications', 'receive_newsletter']
        for field in boolean_fields:
            if field in data:
                if isinstance(data[field], str):
                    data[field] = data[field].lower() == 'true'
        
        print("Received data for business profile update:", data)  # Debug
        
        serializer = BusinessProfileSerializer(business_profile, data=data, partial=True)
        
        if serializer.is_valid():
            updated_profile = serializer.save()
            
            return Response({
                'success': True,
                'business_profile': BusinessProfileSerializer(updated_profile).data,
                'message': 'Business profile updated successfully'
            })
        
        print("Business profile serializer errors:", serializer.errors)  # Debug
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    else:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)



@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, JSONParser])
def api_edit_document(request, document_id):
    """Edit document information"""
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    document = get_object_or_404(Document, id=document_id, profile=profile)
    
    data = request.data.copy()
    
    # Handle empty description
    if 'description' in data and data['description'] == '':
        data['description'] = None
    
    # If a new file is uploaded, update the file_name
    if 'file' in request.FILES:
        data['file_name'] = request.FILES['file'].name
    
    serializer = DocumentCreateSerializer(document, data=data, partial=True)
    if serializer.is_valid():
        updated_document = serializer.save()
        
        return Response({
            'success': True,
            'document': DocumentSerializer(updated_document).data
        })
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_document_detail(request, document_id):
    """Get single document details"""
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    document = get_object_or_404(Document, id=document_id, profile=profile)
    
    serializer = DocumentSerializer(document)
    return Response({
        'success': True,
        'document': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_business_applicant_documents(request, application_id):
    """Get all documents for an applicant in a specific application - for business users"""
    try:
        # Check if user is a business user
        if request.user.user_type != 'admin':
            return Response({'error': 'Unauthorized - Business access only'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get business profile
        business_profile = get_object_or_404(BusinessProfile, user=request.user)
        
        # Get the application and verify it belongs to a job listing from this business
        application = get_object_or_404(
            Application, 
            id=application_id,
            job_listing__company_name=business_profile.company_name
        )
        
        # Get all documents for the applicant
        applicant_profile = application.applicant
        documents = applicant_profile.documents.all()
        
        serializer = DocumentSerializer(documents, many=True)
        
        return Response({
            'success': True,
            'applicant': {
                'id': applicant_profile.id,
                'first_name': applicant_profile.first_name,
                'last_name': applicant_profile.last_name,
                'email': applicant_profile.user.email
            },
            'job_listing': {
                'id': application.job_listing.id,
                'title': application.job_listing.title,
                'reference': application.job_listing.listing_reference
            },
            'documents': serializer.data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_business_applications_with_documents(request):
    """Get list of all applications with document access for a business"""
    try:
        # Check if user is a business user
        if request.user.user_type != 'admin':
            return Response({'error': 'Unauthorized - Business access only'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get business profile
        business_profile = get_object_or_404(BusinessProfile, user=request.user)
        
        # Get all applications for this business's job listings
        applications = Application.objects.filter(
            job_listing__company_name=business_profile.company_name
        ).select_related(
            'applicant', 
            'applicant__user',
            'job_listing'
        ).order_by('-applied_date')
        
        applications_data = []
        for application in applications:
            applicant_profile = application.applicant
            document_count = applicant_profile.documents.count()
            
            applications_data.append({
                'application_id': str(application.id),
                'applicant': {
                    'id': applicant_profile.id,
                    'first_name': applicant_profile.first_name,
                    'last_name': applicant_profile.last_name,
                    'email': applicant_profile.user.email,
                    'profile_completeness': applicant_profile.profile_completeness
                },
                'job_listing': {
                    'id': application.job_listing.id,
                    'title': application.job_listing.title,
                    'reference': application.job_listing.listing_reference,
                    'location': application.job_listing.location
                },
                'application_status': application.status,
                'applied_date': application.applied_date,
                'document_count': document_count,
                'has_documents': document_count > 0
            })
        
        return Response({
            'success': True,
            'applications': applications_data,
            'company_name': business_profile.company_name,
            'total_applications': len(applications_data)
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


        
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])  # Use MultiPartParser for file uploads
def api_add_document(request):
    """Upload a new document"""
    if request.user.user_type != 'applicant':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    profile = get_object_or_404(ApplicantProfile, user=request.user)
    
    # Create a mutable copy of the data
    data = request.data.copy()
    data['profile'] = profile.id
    
    serializer = DocumentCreateSerializer(data=data)
    if serializer.is_valid():
        document = serializer.save()
        
        return Response({
            'success': True,
            'document': DocumentSerializer(document).data,
            'message': 'Document uploaded successfully'
        })
    
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

#admin functions 
# ==================== ADMIN JOB MANAGEMENT VIEWS ====================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser, MultiPartParser])
def api_admin_jobs(request):
    """Admin job management - list all jobs and create new ones"""
    # Allow both admin and business users
    if not (has_admin_access(request.user) or has_business_access(request.user)):
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    if request.method == 'GET':
        # BUSINESS USER: Can only see their own company's jobs
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                company_name = business_profile.company_name
                jobs = JobListing.objects.filter(company_name=company_name).order_by('-created_at')
            except BusinessProfile.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Business profile not found'
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Superuser/Admin - show all jobs
            jobs = JobListing.objects.all().order_by('-created_at')
        
        # Apply status filter
        status_filter = request.GET.get('status', 'all')
        if status_filter != 'all':
            jobs = jobs.filter(status=status_filter)
        
        serializer = JobListingSerializer(jobs, many=True, context={'request': request})
        
        # Get counts for filters
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                company_name = business_profile.company_name
                total_jobs = JobListing.objects.filter(company_name=company_name).count()
                published_jobs = JobListing.objects.filter(company_name=company_name, status='published').count()
                draft_jobs = JobListing.objects.filter(company_name=company_name, status='draft').count()
                archived_jobs = JobListing.objects.filter(company_name=company_name, status='closed').count()
            except BusinessProfile.DoesNotExist:
                total_jobs = published_jobs = draft_jobs = archived_jobs = 0
        else:
            total_jobs = JobListing.objects.count()
            published_jobs = JobListing.objects.filter(status='published').count()
            draft_jobs = JobListing.objects.filter(status='draft').count()
            archived_jobs = JobListing.objects.filter(status='closed').count()
        
        return Response({
            'success': True,
            'jobs': serializer.data,
            'counts': {
                'total': total_jobs,
                'published': published_jobs,
                'draft': draft_jobs,
                'archived': archived_jobs
            }
        })
    
    elif request.method == 'POST':
        """Create new job listing with proper company isolation"""
        # BUSINESS USER: Must have business profile
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                BusinessProfile.objects.get(user=request.user)
            except BusinessProfile.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Business profile not found. Please complete your business profile first.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        data = request.data.copy()
        
        # Handle file upload for company logo
        if 'company_logo' in request.FILES:
            data['company_logo'] = request.FILES['company_logo']
        elif data.get('company_logo') in ['', 'null', 'undefined']:
            # Handle explicit logo removal
            data['company_logo'] = None
        
        # Remove company_name for business users (serializer will handle it)
        if has_business_access(request.user) and not request.user.is_superuser:
            data.pop('company_name', None)
        
        serializer = AdminJobCreateSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            job = serializer.save()
            
            # Log the creation
            logger.info(f"User {request.user.username} created job: {job.title} for company: {job.company_name}")
            
            return Response({
                'success': True,
                'job': JobListingSerializer(job, context={'request': request}).data,
                'message': 'Job created successfully'
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser, MultiPartParser])
def api_admin_job_detail(request, job_id):
    """Admin job detail management - get, update, or delete specific job"""
    # Allow both admin and business users
    if not (has_admin_access(request.user) or has_business_access(request.user)):
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        job = JobListing.objects.get(id=job_id)
        
        # BUSINESS USER: Can only access their own company's jobs
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                if job.company_name != business_profile.company_name:
                    return Response({
                        'success': False,
                        'error': 'Unauthorized - You can only access jobs from your own company'
                    }, status=status.HTTP_403_FORBIDDEN)
            except BusinessProfile.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Business profile not found'
                }, status=status.HTTP_400_BAD_REQUEST)
                
    except JobListing.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Job not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = JobListingSerializer(job, context={'request': request})
        
        # Get application count for this job
        application_count = Application.objects.filter(job_listing=job).count()
        
        return Response({
            'success': True,
            'job': serializer.data,
            'application_count': application_count
        })
    
    elif request.method == 'PUT':
        """Update job with proper company isolation"""
        data = request.data.copy()
        
        # Handle file upload for company logo
        if 'company_logo' in request.FILES:
            data['company_logo'] = request.FILES['company_logo']
        elif data.get('company_logo') in ['', 'null', 'undefined']:
            # Handle explicit logo removal
            data['company_logo'] = None
        
        # BUSINESS USER: Prevent changing company name
        if has_business_access(request.user) and not request.user.is_superuser:
            data.pop('company_name', None)
        
        serializer = AdminJobCreateSerializer(job, data=data, partial=True, context={'request': request})
        if serializer.is_valid():
            updated_job = serializer.save()
            
            logger.info(f"User {request.user.username} updated job: {job.title} for company: {job.company_name}")
            
            return Response({
                'success': True,
                'job': JobListingSerializer(updated_job, context={'request': request}).data,
                'message': 'Job updated successfully'
            })
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        # BUSINESS USER: Can only delete their own company's jobs
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                if job.company_name != business_profile.company_name:
                    return Response({
                        'success': False,
                        'error': 'Unauthorized - You can only delete jobs from your own company'
                    }, status=status.HTTP_403_FORBIDDEN)
            except BusinessProfile.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Business profile not found'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        job_title = job.title
        company_name = job.company_name
        job.delete()
        
        logger.info(f"User {request.user.username} deleted job: {job_title} from company: {company_name}")
        
        return Response({
            'success': True,
            'message': 'Job deleted successfully'
        })
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_job_status(request, job_id):
    """Update job status (publish, unpublish, archive, etc.)"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        job = JobListing.objects.get(id=job_id)
        
        # Business users can only access their own jobs
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                if job.company_name != business_profile.company_name:
                    return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
            except BusinessProfile.DoesNotExist:
                return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
                
    except JobListing.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Job not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    new_status = request.data.get('status')
    valid_statuses = ['draft', 'under_review', 'published', 'closed']
    
    if new_status not in valid_statuses:
        return Response({
            'success': False,
            'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    old_status = job.status
    job.status = new_status
    job.save()
    
    logger.info(f"Admin {request.user.username} changed job {job.title} status from {old_status} to {new_status}")
    
    return Response({
        'success': True,
        'message': f'Job status updated to {new_status}',
        'job': JobListingSerializer(job, context={'request': request}).data
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_job_applications(request, job_id):
    """Get all applications for a specific job"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        job = JobListing.objects.get(id=job_id)
        
        # Business users can only access their own jobs
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                if job.company_name != business_profile.company_name:
                    return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
            except BusinessProfile.DoesNotExist:
                return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
                
    except JobListing.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Job not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    applications = Application.objects.filter(job_listing=job).select_related(
        'applicant', 'applicant__user'
    ).order_by('-applied_date')
    
    application_data = []
    for app in applications:
        application_data.append({
            'id': str(app.id),
            'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}",
            'applicant_email': app.applicant.user.email,
            'status': app.status,
            'applied_date': app.applied_date,
            'cover_letter': app.cover_letter,
            'reference_number': f"APP-{app.id.hex[:8].upper()}"
        })
    
    return Response({
        'success': True,
        'job_title': job.title,
        'applications': application_data,
        'total_applications': len(application_data)
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_application_status(request, application_id):
    """Update application status"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        application = Application.objects.get(id=application_id)
    except Application.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Application not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    serializer = AdminApplicationStatusSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    new_status = serializer.validated_data['status']
    old_status = application.status
    application.status = new_status
    application.save()
    
    # Send notification to applicant
    try:
        Alert.objects.create(
            applicant=application.applicant,
            title="Application Status Updated",
            message=f"Your application for {application.job_listing.title} has been updated from {old_status} to {new_status}."
        )
    except Exception as e:
        logger.error(f"Failed to send notification for application status update: {str(e)}")
    
    logger.info(f"Admin {request.user.username} updated application {application_id} status to {new_status}")
    
    return Response({
        'success': True,
        'message': f'Application status updated to {new_status}',
        'application': {
            'id': application.id,
            'applicant_name': f"{application.applicant.first_name} {application.applicant.last_name}",
            'job_title': application.job_listing.title,
            'status': application.status
        }
    })

# ==================== ADMIN PAGE VIEWS ====================

@admin_required  # FIXED: Use the decorator
def admin_jobs_page(request):
    """Render admin jobs management page"""
    return render(request, 'hiring/admin_jobs.html')


@admin_required  # FIXED: Use the decorator
def admin_applications_page(request):
    """Render admin applications management page"""
    return render(request, 'hiring/admin_applications.html')


@admin_required  # FIXED: Use the decorator
def admin_analytics_page(request):
    """Render admin analytics page"""
    return render(request, 'hiring/admin_analytics.html')

#api edit job
#api edit job
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser, MultiPartParser])
def api_edit_job(request, job_id):
    """Edit existing job record with proper data handling for ALL fields"""
    if not has_admin_access(request.user):
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        job = JobListing.objects.get(id=job_id)
        
        # Business users can only access their own jobs
        if has_business_access(request.user) and not request.user.is_superuser:
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                if job.company_name != business_profile.company_name:
                    return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
            except BusinessProfile.DoesNotExist:
                return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
                
    except JobListing.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Job not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    data = request.data.copy()
    
    # Debug logging
    print(f"=== DEBUG: Editing job {job_id} ===")
    print(f"User: {request.user.username}")
    print(f"Is superuser: {request.user.is_superuser}")
    print(f"Has business access: {has_business_access(request.user)}")
    print(f"Request data keys: {data.keys()}")
    print(f"Request FILES: {request.FILES}")
    print(f"Current job company_name: {job.company_name}")
    print(f"Current job company_logo: {job.company_logo}")
    
    # Handle file upload for company logo - FIXED
    if 'company_logo' in request.FILES:
        print(f"DEBUG: New logo file uploaded: {request.FILES['company_logo'].name}")
        data['company_logo'] = request.FILES['company_logo']
    elif 'company_logo' in data:
        # Handle logo removal or keep existing
        if data['company_logo'] == '' or data['company_logo'] == 'null':
            print("DEBUG: Logo removal requested")
            
            # If business user, revert to business profile logo when removing
            if has_business_access(request.user) and not request.user.is_superuser:
                try:
                    business_profile = BusinessProfile.objects.get(user=request.user)
                    if business_profile.company_logo:
                        print(f"DEBUG: Using business profile logo: {business_profile.company_logo}")
                        data['company_logo'] = business_profile.company_logo
                    else:
                        data['company_logo'] = None
                except BusinessProfile.DoesNotExist:
                    data['company_logo'] = None
            else:
                data['company_logo'] = None
        elif data['company_logo'] is None:
            # Keep existing logo
            data.pop('company_logo', None)
    
    # Remove company_name for business users (cannot change it)
    if has_business_access(request.user) and not request.user.is_superuser:
        data.pop('company_name', None)
        print("DEBUG: Removed company_name for business user")
    
    # Handle empty fields - set to empty string instead of None for text fields
    text_fields = [
        'industry', 'job_category', 'contract_type', 'company_description',
        'knowledge_requirements', 'skills_requirements', 'competencies_requirements',
        'experience_requirements', 'education_requirements', 'position_summary',
        'job_description'
    ]
    
    for field in text_fields:
        if field in data and data[field] == '':
            data[field] = ''
    
    # Handle boolean field
    if 'ee_position' in data:
        if isinstance(data['ee_position'], str):
            data['ee_position'] = data['ee_position'].lower() in ['true', '1', 'yes']
        elif isinstance(data['ee_position'], bool):
            data['ee_position'] = data['ee_position']
    
    # Handle date field
    if 'apply_by' in data and not data['apply_by']:
        # Don't allow empty apply_by date
        return Response({
            'success': False,
            'error': 'Apply by date is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    print(f"DEBUG: Final data for update: {data}")
    
    # Use the updated serializer that includes all fields
    serializer = AdminJobCreateSerializer(job, data=data, partial=True, context={'request': request})
    
    if serializer.is_valid():
        updated_job = serializer.save()
        
        print(f"DEBUG: Job updated successfully")
        print(f"DEBUG: New company_logo: {updated_job.company_logo}")
        
        logger.info(f"Admin {request.user.username} updated job: {job.title}")
        
        return Response({
            'success': True,
            'job': JobListingSerializer(updated_job, context={'request': request}).data,
            'message': 'Job updated successfully'
        })
    
    print(f"DEBUG: Serializer errors: {serializer.errors}")
    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)


#status jobs
def connections_page(request):
    """Page to find and connect with other users"""
    return render(request, 'connections.html')

def messaging_page(request):
    """Main messaging page"""
    return render(request, 'messaging.html') 


def job_detail_page(request, job_id):
    """Render job detail page"""
    return render(request, 'hiring/job_detail.html')


@admin_required  # FIXED: Use the decorator
def admin_job_edit_page(request):
    """Render admin job edit page"""
    job_id = request.GET.get('job_id')
    
    # If no job_id provided, show empty form for creating new job
    if not job_id:
        context = {
            'page_title': 'Create New Job - Admin Portal',
            'job': None,
            'is_edit': False
        }
        return render(request, 'hiring/admin_job_edit.html', context)
    
    try:
        # Get the job from database
        job = get_object_or_404(JobListing, id=job_id)
        
        context = {
            'page_title': f'Edit Job - {job.title}',
            'job': job,
            'is_edit': True,
            'job_data': {
                'id': job.id,
                'title': job.title,
                'company_name': job.company_name,
                'location': job.location,
                'position_summary': job.position_summary,
                'job_description': job.job_description,
                'requirements': job.requirements,
                'benefits': job.benefits,
                'salary_range': job.salary_range,
                'employment_type': job.employment_type,
                'experience_level': job.experience_level,
                'education_level': job.education_level,
                'apply_by': job.apply_by.strftime('%Y-%m-%d') if job.apply_by else '',
                'status': job.status,
                'listing_reference': job.listing_reference,
                'contact_email': job.contact_email,
                'contact_phone': job.contact_phone,
                'website': job.website,
                'is_remote': job.is_remote,
                'vacancies': job.vacancies,
            }
        }
        return render(request, 'hiring/admin_job_edit.html', context)
        
    except Exception as e:
        # If there's an error, redirect back to jobs list
        return redirect('/admin-portal/jobs/')

@admin_required  # FIXED: Use the decorator
def save_job(request):
    """Save job data (create or update)"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            job_id = data.get('id')
            
            # Basic validation
            required_fields = ['title', 'location', 'position_summary', 'job_description']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({
                        'success': False,
                        'error': f'{field.replace("_", " ").title()} is required'
                    })
            
            if job_id:
                # Update existing job
                job = get_object_or_404(JobListing, id=job_id)
                action = 'updated'
            else:
                # Create new job
                job = JobListing()
                action = 'created'
            
            # Update job fields
            job.title = data.get('title', '')
            job.company_name = data.get('company_name', 'Benta Group')
            job.location = data.get('location', '')
            job.position_summary = data.get('position_summary', '')
            job.job_description = data.get('job_description', '')
            job.requirements = data.get('requirements', '')
            job.benefits = data.get('benefits', '')
            job.salary_range = data.get('salary_range', '')
            job.employment_type = data.get('employment_type', 'full_time')
            job.experience_level = data.get('experience_level', 'mid_level')
            job.education_level = data.get('education_level', 'bachelors')
            job.apply_by = data.get('apply_by')
            job.status = data.get('status', 'draft')
            job.contact_email = data.get('contact_email', '')
            job.contact_phone = data.get('contact_phone', '')
            job.website = data.get('website', '')
            job.is_remote = data.get('is_remote', False)
            job.vacancies = data.get('vacancies', 1)
            
            # Generate reference if not exists
            if not job.listing_reference:
                job.listing_reference = f"JOB-{job.id or 'NEW'}"
            
            job.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Job {action} successfully!',
                'job_id': job.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error saving job: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

@admin_required  # FIXED: Use the decorator
def get_job_data(request, job_id):
    """Get job data for editing"""
    try:
        job = get_object_or_404(JobListing, id=job_id)
        
        job_data = {
            'id': job.id,
            'title': job.title,
            'company_name': job.company_name,
            'location': job.location,
            'position_summary': job.position_summary,
            'job_description': job.job_description,
            'requirements': job.requirements,
            'benefits': job.benefits,
            'salary_range': job.salary_range,
            'employment_type': job.employment_type,
            'experience_level': job.experience_level,
            'education_level': job.education_level,
            'apply_by': job.apply_by.strftime('%Y-%m-%d') if job.apply_by else '',
            'status': job.status,
            'listing_reference': job.listing_reference,
            'contact_email': job.contact_email,
            'contact_phone': job.contact_phone,
            'website': job.website,
            'is_remote': job.is_remote,
            'vacancies': job.vacancies,
            'created_at': job.created_at.strftime('%Y-%m-%d %H:%M') if job.created_at else '',
            'updated_at': job.updated_at.strftime('%Y-%m-%d %H:%M') if job.updated_at else '',
        }
        
        return JsonResponse({
            'success': True,
            'job': job_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error loading job: {str(e)}'
        })


# ==================== ADMIN APPLICATION MANAGEMENT VIEWS ====================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_admin_applications_list(request):
    """Get list of all applications with filtering and pagination"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    if request.method == 'POST':
        try:
            data = request.data
            page = data.get('page', 1)
            page_size = data.get('page_size', 10)
            
            # Start with all applications
            applications = Application.objects.select_related(
                'applicant', 'applicant__user', 'job_listing'
            ).all().order_by('-applied_date')
            
            # Apply filters
            status_filter = data.get('status')
            if status_filter and status_filter != 'all':
                applications = applications.filter(status=status_filter)
            
            job_filter = data.get('job_id')
            if job_filter and job_filter != 'all':
                applications = applications.filter(job_listing_id=job_filter)
            
            date_filter = data.get('date')
            if date_filter:
                filter_date = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
                applications = applications.filter(applied_date__date=filter_date)
            
            search_query = data.get('search')
            if search_query:
                applications = applications.filter(
                    Q(applicant__first_name__icontains=search_query) |
                    Q(applicant__last_name__icontains=search_query) |
                    Q(applicant__user__email__icontains=search_query) |
                    Q(job_listing__title__icontains=search_query) |
                    Q(job_listing__company_name__icontains=search_query)
                )
            
            # Calculate pagination
            total_applications = applications.count()
            total_pages = (total_applications + page_size - 1) // page_size
            start_index = (page - 1) * page_size
            end_index = start_index + page_size
            
            # Get paginated applications
            paginated_applications = applications[start_index:end_index]
            
            # Prepare application data
            application_data = []
            for app in paginated_applications:
                application_data.append({
                    'id': str(app.id),
                    'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}",
                    'applicant_email': app.applicant.user.email,
                    'applicant_phone': app.applicant.user.mobile_phone,
                    'job_title': app.job_listing.title,
                    'company_name': app.job_listing.company_name,
                    'job_location': app.job_listing.location,
                    'job_reference': app.job_listing.listing_reference,
                    'status': app.status,
                    'applied_date': app.applied_date,
                    'cover_letter': app.cover_letter,
                    'notes': app.notes,
                    'reference_number': f"APP-{app.id.hex[:8].upper()}"
                })
            
            # Get statistics
            stats = {
                'total': Application.objects.count(),
                'submitted': Application.objects.filter(status='submitted').count(),
                'under_review': Application.objects.filter(status='under_review').count(),
                'shortlisted': Application.objects.filter(status='shortlisted').count(),
                'interview': Application.objects.filter(status='interview').count(),
                'successful': Application.objects.filter(status='successful').count(),
                'unsuccessful': Application.objects.filter(status='unsuccessful').count(),
            }
            
            return Response({
                'success': True,
                'applications': application_data,
                'total_pages': total_pages,
                'current_page': page,
                'total_applications': total_applications,
                'stats': stats
            })
            
        except Exception as e:
            logger.error(f"Error loading applications list: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to load applications'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_application_detail(request, application_id):
    """Get detailed information for a specific application"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        application = Application.objects.select_related(
            'applicant', 'applicant__user', 'job_listing'
        ).get(id=application_id)
        
        # Get applicant profile data
        profile = application.applicant
        
        application_data = {
            'id': str(application.id),
            'applicant_name': f"{profile.first_name} {profile.last_name}",
            'applicant_email': profile.user.email,
            'applicant_phone': profile.user.mobile_phone,
            'applicant_title': profile.title,
            'applicant_gender': profile.gender,
            'applicant_ethnicity': profile.ethnicity,
            'applicant_location': profile.current_home_location,
            'job_title': application.job_listing.title,
            'company_name': application.job_listing.company_name,
            'job_location': application.job_listing.location,
            'job_reference': application.job_listing.listing_reference,
            'status': application.status,
            'applied_date': application.applied_date,
            'cover_letter': application.cover_letter,
            'notes': application.notes,
            'reference_number': f"APP-{application.id.hex[:8].upper()}",
            'profile_completeness': profile.profile_completeness
        }
        
        return Response({
            'success': True,
            'application': application_data
        })
        
    except Application.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Application not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error loading application details: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load application details'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_update_application_status(request, application_id):
    """Update application status with notes"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        application = Application.objects.get(id=application_id)
    except Application.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Application not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    serializer = AdminApplicationStatusSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    new_status = serializer.validated_data['status']
    notes = request.data.get('notes', '')
    
    old_status = application.status
    application.status = new_status
    
    # Update notes if provided
    if notes:
        if application.notes:
            application.notes += f"\n\n--- Status Update {timezone.now().strftime('%Y-%m-%d %H:%M')} ---\n{notes}"
        else:
            application.notes = f"--- Status Update {timezone.now().strftime('%Y-%m-%d %H:%M')} ---\n{notes}"
    
    application.save()
    
    # Send notification to applicant
    try:
        Alert.objects.create(
            applicant=application.applicant,
            title="Application Status Updated",
            message=f"Your application for {application.job_listing.title} has been updated from {old_status.replace('_', ' ').title()} to {new_status.replace('_', ' ').title()}."
        )
        
        # Also create a sent notification
        SentNotification.objects.create(
            applicant=application.applicant,
            notification_type='application_status_update',
            subject="Application Status Update",
            message=f"Your application status for {application.job_listing.title} at {application.job_listing.company_name} has been changed to {new_status.replace('_', ' ').title()}.",
            sent_via='in_app'
        )
        
    except Exception as e:
        logger.error(f"Failed to send notification for application status update: {str(e)}")
    
    logger.info(f"Admin {request.user.username} updated application {application_id} status from {old_status} to {new_status}")
    
    return Response({
        'success': True,
        'message': f'Application status updated to {new_status}',
        'application': {
            'id': application.id,
            'applicant_name': f"{application.applicant.first_name} {application.applicant.last_name}",
            'job_title': application.job_listing.title,
            'status': application.status
        }
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_jobs_simple_list(request):
    """Get simple list of all jobs for filters"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        jobs = JobListing.objects.all().values('id', 'title', 'company_name').order_by('title')
        
        return Response({
            'success': True,
            'jobs': list(jobs)
        })
        
    except Exception as e:
        logger.error(f"Error loading jobs list: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load jobs list'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_application_stats(request):
    """Get application statistics for dashboard"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Today's date
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Basic counts
        total_applications = Application.objects.count()
        applications_today = Application.objects.filter(applied_date__date=today).count()
        applications_week = Application.objects.filter(applied_date__date__gte=week_ago).count()
        applications_month = Application.objects.filter(applied_date__date__gte=month_ago).count()
        
        # Status breakdown
        status_breakdown = Application.objects.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        # Recent activity (last 10 applications)
        recent_applications = Application.objects.select_related(
            'applicant', 'job_listing'
        ).order_by('-applied_date')[:10]
        
        recent_data = []
        for app in recent_applications:
            recent_data.append({
                'id': str(app.id),
                'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}",
                'job_title': app.job_listing.title,
                'status': app.status,
                'applied_date': app.applied_date,
                'company_name': app.job_listing.company_name
            })
        
        return Response({
            'success': True,
            'stats': {
                'total': total_applications,
                'today': applications_today,
                'this_week': applications_week,
                'this_month': applications_month,
                'status_breakdown': list(status_breakdown),
                'recent_applications': recent_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading application stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load application statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@admin_required  # FIXED: Use the decorator
def admin_applications_page(request):
    """Render admin applications management page"""
    return render(request, 'hiring/admin_applications.html')


# ==================== ADMIN USERS MANAGEMENT VIEWS ====================
@admin_required # FIXED: Use superuser_required
def admin_users_page(request):
    """Render admin users management page"""
    return render(request, 'hiring/admin_users.html')

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_admin_users_list(request):
    """Get list of all users with filtering and pagination"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        data = request.data
        page = data.get('page', 1)
        page_size = data.get('page_size', 10)
        
        # Start with all users
        users = CustomUser.objects.all().order_by('-date_joined')
        
        # Apply filters
        user_type_filter = data.get('user_type')
        if user_type_filter and user_type_filter != 'all':
            users = users.filter(user_type=user_type_filter)
        
        profile_status_filter = data.get('profile_status')
        if profile_status_filter and profile_status_filter != 'all':
            if profile_status_filter == 'complete':
                users = users.filter(applicantprofile__profile_completeness__gte=80)
            elif profile_status_filter == 'partial':
                users = users.filter(
                    applicantprofile__profile_completeness__gte=50,
                    applicantprofile__profile_completeness__lt=80
                )
            elif profile_status_filter == 'incomplete':
                users = users.filter(applicantprofile__profile_completeness__lt=50)
        
        date_filter = data.get('date')
        if date_filter:
            filter_date = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
            users = users.filter(date_joined__date=filter_date)
        
        search_query = data.get('search')
        if search_query:
            users = users.filter(
                Q(username__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query)
            )
        
        # Calculate pagination
        total_users = users.count()
        total_pages = (total_users + page_size - 1) // page_size
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        
        # Get paginated users
        paginated_users = users[start_index:end_index]
        
        # Prepare user data
        user_data = []
        for user in paginated_users:
            user_info = {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'full_name': user.get_full_name(),
                'user_type': user.user_type,
                'mobile_phone': user.mobile_phone,
                'date_joined': user.date_joined,
                'last_login': user.last_login,
                'has_profile': hasattr(user, 'applicantprofile'),
            }
            
            # Add profile information if exists
            if hasattr(user, 'applicantprofile'):
                profile = user.applicantprofile
                user_info.update({
                    'profile_completeness': profile.profile_completeness,
                    'profile_title': profile.title,
                    'preferred_job_title': profile.preferred_job_title,
                    'location': profile.current_home_location,
                    'profile_created': profile.created_at,
                    'profile_updated': profile.updated_at,
                    'applications_count': user.applicantprofile.applications.count(),
                    'skills_count': user.applicantprofile.skills.count(),
                    'documents_count': user.applicantprofile.documents.count(),
                    'alerts_count': user.applicantprofile.alerts.count(),
                })
            else:
                user_info.update({
                    'profile_completeness': None,
                    'profile_title': None,
                    'preferred_job_title': None,
                    'location': None,
                    'profile_created': None,
                    'profile_updated': None,
                    'applications_count': 0,
                    'skills_count': 0,
                    'documents_count': 0,
                    'alerts_count': 0,
                })
            
            user_data.append(user_info)
        
        # Get statistics
        today = timezone.now().date()
        stats = {
            'total': CustomUser.objects.count(),
            'applicants': CustomUser.objects.filter(user_type='applicant').count(),
            'admins': CustomUser.objects.filter(user_type='admin').count(),
            'active_today': CustomUser.objects.filter(last_login__date=today).count(),
            'new_today': CustomUser.objects.filter(date_joined__date=today).count(),
        }
        
        return Response({
            'success': True,
            'users': user_data,
            'total_pages': total_pages,
            'current_page': page,
            'total_users': total_users,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Error loading users list: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load users'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_user_detail(request, user_id):
    """Get detailed information for a specific user"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        user = CustomUser.objects.get(id=user_id)
        
        user_data = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'full_name': user.get_full_name(),
            'user_type': user.user_type,
            'mobile_phone': user.mobile_phone,
            'date_joined': user.date_joined,
            'last_login': user.last_login,
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'has_profile': hasattr(user, 'applicantprofile'),
        }
        
        # Add profile information if exists
        if hasattr(user, 'applicantprofile'):
            profile = user.applicantprofile
            user_data.update({
                'profile_completeness': profile.profile_completeness,
                'profile_title': profile.title,
                'profile_gender': profile.gender,
                'profile_ethnicity': profile.ethnicity,
                'preferred_job_title': profile.preferred_job_title,
                'location': profile.current_home_location,
                'has_drivers_license': profile.has_drivers_license,
                'has_own_transport': profile.has_own_transport,
                'willing_to_relocate': profile.willing_to_relocate,
                'availability': profile.availability,
                'current_salary': profile.current_salary,
                'desired_salary': profile.desired_salary,
                'introduction': profile.introduction,
                'profile_created': profile.created_at,
                'profile_updated': profile.updated_at,
                'applications_count': user.applicantprofile.applications.count(),
                'skills_count': user.applicantprofile.skills.count(),
                'employment_count': user.applicantprofile.employment_history.count(),
                'education_count': user.applicantprofile.education.count(),
                'documents_count': user.applicantprofile.documents.count(),
                'alerts_count': user.applicantprofile.alerts.count(),
            })
        
        return Response({
            'success': True,
            'user': user_data
        })
        
    except CustomUser.DoesNotExist:
        return Response({
            'success': False,
            'error': 'User not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error loading user details: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load user details'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_admin_update_user(request, user_id):
    """Update user information"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return Response({
            'success': False,
            'error': 'User not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    try:
        data = request.data
        
        # Update basic user fields
        if 'username' in data:
            user.username = data['username']
        if 'email' in data:
            user.email = data['email']
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'user_type' in data:
            user.user_type = data['user_type']
        if 'mobile_phone' in data:
            user.mobile_phone = data['mobile_phone']
        
        user.save()
        
        logger.info(f"Admin {request.user.username} updated user: {user.username}")
        
        return Response({
            'success': True,
            'message': 'User updated successfully',
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'user_type': user.user_type
            }
        })
        
    except Exception as e:
        logger.error(f"Error updating user {user_id}: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to update user: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_admin_delete_user(request, user_id):
    """Delete a user"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    # Prevent users from deleting themselves
    if request.user.id == user_id:
        return Response({
            'success': False,
            'error': 'Cannot delete your own account'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user = CustomUser.objects.get(id=user_id)
        username = user.username
        user.delete()
        
        logger.info(f"Admin {request.user.username} deleted user: {username}")
        
        return Response({
            'success': True,
            'message': 'User deleted successfully'
        })
        
    except CustomUser.DoesNotExist:
        return Response({
            'success': False,
            'error': 'User not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error deleting user {user_id}: {str(e)}")
        return Response({
            'success': False,
            'error': f'Failed to delete user: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@admin_required  # FIXED: Use superuser_required
def admin_users_page(request):
    """Render admin users management page"""
    return render(request, 'hiring/admin_users.html')

# ==================== ADMIN ANALYTICS VIEWS ====================

@admin_required  # FIXED: Use the decorator
def admin_analytics_page(request):
    """Render admin analytics page"""
    return render(request, 'hiring/admin_analytics.html')

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def api_admin_analytics(request):
    """Get comprehensive analytics data"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        data = request.data
        days = data.get('days', 30)
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Calculate date range
        if start_date and end_date:
            start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=days)
        
        # Basic statistics
        total_users = CustomUser.objects.filter(user_type='applicant').count()
        total_jobs = JobListing.objects.filter(status='published').count()
        total_applications = Application.objects.count()
        
        # Today's metrics
        today = timezone.now().date()
        new_users_today = CustomUser.objects.filter(
            user_type='applicant', 
            date_joined__date=today
        ).count()
        new_applications_today = Application.objects.filter(
            applied_date__date=today
        ).count()
        active_jobs_today = JobListing.objects.filter(
            status='published',
            apply_by__gte=today
        ).count()
        
        # Calculate conversion rate (applications per job)
        conversion_rate = 0
        if total_jobs > 0:
            conversion_rate = round((total_applications / total_jobs) * 100, 1)
        
        # Average profile completeness
        avg_profile_complete = ApplicantProfile.objects.aggregate(
            avg_completeness=Avg('profile_completeness')
        )['avg_completeness'] or 0
        avg_profile_complete = round(avg_profile_complete, 1)
        
        # Average response time (simplified - could be based on actual response times)
        avg_response_time = 24  # Placeholder - in hours
        
        # Trends (simplified - in real app, compare with previous period)
        trends = {
            'user_growth': 5.2,
            'job_growth': 2.1,
            'application_growth': 12.5,
            'conversion_trend': 1.8,
            'profile_completion_trend': 0.5,
            'response_time_trend': -2.3,
        }
        
        # User growth data (last 30 days)
        user_growth_data = []
        user_growth_labels = []
        for i in range(days, 0, -1):
            date = end_date - timedelta(days=i)
            count = CustomUser.objects.filter(
                user_type='applicant',
                date_joined__date=date
            ).count()
            user_growth_data.append(count)
            user_growth_labels.append(date.strftime('%m/%d'))
        
        # Applications overview (last 30 days)
        applications_data = []
        applications_labels = []
        for i in range(days, 0, -1):
            date = end_date - timedelta(days=i)
            count = Application.objects.filter(
                applied_date__date=date
            ).count()
            applications_data.append(count)
            applications_labels.append(date.strftime('%m/%d'))
        
        # Application status distribution
        status_distribution = Application.objects.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        status_labels = []
        status_data = []
        for item in status_distribution:
            status_labels.append(item['status'].replace('_', ' ').title())
            status_data.append(item['count'])
        
        # Top job locations
        top_locations = JobListing.objects.filter(status='published').values('location').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        location_labels = [item['location'] for item in top_locations]
        location_data = [item['count'] for item in top_locations]
        
        # Profile completion statistics
        profile_completion = {
            'high': ApplicantProfile.objects.filter(profile_completeness__gte=80).count(),
            'medium': ApplicantProfile.objects.filter(
                profile_completeness__gte=50, 
                profile_completeness__lt=80
            ).count(),
            'low': ApplicantProfile.objects.filter(profile_completeness__lt=50).count()
        }
        
        # Calculate percentages
        total_profiles = sum(profile_completion.values())
        if total_profiles > 0:
            profile_completion = {
                'high': round((profile_completion['high'] / total_profiles) * 100, 1),
                'medium': round((profile_completion['medium'] / total_profiles) * 100, 1),
                'low': round((profile_completion['low'] / total_profiles) * 100, 1)
            }
        
        # Activity timeline (registrations vs applications)
        activity_timeline = {
            'labels': [],
            'registrations': [],
            'applications': []
        }
        
        for i in range(7, 0, -1):
            date = end_date - timedelta(days=i)
            registrations = CustomUser.objects.filter(
                user_type='applicant',
                date_joined__date=date
            ).count()
            applications = Application.objects.filter(
                applied_date__date=date
            ).count()
            
            activity_timeline['labels'].append(date.strftime('%a'))
            activity_timeline['registrations'].append(registrations)
            activity_timeline['applications'].append(applications)
        
        # Popular jobs (most applications)
        popular_jobs = JobListing.objects.annotate(
            application_count=Count('applications')
        ).order_by('-application_count')[:5]
        
        popular_jobs_data = []
        for job in popular_jobs:
            successful_apps = job.applications.filter(status='successful').count()
            success_rate = 0
            if job.application_count > 0:
                success_rate = round((successful_apps / job.application_count) * 100, 1)
            
            popular_jobs_data.append({
                'title': job.title,
                'company': job.company_name,
                'applications': job.application_count,
                'success_rate': success_rate
            })
        
        analytics_data = {
            'stats': {
                'total_users': total_users,
                'active_jobs': total_jobs,
                'total_applications': total_applications,
                'conversion_rate': conversion_rate,
                'avg_profile_complete': avg_profile_complete,
                'avg_response_time': avg_response_time,
                'new_users_today': new_users_today,
                'new_applications_today': new_applications_today,
                'active_jobs_today': active_jobs_today,
            },
            'trends': trends,
            'user_growth': {
                'labels': user_growth_labels,
                'data': user_growth_data
            },
            'applications_overview': {
                'labels': applications_labels,
                'data': applications_data
            },
            'status_distribution': {
                'labels': status_labels,
                'data': status_data
            },
            'top_locations': {
                'labels': location_labels,
                'data': location_data
            },
            'profile_completion': profile_completion,
            'activity_timeline': activity_timeline,
            'popular_jobs': popular_jobs_data
        }
        
        return Response({
            'success': True,
            'analytics': analytics_data,
            'date_range': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading analytics data: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load analytics data'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@admin_required  # FIXED: Use the decorator
def admin_analytics_page(request):
    """Render admin analytics page"""
    return render(request, 'hiring/admin_analytics.html')


# ==================== ADMIN DASHBOARD VIEWS ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_stats(request):
    """Get admin statistics"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Calculate basic statistics - MAKE SURE TO FILTER BY BUSINESS
        if has_business_access(request.user) and not request.user.is_superuser:
            # Business user - only show their company's data
            try:
                business_profile = BusinessProfile.objects.get(user=request.user)
                total_jobs = JobListing.objects.filter(company_name=business_profile.company_name, status='published').count()
                # Get applications for this business's jobs
                business_jobs = JobListing.objects.filter(company_name=business_profile.company_name)
                total_applications = Application.objects.filter(job_listing__in=business_jobs).count()
            except BusinessProfile.DoesNotExist:
                total_jobs = 0
                total_applications = 0
        else:
            # Superuser - show all data
            total_jobs = JobListing.objects.filter(status='published').count()
            total_applications = Application.objects.count()
        
        total_users = CustomUser.objects.filter(user_type='applicant').count()
        total_companies = JobListing.objects.values('company_name').distinct().count()
        
        return Response({
            'success': True,
            'stats': {
                'total_users': total_users,
                'total_jobs': total_jobs,
                'total_applications': total_applications,
                'total_companies': total_companies
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading admin stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load admin statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_recent_activity(request):
    """Get recent system activity for admin dashboard"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Get recent user registrations (last 10)
        recent_users = CustomUser.objects.filter(
            user_type='applicant'
        ).order_by('-date_joined')[:5]
        
        # Get recent applications (last 10)
        recent_applications = Application.objects.select_related(
            'applicant', 'job_listing'
        ).order_by('-applied_date')[:5]
        
        # Get recent jobs (last 10)
        recent_jobs = JobListing.objects.order_by('-created_at')[:5]
        
        activity_data = []
        
        # Add user registrations
        for user in recent_users:
            activity_data.append({
                'action': 'User Registration',
                'description': f'{user.get_full_name() or user.username} registered as applicant',
                'timestamp': user.date_joined,
                'user': 'System'
            })
        
        # Add applications
        for app in recent_applications:
            activity_data.append({
                'action': 'Job Application',
                'description': f'{app.applicant.first_name} {app.applicant.last_name} applied for {app.job_listing.title}',
                'timestamp': app.applied_date,
                'user': f'{app.applicant.first_name} {app.applicant.last_name}'
            })
        
        # Add job listings
        for job in recent_jobs:
            activity_data.append({
                'action': 'Job Listing',
                'description': f'New job published: {job.title} at {job.company_name}',
                'timestamp': job.created_at,
                'user': 'Admin'
            })
        
        # Add some system alerts
        recent_alerts = Alert.objects.select_related('applicant').order_by('-created_at')[:3]
        for alert in recent_alerts:
            activity_data.append({
                'action': 'System Alert',
                'description': f'Alert sent to {alert.applicant.first_name} {alert.applicant.last_name}: {alert.title}',
                'timestamp': alert.created_at,
                'user': 'System'
            })
        
        # Sort by timestamp and take top 8
        activity_data.sort(key=lambda x: x['timestamp'], reverse=True)
        activity_data = activity_data[:8]
        
        return Response({
            'success': True,
            'activity': activity_data
        })
        
    except Exception as e:
        logger.error(f"Error loading recent activity: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load recent activity'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_quick_stats(request):
    """Get quick statistics for admin dashboard"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        
        # Today's applications
        today_applications = Application.objects.filter(
            applied_date__date=today
        ).count()
        
        # This week's applications
        week_applications = Application.objects.filter(
            applied_date__date__gte=week_ago
        ).count()
        
        # Pending jobs (draft status)
        pending_jobs = JobListing.objects.filter(status='draft').count()
        
        # New users in last 7 days
        new_users = CustomUser.objects.filter(
            user_type='applicant',
            date_joined__date__gte=week_ago
        ).count()
        
        # Active jobs (not expired)
        active_jobs = JobListing.objects.filter(
            status='published',
            apply_by__gte=today
        ).count()
        
        # Expired jobs
        expired_jobs = JobListing.objects.filter(
            apply_by__lt=today
        ).count()
        
        # Profile completion stats
        profile_stats = ApplicantProfile.objects.aggregate(
            avg_completeness=Avg('profile_completeness'),
            high_completeness=Count('id', filter=Q(profile_completeness__gte=80)),
            low_completeness=Count('id', filter=Q(profile_completeness__lt=50))
        )
        
        return Response({
            'success': True,
            'stats': {
                'today_applications': today_applications,
                'week_applications': week_applications,
                'pending_jobs': pending_jobs,
                'new_users': new_users,
                'active_jobs': active_jobs,
                'expired_jobs': expired_jobs,
                'avg_profile_complete': round(profile_stats['avg_completeness'] or 0, 1),
                'high_complete_profiles': profile_stats['high_completeness'] or 0,
                'low_complete_profiles': profile_stats['low_completeness'] or 0
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading quick stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load quick statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ==================== USER ALERTS VIEWS ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_user_alerts(request):
    """Get alerts for both applicants and business users - FIXED FOR TEMPLATES"""
    print(f"DEBUG: User: {request.user.username}, User Type: {request.user.user_type}")
    
    # ✅ ALLOW BOTH applicants AND business users
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if request.user.user_type == 'applicant':
            # ✅ APPLICANT LOGIC
            profile = get_object_or_404(ApplicantProfile, user=request.user)
            alerts = Alert.objects.filter(applicant=profile).order_by('-created_at')
            
            alert_data = []
            for alert in alerts:
                alert_data.append({
                    'id': alert.id,
                    'title': alert.title,
                    'message': alert.message,
                    'is_read': alert.is_read,
                    'created_at': alert.created_at,
                    'type': 'system',
                    'user_type': 'applicant'
                })
            
            return Response({
                'success': True,
                'alerts': alert_data,
                'user_type': 'applicant'
            })
            
        else:
            # ✅ BUSINESS USER LOGIC
            business_profile = BusinessProfile.objects.get(user=request.user)
            
            # Get business-specific alerts
            business_alerts = BusinessAlert.objects.filter(
                business=business_profile
            ).order_by('-created_at')
            
            # Also get system alerts
            system_alerts = Alert.objects.filter(
                Q(message__icontains=business_profile.company_name) |
                Q(title__icontains=business_profile.company_name)
            ).order_by('-created_at')[:50]
            
            alert_data = []
            
            # Add business alerts
            for alert in business_alerts:
                alert_data.append({
                    'id': alert.id,
                    'title': alert.title,
                    'message': alert.message,
                    'is_read': alert.is_read,
                    'created_at': alert.created_at,
                    'type': alert.alert_type,
                    'company_name': business_profile.company_name,
                    'user_type': 'business'
                })
            
            # Add system alerts
            for alert in system_alerts:
                alert_data.append({
                    'id': f"sys_{alert.id}",
                    'title': alert.title,
                    'message': alert.message,
                    'is_read': alert.is_read,
                    'created_at': alert.created_at,
                    'type': 'system',
                    'company_name': business_profile.company_name,
                    'user_type': 'business'
                })
            
            return Response({
                'success': True,
                'alerts': alert_data,
                'user_type': 'business',
                'company_name': business_profile.company_name
            })
            
    except BusinessProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Business profile not found. Please complete your business profile first.'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error loading alerts: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load alerts'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_mark_alert_read(request, alert_id):
    """Mark an alert as read for both user types - FIXED FOR TEMPLATES"""
    print(f"DEBUG: Mark as read - User: {request.user.username}, User Type: {request.user.user_type}")
    
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if request.user.user_type == 'applicant':
            profile = get_object_or_404(ApplicantProfile, user=request.user)
            alert = get_object_or_404(Alert, id=alert_id, applicant=profile)
        else:
            business_profile = BusinessProfile.objects.get(user=request.user)
            # Try BusinessAlert first, then fall back to system Alert
            try:
                alert = BusinessAlert.objects.get(id=alert_id, business=business_profile)
            except BusinessAlert.DoesNotExist:
                alert = get_object_or_404(Alert, id=alert_id)
        
        alert.is_read = True
        alert.save()
        
        return Response({
            'success': True,
            'message': 'Alert marked as read'
        })
        
    except Exception as e:
        logger.error(f"Error marking alert as read: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to mark alert as read'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_alert(request, alert_id):
    """Delete an alert for both user types - FIXED FOR TEMPLATES"""
    print(f"DEBUG: Delete alert - User: {request.user.username}, User Type: {request.user.user_type}")
    
    if request.user.user_type not in ['applicant', 'admin']:
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if request.user.user_type == 'applicant':
            profile = get_object_or_404(ApplicantProfile, user=request.user)
            alert = get_object_or_404(Alert, id=alert_id, applicant=profile)
        else:
            business_profile = BusinessProfile.objects.get(user=request.user)
            # Try BusinessAlert first, then fall back to system Alert
            try:
                alert = BusinessAlert.objects.get(id=alert_id, business=business_profile)
            except BusinessAlert.DoesNotExist:
                alert = get_object_or_404(Alert, id=alert_id)
        
        alert.delete()
        
        return Response({
            'success': True,
            'message': 'Alert deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting alert: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to delete alert'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Helper function ensures consistent user data everywhere
def serialize_user(user, include_online=True):
    return {
        'id': str(user.id),
        'username': user.username,
        'display_name': get_user_display_name(user),
        'is_online': user.profile.is_online if hasattr(user, 'profile') else False
    }



#Business Functions
# ==================== BUSINESS ADMIN VIEWS ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_business_dashboard_stats(request):
    """Get business-specific dashboard statistics"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Get business profile - this should exist for business admins
        business_profile = BusinessProfile.objects.get(user=request.user)
        company_name = business_profile.company_name
        
        # Calculate statistics for this business only
        total_jobs = JobListing.objects.filter(company_name=company_name, status='published').count()
        draft_jobs = JobListing.objects.filter(company_name=company_name, status='draft').count()
        
        # Get applications for this business's jobs
        business_jobs = JobListing.objects.filter(company_name=company_name)
        total_applications = Application.objects.filter(job_listing__in=business_jobs).count()
        
        # Recent applications
        recent_applications = Application.objects.filter(
            job_listing__in=business_jobs
        ).order_by('-applied_date')[:5]
        
        # Application status breakdown
        application_statuses = Application.objects.filter(
            job_listing__in=business_jobs
        ).values('status').annotate(count=Count('id')).order_by('status')
        
        return Response({
            'success': True,
            'stats': {
                'company_name': company_name,
                'total_jobs': total_jobs,
                'draft_jobs': draft_jobs,
                'total_applications': total_applications,
                'application_statuses': list(application_statuses)
            },
            'recent_applications': [
                {
                    'id': str(app.id),
                    'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}",
                    'job_title': app.job_listing.title,
                    'status': app.status,
                    'applied_date': app.applied_date
                }
                for app in recent_applications
            ]
        })
        
    except BusinessProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Business profile not found. Please complete your business profile.'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error loading business dashboard stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load dashboard statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_business_jobs(request):
    """Get jobs for the current business"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        business_profile = BusinessProfile.objects.get(user=request.user)
        company_name = business_profile.company_name
        
        # Get all jobs for this business
        jobs = JobListing.objects.filter(company_name=company_name).order_by('-created_at')
        
        serializer = JobListingSerializer(jobs, many=True, context={'request': request})
        
        return Response({
            'success': True,
            'jobs': serializer.data,
            'company_name': company_name
        })
        
    except BusinessProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Business profile not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_business_applications(request):
    """Get applications for the current business's jobs"""
    if not has_admin_access(request.user):  # FIXED: Use has_admin_access
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        business_profile = BusinessProfile.objects.get(user=request.user)
        company_name = business_profile.company_name
        
        # Get jobs for this business
        business_jobs = JobListing.objects.filter(company_name=company_name)
        
        # Get applications for these jobs
        applications = Application.objects.filter(
            job_listing__in=business_jobs
        ).select_related('applicant', 'job_listing').order_by('-applied_date')
        
        application_data = []
        for app in applications:
            application_data.append({
                'id': str(app.id),
                'applicant_name': f"{app.applicant.first_name} {app.applicant.last_name}",
                'applicant_email': app.applicant.user.email,
                'job_title': app.job_listing.title,
                'status': app.status,
                'applied_date': app.applied_date,
                'cover_letter': app.cover_letter,
                'reference_number': f"APP-{app.id.hex[:8].upper()}"
            })
        
        return Response({
            'success': True,
            'applications': application_data,
            'company_name': company_name,
            'total_applications': len(application_data)
        })
        
    except BusinessProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Business profile not found'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_business_stats(request):
    """Get dynamic business statistics for the current business user"""
    if not has_admin_access(request.user):
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Get business profile
        business_profile = BusinessProfile.objects.get(user=request.user)
        company_name = business_profile.company_name
        
        # Calculate active jobs (published and not expired)
        today = timezone.now().date()
        active_jobs = JobListing.objects.filter(
            company_name=company_name,
            status='published',
            apply_by__gte=today
        ).count()
        
        # Calculate total applications for this business
        business_jobs = JobListing.objects.filter(company_name=company_name)
        total_applications = Application.objects.filter(job_listing__in=business_jobs).count()
        
        # Calculate profile views (you'll need to implement this tracking)
        # For now, using a placeholder - you should implement actual view tracking
        profile_views = BusinessProfileView.objects.filter(
            business_profile=business_profile
        ).count() if hasattr(business_profile, 'businessprofileview') else 0
        
        # Calculate success rate (applications that resulted in successful hires)
        successful_applications = Application.objects.filter(
            job_listing__in=business_jobs,
            status='successful'
        ).count()
        
        success_rate = 0
        if total_applications > 0:
            success_rate = round((successful_applications / total_applications) * 100)
        
        return Response({
            'success': True,
            'stats': {
                'active_jobs': active_jobs,
                'total_applications': total_applications,
                'profile_views': profile_views,
                'success_rate': f'{success_rate}%',
                'company_name': company_name
            }
        })
        
    except BusinessProfile.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Business profile not found. Please complete your business profile.'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error loading business stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load business statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['POST'])
@permission_classes([AllowAny])
def api_record_business_profile_view(request, business_id):
    """Record when someone views a business profile"""
    try:
        business_profile = BusinessProfile.objects.get(id=business_id)
        
        BusinessProfileView.objects.create(
            business_profile=business_profile,
            viewer=request.user if request.user.is_authenticated else None,
            ip_address=get_client_ip(request)
        )
        
        return Response({'success': True})
        
    except BusinessProfile.DoesNotExist:
        return Response({'success': False, 'error': 'Business not found'})


#==================== post functions ======================

# ==================== POST FEED SYSTEM ====================

@api_view(['GET'])
@permission_classes([AllowAny])
def api_home_feed(request):
    """Get posts for home page feed with pagination and filtering - ALL users"""
    try:
        # Get and validate query parameters
        try:
            page = max(1, int(request.GET.get('page', 1)))
            page_size = min(100, max(1, int(request.GET.get('page_size', 10))))
        except ValueError:
            return error_response('Invalid page or page_size parameter')
        
        post_type = request.GET.get('type', 'all')
        sort_by = request.GET.get('sort', 'newest')
        search = request.GET.get('search', '').strip()
        
        # Validate parameters
        if post_type not in ALLOWED_POST_TYPES:
            return error_response(
                f'Invalid post type. Allowed: {", ".join(ALLOWED_POST_TYPES)}'
            )
        
        if sort_by not in ALLOWED_SORT_OPTIONS:
            return error_response(
                f'Invalid sort option. Allowed: {", ".join(ALLOWED_SORT_OPTIONS)}'
            )
        
        # Apply visibility filters for ALL users
        visibility_filters = get_user_visibility_filters(request.user)
        posts = Post.objects.filter(visibility_filters)
        
        # Filter by post type
        if post_type != 'all':
            posts = posts.filter(post_type=post_type)
        
        # Search filter (case-insensitive)
        if search:
            search_filter = (
                Q(title__icontains=search) |
                Q(content__icontains=search) |
                Q(tags__icontains=search) |
                Q(author__username__icontains=search) |
                Q(author__first_name__icontains=search) |
                Q(author__last_name__icontains=search)
            )
            posts = posts.filter(search_filter)
        
        # Apply sorting
        if sort_by == 'popular':
            # Calculate popularity score
            posts = posts.annotate(
                popularity_score=(
                    Count('likes') + 
                    Count('comments') * 2 + 
                    F('views') * 0.1 + 
                    F('shares') * 3
                )
            ).order_by('-popularity_score', '-created_at')
        elif sort_by == 'top':
            # Top posts based on rating
            posts = posts.filter(rating_count__gte=3).order_by('-average_rating', '-created_at')
        else:  # newest
            posts = posts.order_by('-created_at')
        
        # Get paginated data using helper
        paginated_data = get_paginated_data(
            queryset=posts,
            page=page,
            page_size=page_size,
            serializer_class=PostSerializer,
            context={'request': request}
        )
        
        # Get user stats if authenticated
        user_stats = None
        if request.user.is_authenticated:
            try:
                user = request.user
                
                # Check user type based on your custom user model
                if hasattr(user, 'user_type'):
                    if user.user_type == 'applicant':
                        try:
                            profile = ApplicantProfile.objects.get(user=user)
                            user_stats = {
                                'profile_completeness': getattr(profile, 'profile_completeness', 0),
                                'applications_count': getattr(profile.applications, 'count', lambda: 0)(),
                                'skills_count': getattr(profile.skills, 'count', lambda: 0)() if hasattr(profile, 'skills') else 0
                            }
                        except ApplicantProfile.DoesNotExist:
                            user_stats = {'error': 'Applicant profile not found'}
                    
                    elif user.user_type == 'business':
                        try:
                            business_profile = BusinessProfile.objects.get(user=user)
                            user_stats = {
                                'company_name': business_profile.company_name,
                                'jobs_count': JobListing.objects.filter(
                                    company=business_profile
                                ).count() if hasattr(JobListing, 'company') else 0
                            }
                        except BusinessProfile.DoesNotExist:
                            user_stats = {'error': 'Business profile not found'}
                    
                    else:  # admin, staff, or other types
                        user_stats = {
                            'user_type': user.user_type,
                            'is_staff': user.is_staff,
                            'is_superuser': user.is_superuser,
                            'total_posts': Post.objects.filter(author=user).count()
                        }
                else:
                    # Fallback for users without user_type
                    user_stats = {
                        'username': user.username,
                        'is_staff': user.is_staff,
                        'is_superuser': user.is_superuser
                    }
                    
            except Exception as e:
                logger.warning(f"Could not fetch user stats: {str(e)}")
                user_stats = {'error': 'Could not load stats'}
        
        return Response({
            'success': True,
            'posts': paginated_data['data'],
            'pagination': paginated_data['pagination'],
            'filters': {
                'current_type': post_type,
                'current_sort': sort_by,
                'search_query': search if search else None,
                'allowed_types': ALLOWED_POST_TYPES,
                'allowed_sort_options': ALLOWED_SORT_OPTIONS
            },
            'user_stats': user_stats,
            'user_info': {
                'is_authenticated': request.user.is_authenticated,
                'username': request.user.username if request.user.is_authenticated else None,
                'user_type': getattr(request.user, 'user_type', None) if request.user.is_authenticated else None
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading home feed: {str(e)}", exc_info=True)
        return error_response(
            'Failed to load feed. Please try again later.',
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_post_like_dislike(request, post_id):
    """Like, dislike, or remove reaction from a post - Using your existing ManyToMany fields"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can view this post
        if not can_user_view_post(post, request.user):
            return error_response(
                'You do not have permission to interact with this post',
                status_code=status.HTTP_403_FORBIDDEN
            )
        
        # Get action from request
        action = request.data.get('action', '').lower()
        
        if action not in ['like', 'dislike', 'remove']:
            return error_response('Invalid action. Use "like", "dislike", or "remove"')
        
        # Process the action
        if action == 'like':
            # Remove from dislikes if present
            post.dislikes.remove(request.user)
            # Add to likes
            post.likes.add(request.user)
            message = 'Post liked'
            
        elif action == 'dislike':
            # Remove from likes if present
            post.likes.remove(request.user)
            # Add to dislikes
            post.dislikes.add(request.user)
            message = 'Post disliked'
            
        else:  # remove
            # Remove from both
            post.likes.remove(request.user)
            post.dislikes.remove(request.user)
            message = 'Reaction removed'
        
        # Refresh from database to get accurate counts
        post.refresh_from_db()
        
        # Get counts
        likes_count = post.likes.count()
        dislikes_count = post.dislikes.count()
        
        # Create notification for like (not for dislike or remove)
        if action == 'like' and post.author != request.user:
            try:
                # Create alert based on user type
                if hasattr(post.author, 'applicantprofile'):
                    Alert.objects.create(
                        applicant=post.author.applicantprofile,
                        title="New Like",
                        message=f"{request.user.username} liked your post: '{post.title[:50]}...'"
                    )
                elif hasattr(post.author, 'business_profile'):
                    BusinessAlert.objects.create(
                        business=post.author.business_profile,
                        title="New Like",
                        message=f"{request.user.username} liked your post: '{post.title[:50]}...'",
                        alert_type='like'
                    )
            except Exception as notify_error:
                logger.warning(f"Notification error (non-critical): {notify_error}")
        
        return success_response(message, {
            'likes_count': likes_count,
            'dislikes_count': dislikes_count,
            'user_has_liked': post.likes.filter(id=request.user.id).exists(),
            'user_has_disliked': post.dislikes.filter(id=request.user.id).exists(),
            'action_taken': action
        })
        
    except Exception as e:
        logger.error(f"Error in post like/dislike: {str(e)}", exc_info=True)
        return error_response(
            'Failed to process reaction',
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )




def error_response(message, status_code=status.HTTP_400_BAD_REQUEST):
    """Helper function for error responses"""
    return {
        'success': False,
        'error': message
    }

def success_response(message, data=None, status_code=status.HTTP_200_OK):
    """Helper function for success responses"""
    response = {
        'success': True,
        'message': message
    }
    if data is not None:
        response.update(data)
    return response

def can_user_view_post(post, user):
    """Check if user can view a post"""
    if not user.is_authenticated:
        return post.visibility == 'public'
    
    if post.visibility == 'public':
        return True
    elif post.visibility == 'private':
        return user == post.author
    elif post.visibility == 'company':
        return user == post.author or (user.user_type == 'admin' and post.company and post.company.user == user)
    elif post.visibility == 'connections':
        return user == post.author  # You can add more logic here later
    return False



@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_post_comments(request, post_id):
    """Get or add comments to a post"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can view this post
        if not can_user_view_post(post, request.user):
            return Response({
                'success': False,
                'error': 'You do not have permission to view comments on this post'
            }, status=403)
        
        if request.method == 'GET':
            # Get all comments for this post (no parent = top-level comments)
            comments = Comment.objects.filter(post=post, parent_comment__isnull=True).order_by('-created_at')
            
            # Serialize comments
            from .serializers import CommentSerializer
            serializer = CommentSerializer(comments, many=True, context={'request': request})
            
            return Response({
                'success': True,
                'message': 'Comments loaded',
                'comments': serializer.data,
                'total_comments': post.comment_count,
                'post_id': post_id,
                'post_title': post.title
            }, status=200)
        
        elif request.method == 'POST':
            # Get comment content from request
            content = request.data.get('content', '').strip()
            if not content:
                return Response({
                    'success': False,
                    'error': 'Comment cannot be empty'
                }, status=400)
            
            # Check for parent comment ID
            parent_comment_id = request.data.get('parent_comment_id')
            parent_comment = None
            
            if parent_comment_id:
                try:
                    parent_comment = Comment.objects.get(id=parent_comment_id, post=post)
                except Comment.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'Parent comment not found'
                    }, status=404)
            
            # Create the comment
            comment = Comment.objects.create(
                post=post,
                author=request.user,
                content=content,
                parent_comment=parent_comment
            )
            
            # UPDATE COMMENT COUNT ON POST
            post.comment_count = Comment.objects.filter(post=post).count()
            post.save(update_fields=['comment_count'])
            
            # Create notification for post author (if not commenting on own post)
            if post.author != request.user:
                try:
                    if hasattr(post.author, 'applicantprofile'):
                        Alert.objects.create(
                            applicant=post.author.applicantprofile,
                            title="New Comment",
                            message=f"{request.user.username} commented on your post: '{post.title[:50]}...'"
                        )
                    elif hasattr(post.author, 'business_profile'):
                        BusinessAlert.objects.create(
                            business=post.author.business_profile,
                            title="New Comment",
                            message=f"{request.user.username} commented on your post: '{post.title[:50]}...'",
                            alert_type='comment'
                        )
                except Exception as notify_error:
                    logger.warning(f"Notification error: {notify_error}")
            
            # Also notify parent comment author if replying
            if parent_comment and parent_comment.author != request.user:
                try:
                    if hasattr(parent_comment.author, 'applicantprofile'):
                        Alert.objects.create(
                            applicant=parent_comment.author.applicantprofile,
                            title="Reply to Your Comment",
                            message=f"{request.user.username} replied to your comment on post: '{post.title[:50]}...'"
                        )
                    elif hasattr(parent_comment.author, 'business_profile'):
                        BusinessAlert.objects.create(
                            business=parent_comment.author.business_profile,
                            title="Reply to Your Comment",
                            message=f"{request.user.username} replied to your comment on post: '{post.title[:50]}...'",
                            alert_type='comment'
                        )
                except Exception as notify_error:
                    logger.warning(f"Notification error: {notify_error}")
            
            # Serialize the new comment
            from .serializers import CommentSerializer
            serializer = CommentSerializer(comment, context={'request': request})
            
            return Response({
                'success': True,
                'message': 'Comment added successfully',
                'comment': serializer.data,
                'comment_count': post.comment_count
            }, status=201)
            
    except Exception as e:
        logger.error(f"Error in post comments: {str(e)}", exc_info=True)
        return Response({
            'success': False,
            'error': 'Failed to process comments'
        }, status=500)
    


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_job_like_dislike(request, job_id):
    """Like, dislike, or remove reaction from a job posting - Using your existing ManyToMany fields"""
    try:
        job = get_object_or_404(JobListing, id=job_id)
        
        # Get action from request
        action = request.data.get('action', '').lower()
        
        if action not in ['like', 'dislike', 'remove']:
            return error_response('Invalid action. Use "like", "dislike", or "remove"')
        
        # Process the action using your existing ManyToMany fields
        if action == 'like':
            # Remove from dislikes if present
            job.disliked_by.remove(request.user)
            # Add to likes
            job.liked_by.add(request.user)
            message = 'Job liked'
            
        elif action == 'dislike':
            # Remove from likes if present
            job.liked_by.remove(request.user)
            # Add to dislikes
            job.disliked_by.add(request.user)
            message = 'Job disliked'
            
        else:  # remove
            # Remove from both
            job.liked_by.remove(request.user)
            job.disliked_by.remove(request.user)
            message = 'Reaction removed'
        
        # Refresh from database
        job.refresh_from_db()
        
        # Get updated counts
        likes_count = job.liked_by.count()
        dislikes_count = job.disliked_by.count()
        
        # Create notification for like (if job creator is different from liker)
        if action == 'like' and hasattr(job, 'created_by') and job.created_by != request.user:
            try:
                if hasattr(job.created_by, 'business_profile'):
                    BusinessAlert.objects.create(
                        business=job.created_by.business_profile,
                        title="Job Liked",
                        message=f"{request.user.username} liked your job: '{job.title[:50]}...'",
                        alert_type='like'
                    )
            except Exception as notify_error:
                logger.warning(f"Notification error (non-critical): {notify_error}")
        
        return success_response(message, {
            'likes_count': likes_count,
            'dislikes_count': dislikes_count,
            'user_has_liked': job.liked_by.filter(id=request.user.id).exists(),
            'user_has_disliked': job.disliked_by.filter(id=request.user.id).exists(),
            'action_taken': action
        })
        
    except Exception as e:
        logger.error(f"Error in job like/dislike: {str(e)}", exc_info=True)
        return error_response(
            'Failed to process reaction',
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )



#post function to post the job updates
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, JSONParser])
def api_posts(request):
    """Get ALL posts or create a new post"""
    
    # ===== GET REQUEST =====
    if request.method == 'GET':
        try:
            # Get query parameters
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            post_type = request.GET.get('type', 'all')
            show_all = request.GET.get('show_all', 'false').lower() == 'true'
            
            # Get posts - ALL posts visible to user
            if show_all:
                # Show all public posts and user's own posts
                posts = Post.objects.filter(
                    models.Q(is_published=True) | 
                    models.Q(author=request.user)
                ).distinct()
            else:
                # Show only user's own posts
                posts = Post.objects.filter(author=request.user)
            
            # Filter by post type
            if post_type != 'all':
                posts = posts.filter(post_type=post_type)
            
            # Apply pagination
            total_posts = posts.count()
            total_pages = (total_posts + page_size - 1) // page_size
            
            start_index = (page - 1) * page_size
            end_index = start_index + page_size
            paginated_posts = posts.order_by('-created_at')[start_index:end_index]
            
            # Serialize posts
            serializer = PostSerializer(
                paginated_posts, 
                many=True, 
                context={'request': request}
            )
            
            return Response({
                'success': True,
                'posts': serializer.data,
                'pagination': {
                    'current_page': page,
                    'total_pages': total_pages,
                    'total_posts': total_posts,
                    'page_size': page_size
                },
                'show_all': show_all
            })
            
        except Exception as e:
            logger.error(f"Error loading posts: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to load posts'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    # ===== POST REQUEST =====
    elif request.method == 'POST':
        """Create a new post - ANY authenticated user can post"""
        try:
            # Log request details for debugging
            print(f"\n{'='*50}")
            print(f"POST REQUEST FROM: {request.user.username}")
            print(f"User Type: {request.user.user_type}")
            print(f"Data fields: {list(request.data.keys())}")
            print(f"{'='*50}\n")
            
            # Create a mutable copy of request.data
            post_data = request.data.copy()
            
            # Check if title is provided
            if 'title' not in post_data or not post_data.get('title'):
                # Try to generate a title from content
                content = post_data.get('content', '')
                if content:
                    # Generate title from first few words of content
                    words = content.strip().split()[:5]
                    generated_title = ' '.join(words) + '...'
                    post_data['title'] = generated_title
                    print(f"Generated title: {generated_title}")
                else:
                    return Response({
                        'success': False,
                        'error': 'Title is required. Please provide a title for your post.',
                        'tip': 'Either add a "title" field or ensure you have content to auto-generate title'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Add default values if not provided
            if 'post_type' not in post_data:
                post_data['post_type'] = 'general'
            
            if 'visibility' not in post_data:
                post_data['visibility'] = 'public'
            
            if 'is_published' not in post_data:
                post_data['is_published'] = True
            
            print(f"Final post data: {post_data}")
            
            # Create serializer with context
            serializer = PostCreateSerializer(
                data=post_data,
                context={'request': request}
            )
            
            # Validate the data
            if serializer.is_valid():
                print(f"Serializer is VALID")
                print(f"Validated data: {serializer.validated_data}")
                
                # Save the post
                post = serializer.save(author=request.user)
                
                print(f"✓ Post created successfully! ID: {post.id}")
                print(f"Post created successfully! ID: {post.id}, Title: '{post.title}'")
                
                # Create notification if published
                if post.is_published:
                    try:
                        # For business users
                        if hasattr(request.user, 'business_profile'):
                            BusinessAlert.objects.create(
                                business=request.user.business_profile,
                                title="New Post Published",
                                message=f"Your post '{post.title}' has been published successfully.",
                                alert_type='custom'
                            )
                        # For applicants
                        elif hasattr(request.user, 'applicantprofile'):
                            Alert.objects.create(
                                applicant=request.user.applicantprofile,
                                title="New Post Published",
                                message=f"Your post '{post.title}' has been published successfully."
                            )
                        # For admins/superusers without profiles
                        else:
                            print(f"Post published by admin/superuser: {post.title}")
                    except Exception as notify_error:
                        print(f"Notification error (non-critical): {notify_error}")
                        # Don't fail the post creation because of notification error
                
                # Increment user activity
                request.user.last_activity = timezone.now()
                request.user.save()
                
                # Get the full post data for response
                post_serializer = PostSerializer(post, context={'request': request})
                
                return Response({
                    'success': True,
                    'message': 'Post created successfully!',
                    'post': post_serializer.data,
                    'user_info': {
                        'username': request.user.username,
                        'user_type': request.user.user_type,
                        'is_superuser': request.user.is_superuser
                    }
                }, status=status.HTTP_201_CREATED)
            
            else:
                # Validation failed
                print(f"SERIALIZER ERRORS: {serializer.errors}")
                
                # Format better error messages
                error_list = []
                for field, errors in serializer.errors.items():
                    if isinstance(errors, list):
                        for error in errors:
                            error_list.append(f"{field}: {error}")
                    else:
                        error_list.append(f"{field}: {errors}")
                
                return Response({
                    'success': False,
                    'message': 'Please fix the errors below',
                    'errors': serializer.errors,
                    'error_list': error_list,
                    'required_fields': ['title', 'content'],
                    'tip': 'Make sure you include both "title" and "content" fields in your request'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            # Log the full error with traceback
            import traceback
            error_details = traceback.format_exc()
            print(f"\n{'!'*50}")
            print(f"ERROR in api_posts POST:")
            print(f"Error: {str(e)}")
            print(f"{'!'*50}\n")
            
            logger.error(f"Error creating post: {str(e)}\n{error_details}")
            
            return Response({
                'success': False,
                'message': 'An error occurred while creating your post',
                'error': str(e),
                'user': request.user.username if request.user else 'Anonymous'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, JSONParser])
def api_post_detail(request, post_id):
    """Get, update, or delete a specific post"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can access this post
        if not post.can_user_view(request.user):
            return Response({
                'success': False,
                'error': 'You do not have permission to access this post'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if request.method == 'GET':
            # Increment view count if not the author
            if request.user != post.author:
                post.increment_views()
            
            serializer = PostSerializer(post, context={'request': request})
            
            # Get comments for this post
            comments = Comment.objects.filter(
                post=post, 
                parent_comment__isnull=True
            ).order_by('-created_at')
            comment_serializer = CommentSerializer(
                comments, 
                many=True, 
                context={'request': request}
            )
            
            return Response({
                'success': True,
                'post': serializer.data,
                'comments': comment_serializer.data,
                'can_edit': request.user == post.author or request.user.is_staff,
                'can_delete': request.user == post.author or request.user.is_staff
            })
        
        elif request.method == 'PUT':
            # Check if user can edit
            if request.user != post.author and not request.user.is_staff:
                return Response({
                    'success': False,
                    'error': 'You do not have permission to edit this post'
                }, status=status.HTTP_403_FORBIDDEN)
            
            serializer = PostUpdateSerializer(
                post, 
                data=request.data, 
                partial=True,
                context={'request': request}
            )
            
            if serializer.is_valid():
                updated_post = serializer.save()
                
                return Response({
                    'success': True,
                    'message': 'Post updated successfully!',
                    'post': PostSerializer(updated_post, context={'request': request}).data
                })
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            # Check if user can delete
            if request.user != post.author and not request.user.is_staff:
                return Response({
                    'success': False,
                    'error': 'You do not have permission to delete this post'
                }, status=status.HTTP_403_FORBIDDEN)
            
            post_title = post.title
            post.delete()
            
            return Response({
                'success': True,
                'message': f'Post "{post_title}" deleted successfully'
            })
            
    except Exception as e:
        logger.error(f"Error in post detail view: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process request'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_post_like_dislike(request, post_id):
    """Like, dislike, or remove reaction from a post"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        if not post.can_user_view(request.user):
            return Response({
                'success': False,
                'error': 'You do not have permission to interact with this post'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = LikeDislikeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        action = serializer.validated_data['action']
        
        if action == 'like':
            # Remove from dislikes if present
            post.dislikes.remove(request.user)
            # Add to likes
            post.likes.add(request.user)
            message = 'Post liked'
            
        elif action == 'dislike':
            # Remove from likes if present
            post.likes.remove(request.user)
            # Add to dislikes
            post.dislikes.add(request.user)
            message = 'Post disliked'
            
        else:  # remove
            # Remove from both
            post.likes.remove(request.user)
            post.dislikes.remove(request.user)
            message = 'Reaction removed'
        
        # Update like/dislike counts
        likes_count = post.likes.count()
        dislikes_count = post.dislikes.count()
        
        return Response({
            'success': True,
            'message': message,
            'likes_count': likes_count,
            'dislikes_count': dislikes_count,
            'user_has_liked': post.likes.filter(id=request.user.id).exists(),
            'user_has_disliked': post.dislikes.filter(id=request.user.id).exists()
        })
        
    except Exception as e:
        logger.error(f"Error in post like/dislike: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process reaction'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_feed_posts(request):
    """Get feed of all public posts"""
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        # Get all public posts and posts from people user follows
        posts = Post.objects.filter(
            is_published=True,
            visibility='public'
        ).order_by('-created_at')
        
        # Apply pagination
        total_posts = posts.count()
        total_pages = (total_posts + page_size - 1) // page_size
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_posts = posts[start_index:end_index]
        
        serializer = PostSerializer(
            paginated_posts, 
            many=True, 
            context={'request': request}
        )
        
        return Response({
            'success': True,
            'posts': serializer.data,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_posts': total_posts,
                'page_size': page_size
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading feed posts: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load feed'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_post_share(request, post_id):
    """Share a post"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        if not post.can_user_view(request.user):
            return Response({
                'success': False,
                'error': 'You do not have permission to share this post'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Increment share count
        post.increment_shares()
        
        return Response({
            'success': True,
            'message': 'Post shared successfully',
            'shares_count': post.shares
        })
        
    except Exception as e:
        logger.error(f"Error sharing post: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to share post'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_post_rating(request, post_id):
    """Rate a post (1-5 stars)"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        if not post.can_user_view(request.user):
            return Response({
                'success': False,
                'error': 'You do not have permission to rate this post'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = RatingSerializer(
            data=request.data, 
            context={'request': request, 'post': post}
        )
        
        if serializer.is_valid():
            rating = serializer.save()
            
            return Response({
                'success': True,
                'message': 'Rating submitted successfully',
                'average_rating': post.average_rating,
                'rating_count': post.rating_count,
                'user_rating': rating.rating
            })
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        logger.error(f"Error rating post: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to submit rating'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_post_comments(request, post_id):
    """Get or add comments to a post"""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can view this post
        if not can_user_view_post(post, request.user):
            return Response({
                'success': False,
                'error': 'You do not have permission to view comments on this post'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if request.method == 'GET':
            # Get query parameters for pagination
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            
            # Get all comments for this post (no parent = top-level comments)
            comments = Comment.objects.filter(post=post, parent_comment__isnull=True)
            total_comments = comments.count()
            total_pages = (total_comments + page_size - 1) // page_size
            
            # Apply pagination
            start_index = (page - 1) * page_size
            end_index = start_index + page_size
            paginated_comments = comments.order_by('-created_at')[start_index:end_index]
            
            # Serialize comments
            from .serializers import CommentSerializer
            serializer = CommentSerializer(paginated_comments, many=True, context={'request': request})
            
            return Response({
                'success': True,
                'message': 'Comments loaded',
                'comments': serializer.data,
                'total_comments': total_comments,
                'pagination': {
                    'current_page': page,
                    'total_pages': total_pages,
                    'page_size': page_size
                },
                'post_id': post_id,
                'post_title': post.title
            }, status=status.HTTP_200_OK)
        
        elif request.method == 'POST':
            # Get comment content from request
            content = request.data.get('content', '').strip()
            if not content:
                return Response({
                    'success': False,
                    'error': 'Comment cannot be empty'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check for parent comment ID
            parent_comment_id = request.data.get('parent_comment_id')
            parent_comment = None
            
            if parent_comment_id:
                try:
                    parent_comment = Comment.objects.get(id=parent_comment_id, post=post)
                except Comment.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'Parent comment not found'
                    }, status=status.HTTP_404_NOT_FOUND)
            
            # Create the comment
            comment = Comment.objects.create(
                post=post,
                author=request.user,
                content=content,
                parent_comment=parent_comment
            )
            
            # UPDATE COMMENT COUNT ON POST
            post.comment_count = Comment.objects.filter(post=post).count()
            post.save(update_fields=['comment_count'])
            
            # Create notification for post author (if not commenting on own post)
            if post.author != request.user:
                try:
                    if hasattr(post.author, 'applicantprofile'):
                        Alert.objects.create(
                            applicant=post.author.applicantprofile,
                            title="New Comment",
                            message=f"{request.user.username} commented on your post: '{post.title[:50]}...'"
                        )
                    elif hasattr(post.author, 'business_profile'):
                        BusinessAlert.objects.create(
                            business=post.author.business_profile,
                            title="New Comment",
                            message=f"{request.user.username} commented on your post: '{post.title[:50]}...'",
                            alert_type='comment'
                        )
                except Exception as notify_error:
                    logger.warning(f"Notification error: {notify_error}")
            
            # Also notify parent comment author if replying
            if parent_comment and parent_comment.author != request.user:
                try:
                    if hasattr(parent_comment.author, 'applicantprofile'):
                        Alert.objects.create(
                            applicant=parent_comment.author.applicantprofile,
                            title="Reply to Your Comment",
                            message=f"{request.user.username} replied to your comment on post: '{post.title[:50]}...'"
                        )
                    elif hasattr(parent_comment.author, 'business_profile'):
                        BusinessAlert.objects.create(
                            business=parent_comment.author.business_profile,
                            title="Reply to Your Comment",
                            message=f"{request.user.username} replied to your comment on post: '{post.title[:50]}...'",
                            alert_type='comment'
                        )
                except Exception as notify_error:
                    logger.warning(f"Notification error: {notify_error}")
            
            # Serialize the new comment
            from .serializers import CommentSerializer
            serializer = CommentSerializer(comment, context={'request': request})
            
            return Response({
                'success': True,
                'message': 'Comment added successfully',
                'comment': serializer.data,
                'comment_count': post.comment_count
            }, status=status.HTTP_201_CREATED)
            
    except Exception as e:
        logger.error(f"Error in post comments: {str(e)}", exc_info=True)
        return Response({
            'success': False,
            'error': 'Failed to process comments'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





@api_view(['GET'])
@permission_classes([AllowAny])
def api_post_stats(request):
    """Get post statistics for dashboard"""
    try:
        total_posts = Post.objects.filter(is_published=True).count()
        total_comments = Comment.objects.count()
        
        # Post type distribution
        post_types = Post.objects.filter(is_published=True).values('post_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Recent activity
        recent_posts = Post.objects.filter(is_published=True).order_by('-created_at')[:5]
        recent_posts_data = PostSerializer(
            recent_posts, 
            many=True, 
            context={'request': request}
        ).data
        
        # Most liked posts
        most_liked = Post.objects.filter(is_published=True).annotate(
            likes_count=Count('likes')
        ).order_by('-likes_count')[:5]
        
        most_liked_data = []
        for post in most_liked:
            most_liked_data.append({
                'id': post.id,
                'title': post.title,
                'author': post.author.username,
                'likes_count': post.likes.count(),
                'post_type': post.post_type
            })
        
        return Response({
            'success': True,
            'stats': {
                'total_posts': total_posts,
                'total_comments': total_comments,
                'post_type_distribution': list(post_types),
                'recent_posts': recent_posts_data,
                'most_liked_posts': most_liked_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting post stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load post statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_user_post_stats(request):
    """Get user-specific post statistics"""
    try:
        user = request.user
        
        # User's posts
        user_posts = Post.objects.filter(author=user)
        total_posts = user_posts.count()
        published_posts = user_posts.filter(is_published=True).count()
        
        # Engagement stats
        total_likes = 0
        total_comments = 0
        total_views = 0
        
        for post in user_posts:
            total_likes += post.likes.count()
            total_comments += post.comments.count()
            total_views += post.views
        
        # Recent posts
        recent_posts = user_posts.order_by('-created_at')[:5]
        recent_posts_data = PostSerializer(
            recent_posts, 
            many=True, 
            context={'request': request}
        ).data
        
        # Most popular post
        most_popular = user_posts.annotate(
            engagement=Count('likes') + Count('comments') + F('shares')
        ).order_by('-engagement').first()
        
        most_popular_data = None
        if most_popular:
            most_popular_data = {
                'id': most_popular.id,
                'title': most_popular.title,
                'engagement': most_popular.total_engagement(),
                'views': most_popular.views
            }
        
        return Response({
            'success': True,
            'stats': {
                'total_posts': total_posts,
                'published_posts': published_posts,
                'total_likes': total_likes,
                'total_comments': total_comments,
                'total_views': total_views,
                'recent_posts': recent_posts_data,
                'most_popular_post': most_popular_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting user post stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load your post statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# comment and likes sections
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_job_comments(request, job_id):
    """Get or add comments to a job posting - Using your existing Comment model"""
    try:
        job = get_object_or_404(JobListing, id=job_id)
        
        if request.method == 'GET':
            # Get query parameters
            try:
                page = max(1, int(request.GET.get('page', 1)))
                page_size = min(50, max(1, int(request.GET.get('page_size', 20))))
            except ValueError:
                return error_response('Invalid page or page_size parameter')
            
            # Get comments for this job
            comments = Comment.objects.filter(
                job_posting=job,  # Using your actual field name
                parent_comment__isnull=True  # Only top-level comments
            ).order_by('-created_at')
            
            # Get paginated comments
            paginated_data = get_paginated_data(
                queryset=comments,
                page=page,
                page_size=page_size,
                serializer_class=CommentSerializer,
                context={'request': request}
            )
            
            # Get total comment count
            total_comments = Comment.objects.filter(job_posting=job).count()
            
            return success_response('Comments loaded', {
                'comments': paginated_data['data'],
                'pagination': paginated_data['pagination'],
                'total_comments': total_comments,
                'job_id': job_id,
                'job_title': job.title
            })
        
        elif request.method == 'POST':
            # Validate comment content
            content = request.data.get('content', '').strip()
            if not content:
                return error_response('Comment cannot be empty')
            
            # Check for parent comment ID
            parent_comment_id = request.data.get('parent_comment_id')
            parent_comment = None
            
            if parent_comment_id:
                try:
                    parent_comment = Comment.objects.get(id=parent_comment_id, job_posting=job)
                except Comment.DoesNotExist:
                    return error_response('Parent comment not found')
            
            # Create the comment
            comment = Comment.objects.create(
                job_posting=job,
                author=request.user,
                content=content,
                parent_comment=parent_comment
            )
            
            # Create notification for job poster
            if hasattr(job, 'created_by') and job.created_by != request.user:
                try:
                    if hasattr(job.created_by, 'business_profile'):
                        BusinessAlert.objects.create(
                            business=job.created_by.business_profile,
                            title="New Comment on Job",
                            message=f"{request.user.username} commented on your job: '{job.title[:50]}...'",
                            alert_type='comment'
                        )
                except Exception as notify_error:
                    logger.warning(f"Notification error (non-critical): {notify_error}")
            
            # Also notify parent comment author if replying
            if parent_comment and parent_comment.author != request.user:
                try:
                    if hasattr(parent_comment.author, 'applicantprofile'):
                        Alert.objects.create(
                            applicant=parent_comment.author.applicantprofile,
                            title="Reply to Your Comment",
                            message=f"{request.user.username} replied to your comment on job: '{job.title[:50]}...'"
                        )
                    elif hasattr(parent_comment.author, 'business_profile'):
                        BusinessAlert.objects.create(
                            business=parent_comment.author.business_profile,
                            title="Reply to Your Comment",
                            message=f"{request.user.username} replied to your comment on job: '{job.title[:50]}...'",
                            alert_type='comment'
                        )
                except Exception as notify_error:
                    logger.warning(f"Notification error (non-critical): {notify_error}")
            
            return success_response(
                'Comment added successfully',
                {
                    'comment': CommentSerializer(comment, context={'request': request}).data
                },
                status_code=status.HTTP_201_CREATED
            )
            
    except Exception as e:
        logger.error(f"Error in job comments: {str(e)}", exc_info=True)
        return error_response(
            'Failed to process comments',
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def rate_limit_user(request, action, limit, period):
    """Rate limiting helper for all user types"""
    if not request.user.is_authenticated:
        return False, "Authentication required"
    
    user_id = request.user.id
    cache_key = f"rate_limit:{action}:{user_id}"
    
    # Get current usage
    current_usage = cache.get(cache_key, [])
    current_time = time.time()
    
    # Clean old entries
    current_usage = [timestamp for timestamp in current_usage 
                     if current_time - timestamp < period]
    
    # Check if limit exceeded
    if len(current_usage) >= limit:
        return True, f"Rate limit exceeded. Try again in {period} seconds"
    
    # Add current request
    current_usage.append(current_time)
    cache.set(cache_key, current_usage, period)
    
    return False, ""


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_job_like_dislike(request, job_id):
    """Like, dislike, or remove reaction from a job posting"""
    try:
        job = get_object_or_404(JobPosting, id=job_id)
        
        serializer = LikeDislikeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        action = serializer.validated_data['action']
        
        # Remove existing reactions
        job.liked_by.remove(request.user)
        job.disliked_by.remove(request.user)
        
        if action == 'like':
            job.liked_by.add(request.user)
            message = 'Job liked'
            
        elif action == 'dislike':
            job.disliked_by.add(request.user)
            message = 'Job disliked'
            
        else:  # remove
            message = 'Reaction removed'
        
        # Get updated counts
        likes_count = job.liked_by.count()
        dislikes_count = job.disliked_by.count()
        
        return Response({
            'success': True,
            'message': message,
            'likes_count': likes_count,
            'dislikes_count': dislikes_count,
            'user_has_liked': job.liked_by.filter(id=request.user.id).exists(),
            'user_has_disliked': job.disliked_by.filter(id=request.user.id).exists()
        })
        
    except Exception as e:
        logger.error(f"Error in job like/dislike: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process reaction'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def api_job_listings_with_interactions(request):
    """Get all job listings with interaction counts"""
    try:
        jobs = JobListing.objects.filter(status='published').order_by('-created_at')
        
        # Pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(jobs, 10)  # 10 jobs per page
        try:
            jobs_page = paginator.page(page)
        except PageNotAnInteger:
            jobs_page = paginator.page(1)
        except EmptyPage:
            jobs_page = paginator.page(paginator.num_pages)
        
        serializer = JobListingInteractionSerializer(
            jobs_page, 
            many=True, 
            context={'request': request}
        )
        
        return Response({
            'success': True,
            'jobs': serializer.data,
            'total_pages': paginator.num_pages,
            'current_page': jobs_page.number,
            'total_jobs': jobs.count()
        })
        
    except Exception as e:
        logger.error(f"Error loading jobs with interactions: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load jobs'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_job_interactions(request, job_id):
    """Get or add interactions (likes, dislikes, comments) for a job"""
    try:
        job = get_object_or_404(JobListing, id=job_id)
        
        if request.method == 'GET':
            # Get all interactions for this job
            interactions = JobInteraction.objects.filter(
                job_listing=job,
                parent_interaction__isnull=True  # Only get parent interactions
            ).order_by('-created_at')
            
            serializer = JobInteractionSerializer(
                interactions, 
                many=True, 
                context={'request': request}
            )
            
            # Get counts
            likes_count = JobInteraction.objects.filter(
                job_listing=job, 
                interaction_type='like'
            ).count()
            
            dislikes_count = JobInteraction.objects.filter(
                job_listing=job, 
                interaction_type='dislike'
            ).count()
            
            comments_count = JobInteraction.objects.filter(
                job_listing=job, 
                interaction_type='comment',
                parent_interaction__isnull=True
            ).count()
            
            return Response({
                'success': True,
                'interactions': serializer.data,
                'counts': {
                    'likes': likes_count,
                    'dislikes': dislikes_count,
                    'comments': comments_count
                }
            })
        
        elif request.method == 'POST':
            data = request.data.copy()
            data['user'] = request.user.id
            data['job_listing'] = job_id
            
            serializer = JobInteractionSerializer(
                data=data, 
                context={'request': request}
            )
            
            if serializer.is_valid():
                interaction = serializer.save()
                
                # For likes/dislikes, remove opposite reaction if exists
                interaction_type = data.get('interaction_type')
                if interaction_type in ['like', 'dislike']:
                    opposite_type = 'dislike' if interaction_type == 'like' else 'like'
                    
                    # Delete opposite reaction if exists
                    JobInteraction.objects.filter(
                        job_listing=job,
                        user=request.user,
                        interaction_type=opposite_type
                    ).delete()
                
                # Get updated counts
                likes_count = JobInteraction.objects.filter(
                    job_listing=job, 
                    interaction_type='like'
                ).count()
                
                dislikes_count = JobInteraction.objects.filter(
                    job_listing=job, 
                    interaction_type='dislike'
                ).count()
                
                return Response({
                    'success': True,
                    'message': f'Job {interaction_type}d successfully',
                    'interaction': serializer.data,
                    'counts': {
                        'likes': likes_count,
                        'dislikes': dislikes_count
                    }
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error in job interactions: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process interaction'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_job_remove_reaction(request, job_id):
    """Remove like/dislike from a job"""
    try:
        job = get_object_or_404(JobListing, id=job_id)
        
        interaction_type = request.data.get('interaction_type')
        
        if interaction_type not in ['like', 'dislike']:
            return Response({
                'success': False,
                'error': 'Invalid interaction type'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Delete the interaction
        deleted_count, _ = JobInteraction.objects.filter(
            job_listing=job,
            user=request.user,
            interaction_type=interaction_type
        ).delete()
        
        if deleted_count > 0:
            # Get updated counts
            likes_count = JobInteraction.objects.filter(
                job_listing=job, 
                interaction_type='like'
            ).count()
            
            dislikes_count = JobInteraction.objects.filter(
                job_listing=job, 
                interaction_type='dislike'
            ).count()
            
            return Response({
                'success': True,
                'message': f'Reaction removed successfully',
                'counts': {
                    'likes': likes_count,
                    'dislikes': dislikes_count
                }
            })
        else:
            return Response({
                'success': False,
                'error': 'No reaction found to remove'
            }, status=status.HTTP_404_NOT_FOUND)
        
    except Exception as e:
        logger.error(f"Error removing job reaction: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to remove reaction'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_job_comment_replies(request, interaction_id):
    """Get or add replies to a comment"""
    try:
        parent_comment = get_object_or_404(JobInteraction, id=interaction_id)
        
        if request.method == 'GET':
            replies = JobInteraction.objects.filter(
                parent_interaction=parent_comment
            ).order_by('created_at')
            
            serializer = JobInteractionSerializer(
                replies, 
                many=True, 
                context={'request': request}
            )
            
            return Response({
                'success': True,
                'replies': serializer.data
            })
        
        elif request.method == 'POST':
            data = request.data.copy()
            data['user'] = request.user.id
            data['job_listing'] = parent_comment.job_listing.id
            data['parent_interaction'] = interaction_id
            data['interaction_type'] = 'comment'  # Replies are always comments
            
            serializer = JobInteractionSerializer(
                data=data, 
                context={'request': request}
            )
            
            if serializer.is_valid():
                reply = serializer.save()
                return Response({
                    'success': True,
                    'message': 'Reply added successfully',
                    'reply': serializer.data
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error in comment replies: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process replies'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== FEED PAGE VIEW ====================
def feed_page(request):
    """Render the feed page"""
    context = {
        'user': request.user,
    }
    return render(request, 'feed_page.html', context)

# ==================== FEED API ENDPOINTS ====================
@api_view(['GET'])
@permission_classes([AllowAny])
def api_home_feed(request):
    """Get posts for home page feed with pagination and filtering - ALL users"""
    try:
        # Get and validate query parameters
        try:
            page = max(1, int(request.GET.get('page', 1)))
            page_size = min(100, max(1, int(request.GET.get('page_size', 10))))
        except ValueError:
            return error_response('Invalid page or page_size parameter')
        
        post_type = request.GET.get('type', 'all')
        sort_by = request.GET.get('sort', 'newest')
        search = request.GET.get('search', '').strip()
        
        # Allowed values
        ALLOWED_POST_TYPES = ['all', 'general', 'job-update', 'advice', 'question']
        ALLOWED_SORT_OPTIONS = ['newest', 'popular', 'top']
        
        # Validate parameters
        if post_type not in ALLOWED_POST_TYPES:
            return error_response(
                f'Invalid post type. Allowed: {", ".join(ALLOWED_POST_TYPES)}'
            )
        
        if sort_by not in ALLOWED_SORT_OPTIONS:
            return error_response(
                f'Invalid sort option. Allowed: {", ".join(ALLOWED_SORT_OPTIONS)}'
            )
        
        # Apply visibility filters for ALL users
        # For public feed, only show public posts
        posts = Post.objects.filter(is_published=True, visibility='public')
        
        # Filter by post type
        if post_type != 'all':
            posts = posts.filter(post_type=post_type)
        
        # Search filter (case-insensitive)
        if search:
            search_filter = (
                Q(title__icontains=search) |
                Q(content__icontains=search) |
                Q(tags__icontains=search) |
                Q(author__username__icontains=search) |
                Q(author__first_name__icontains=search) |
                Q(author__last_name__icontains=search)
            )
            posts = posts.filter(search_filter)
        
        # Apply sorting
        if sort_by == 'popular':
            # Calculate popularity score
            posts = posts.annotate(
                popularity_score=(
                    Count('likes') + 
                    Count('comments') * 2 + 
                    F('views') * 0.1 + 
                    F('shares') * 3
                )
            ).order_by('-popularity_score', '-created_at')
        elif sort_by == 'top':
            # Top posts based on rating
            posts = posts.filter(rating_count__gte=3).order_by('-average_rating', '-created_at')
        else:  # newest
            posts = posts.order_by('-created_at')
        
        # Apply pagination
        total_posts = posts.count()
        total_pages = (total_posts + page_size - 1) // page_size
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_posts = posts[start_index:end_index]
        
        # Serialize posts
        from .serializers import PostSerializer
        serializer = PostSerializer(
            paginated_posts, 
            many=True, 
            context={'request': request}
        )
        
        # Check which posts user has liked
        posts_data = serializer.data
        if request.user.is_authenticated:
            liked_post_ids = set(
                request.user.liked_posts.values_list('id', flat=True)
            )
            for post in posts_data:
                post['user_has_liked'] = post['id'] in liked_post_ids
        
        return Response({
            'success': True,
            'posts': posts_data,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_posts': total_posts,
                'page_size': page_size
            },
            'filters': {
                'current_type': post_type,
                'current_sort': sort_by,
                'search_query': search if search else None,
                'allowed_types': ALLOWED_POST_TYPES,
                'allowed_sort_options': ALLOWED_SORT_OPTIONS
            },
            'user_info': {
                'is_authenticated': request.user.is_authenticated,
                'username': request.user.username if request.user.is_authenticated else None,
                'user_type': getattr(request.user, 'user_type', None) if request.user.is_authenticated else None
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading home feed: {str(e)}", exc_info=True)
        return error_response(
            'Failed to load feed. Please try again later.',
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([AllowAny])
def api_feed_posts(request):
    """Get feed of all public posts (simplified version for home page)"""
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        # Get all public published posts
        posts = Post.objects.filter(
            is_published=True,
            visibility='public'
        ).order_by('-created_at')
        
        # Apply pagination
        total_posts = posts.count()
        total_pages = (total_posts + page_size - 1) // page_size
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_posts = posts[start_index:end_index]
        
        # Serialize posts
        from .serializers import PostSerializer
        serializer = PostSerializer(
            paginated_posts, 
            many=True, 
            context={'request': request}
        )
        
        # Check which posts user has liked
        posts_data = serializer.data
        if request.user.is_authenticated:
            liked_post_ids = set(
                request.user.liked_posts.values_list('id', flat=True)
            )
            for post in posts_data:
                post['user_has_liked'] = post['id'] in liked_post_ids
        
        return Response({
            'success': True,
            'posts': posts_data,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_posts': total_posts,
                'page_size': page_size
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading feed posts: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load feed'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def api_post_stats(request):
    """Get post statistics for dashboard"""
    try:
        total_posts = Post.objects.filter(is_published=True).count()
        total_comments = Comment.objects.count()
        
        # Post type distribution
        post_types = Post.objects.filter(is_published=True).values('post_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Recent activity
        recent_posts = Post.objects.filter(is_published=True).order_by('-created_at')[:5]
        
        from .serializers import PostSerializer
        recent_posts_data = PostSerializer(
            recent_posts, 
            many=True, 
            context={'request': request}
        ).data
        
        # Most liked posts
        most_liked = Post.objects.filter(is_published=True).annotate(
            likes_count=Count('likes')
        ).order_by('-likes_count')[:5]
        
        most_liked_data = []
        for post in most_liked:
            most_liked_data.append({
                'id': post.id,
                'title': post.title,
                'author': post.author.username,
                'likes_count': post.likes.count(),
                'post_type': post.post_type
            })
        
        return Response({
            'success': True,
            'stats': {
                'total_posts': total_posts,
                'total_comments': total_comments,
                'post_type_distribution': list(post_types),
                'recent_posts': recent_posts_data,
                'most_liked_posts': most_liked_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting post stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load post statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_user_post_stats(request):
    """Get user-specific post statistics"""
    try:
        user = request.user
        
        # User's posts
        user_posts = Post.objects.filter(author=user)
        total_posts = user_posts.count()
        published_posts = user_posts.filter(is_published=True).count()
        
        # Engagement stats
        total_likes = 0
        total_comments = 0
        total_views = 0
        
        for post in user_posts:
            total_likes += post.likes.count()
            total_comments += post.comments.count()
            total_views += post.views
        
        # Recent posts
        recent_posts = user_posts.order_by('-created_at')[:5]
        
        from .serializers import PostSerializer
        recent_posts_data = PostSerializer(
            recent_posts, 
            many=True, 
            context={'request': request}
        ).data
        
        # Most popular post
        most_popular = user_posts.annotate(
            engagement=Count('likes') + Count('comments') + F('shares')
        ).order_by('-engagement').first()
        
        most_popular_data = None
        if most_popular:
            most_popular_data = {
                'id': most_popular.id,
                'title': most_popular.title,
                'engagement': most_popular.total_engagement() if hasattr(most_popular, 'total_engagement') else 0,
                'views': most_popular.views
            }
        
        return Response({
            'success': True,
            'stats': {
                'total_posts': total_posts,
                'published_posts': published_posts,
                'total_likes': total_likes,
                'total_comments': total_comments,
                'total_views': total_views,
                'recent_posts': recent_posts_data,
                'most_popular_post': most_popular_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting user post stats: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load your post statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ==================== HELPER FUNCTIONS ====================
def error_response(message, status_code=status.HTTP_400_BAD_REQUEST):
    """Helper function for error responses"""
    return Response({
        'success': False,
        'error': message
    }, status=status_code)

def can_user_view_post(post, user):
    """Check if user can view a post"""
    if not user.is_authenticated:
        return post.visibility == 'public'
    
    if post.visibility == 'public':
        return True
    elif post.visibility == 'private':
        return user == post.author
    elif post.visibility == 'company':
        return user == post.author or (user.user_type == 'admin' and post.company and post.company.user == user)
    elif post.visibility == 'connections':
        return user == post.author  # You can add more logic here later
    return False

def get_paginated_data(queryset, page, page_size, serializer_class, context=None):
    """Helper to get paginated data"""
    total = queryset.count()
    total_pages = (total + page_size - 1) // page_size
    
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    
    paginated_items = queryset[start_index:end_index]
    serializer = serializer_class(paginated_items, many=True, context=context)
    
    return {
        'data': serializer.data,
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_items': total,
            'page_size': page_size
        }
    }

def get_user_visibility_filters(user):
    """Get visibility filters based on user authentication"""
    if not user.is_authenticated:
        return Q(visibility='public', is_published=True)
    
    # For authenticated users
    filters = Q(is_published=True) & (
        Q(visibility='public') |
        Q(author=user) |  # User's own posts
        Q(visibility='private', author=user) |
        Q(visibility='connections', author=user)  # Add connection logic later
    )
    
    # For business users, show company posts
    if hasattr(user, 'user_type') and user.user_type == 'admin':
        filters |= Q(visibility='company')
    
    return filters

# ==================== URL PATTERNS TO ADD ====================
@login_required
def posts_page(request):
    """Render the dedicated posts page"""
    return render(request, 'post.html')


class PasswordResetRequestView(APIView):
    """
    Simple password reset request view
    For development/testing - returns reset link in response
    """
    permission_classes = [AllowAny]
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            email = data.get('email', '').strip().lower()
            
            if not email:
                return Response({
                    'success': False,
                    'error': 'Email is required',
                    'errors': {'email': ['Email is required']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if user exists with this email - USE CustomUser
            try:
                user = CustomUser.objects.get(email=email, is_active=True)
            except CustomUser.DoesNotExist:
                # For security, don't reveal if email exists or not
                return Response({
                    'success': True,
                    'message': 'If your email exists in our system, you will receive a password reset link shortly.'
                }, status=status.HTTP_200_OK)
            
            # Generate a simple token (for development only)
            timestamp = int(time.time())
            token = hashlib.sha256(f"{user.id}{timestamp}{settings.SECRET_KEY}".encode()).hexdigest()[:40]
            full_token = f"{timestamp}-{token}"
            
            # Create reset link
            # Update this with your actual domain
            reset_url = f"http://localhost:8000/reset-password/?token={full_token}&email={email}"
            
            # For development: Log the reset link
            print(f"\n{'='*60}")
            print(f"PASSWORD RESET LINK (DEV MODE)")
            print(f"{'='*60}")
            print(f"User: {user.username} ({email})")
            print(f"Reset URL: {reset_url}")
            print(f"{'='*60}\n")
            
            # Store token in session for validation
            request.session[f'reset_token_{user.id}'] = token
            request.session[f'reset_token_timestamp_{user.id}'] = timestamp
            
            return Response({
                'success': True,
                'message': 'Password reset link generated successfully.',
                'reset_url': reset_url  # Only include in development!
            }, status=status.HTTP_200_OK)
            
        except json.JSONDecodeError:
            return Response({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Password reset request error: {str(e)}")
            return Response({
                'success': False,
                'error': 'An unexpected error occurred. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PasswordResetConfirmView(APIView):
    """
    Password reset confirmation view
    """
    permission_classes = [AllowAny]
    
    def validate_password(self, password):
        """Validate password strength"""
        errors = []
        
        if len(password) < 8:
            errors.append('Password must be at least 8 characters long.')
        
        if not any(char.isupper() for char in password):
            errors.append('Password must contain at least one uppercase letter.')
        
        if not any(char.islower() for char in password):
            errors.append('Password must contain at least one lowercase letter.')
        
        if not any(char.isdigit() for char in password):
            errors.append('Password must contain at least one number.')
        
        return errors
    
    def validate_token(self, token, user_id):
        """Validate reset token"""
        try:
            token_parts = token.split('-', 1)
            if len(token_parts) != 2:
                return False, "Invalid token format"
            
            timestamp = int(token_parts[0])
            token_value = token_parts[1]
            
            # Check if token is expired (24 hours)
            current_time = int(time.time())
            if current_time - timestamp > 86400:  # 24 hours in seconds
                return False, "Reset link has expired"
            
            # Check if token matches stored token
            stored_token = self.request.session.get(f'reset_token_{user_id}')
            stored_timestamp = self.request.session.get(f'reset_token_timestamp_{user_id}')
            
            if not stored_token or not stored_timestamp:
                return False, "Invalid reset token"
            
            if stored_token != token_value or stored_timestamp != timestamp:
                return False, "Invalid reset token"
            
            return True, "Valid token"
            
        except (ValueError, IndexError) as e:
            return False, "Invalid token format"
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            token = data.get('token', '')
            email = data.get('email', '').strip().lower()
            new_password = data.get('new_password', '')
            confirm_password = data.get('confirm_password', '')
            
            errors = {}
            
            # Validate required fields
            if not token:
                errors['token'] = ['Reset token is required']
            if not email:
                errors['email'] = ['Email is required']
            if not new_password:
                errors['new_password'] = ['New password is required']
            if not confirm_password:
                errors['confirm_password'] = ['Please confirm your password']
            
            if errors:
                return Response({
                    'success': False,
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if passwords match
            if new_password != confirm_password:
                errors['confirm_password'] = ['Passwords do not match']
                return Response({
                    'success': False,
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate password strength
            password_errors = self.validate_password(new_password)
            if password_errors:
                errors['new_password'] = password_errors
                return Response({
                    'success': False,
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get user - USE CustomUser
            try:
                user = CustomUser.objects.get(email=email, is_active=True)
            except CustomUser.DoesNotExist:
                errors['email'] = ['Invalid email address']
                return Response({
                    'success': False,
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate token
            is_valid, token_message = self.validate_token(token, user.id)
            if not is_valid:
                errors['token'] = [token_message]
                return Response({
                    'success': False,
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Set new password
            try:
                user.set_password(new_password)
                user.save()
                
                # Clear the reset token from session
                request.session.pop(f'reset_token_{user.id}', None)
                request.session.pop(f'reset_token_timestamp_{user.id}', None)
                
                logger.info(f"Password reset successful for user: {user.username}")
                
                return Response({
                    'success': True,
                    'message': 'Password has been reset successfully. You can now login with your new password.'
                }, status=status.HTTP_200_OK)
                
            except Exception as save_error:
                logger.error(f"Failed to save new password for user {user.username}: {str(save_error)}")
                return Response({
                    'success': False,
                    'error': 'Failed to update password. Please try again.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        except json.JSONDecodeError:
            return Response({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Password reset confirm error: {str(e)}")
            return Response({
                'success': False,
                'error': 'An unexpected error occurred. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===== VIDEO FEED VIEWS =====

@api_view(['GET'])
@permission_classes([AllowAny])
def api_video_feed(request):
    """Get TikTok-style video feed"""
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        sort_by = request.GET.get('sort', 'newest')
        
        # Base queryset - only published videos
        videos = Video.objects.filter(is_published=True)
        
        # Apply privacy filters
        if request.user.is_authenticated:
            videos = videos.filter(
                Q(privacy='public') |
                Q(author=request.user)
            )
        else:
            videos = videos.filter(privacy='public')
        
        # Apply sorting
        if sort_by == 'popular':
            videos = videos.annotate(
                popularity=(
                    Count('likes') + 
                    Count('comments') * 2 + 
                    F('shares') * 3 + 
                    F('views') * 0.1
                )
            ).order_by('-popularity')
        elif sort_by == 'trending':
            yesterday = timezone.now() - timedelta(days=1)
            videos = videos.filter(created_at__gte=yesterday).annotate(
                engagement=Count('likes') + Count('comments') * 2 + F('shares') * 3
            ).order_by('-engagement')
        else:  # newest
            videos = videos.order_by('-created_at')
        
        # Pagination
        paginator = Paginator(videos, page_size)
        try:
            videos_page = paginator.page(page)
        except PageNotAnInteger:
            videos_page = paginator.page(1)
        except EmptyPage:
            videos_page = paginator.page(paginator.num_pages)
        
        # Serialize
        serializer = VideoSerializer(
            videos_page, 
            many=True, 
            context={'request': request}
        )
        
        return Response({
            'success': True,
            'videos': serializer.data,
            'pagination': {
                'current_page': videos_page.number,
                'total_pages': paginator.num_pages,
                'total_videos': paginator.count,
                'page_size': page_size,
                'has_next': videos_page.has_next(),
                'has_previous': videos_page.has_previous(),
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading video feed: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load video feed'
        }, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def api_video_detail(request, video_id):
    """Get single video details"""
    try:
        video = get_object_or_404(Video, id=video_id, is_published=True)
        
        # Check privacy
        if video.privacy != 'public' and request.user != video.author:
            if not request.user.is_authenticated:
                return Response({
                    'success': False,
                    'error': 'This video is private'
                }, status=403)
        
        # Increment views
        video.increment_views()
        
        serializer = VideoSerializer(video, context={'request': request})
        
        # Get more videos from same author
        more_videos = Video.objects.filter(
            author=video.author,
            is_published=True,
            privacy='public'
        ).exclude(id=video.id)[:5]
        
        more_serializer = VideoSerializer(
            more_videos, 
            many=True, 
            context={'request': request}
        )
        
        return Response({
            'success': True,
            'video': serializer.data,
            'more_from_author': more_serializer.data,
            'author': {
                'id': video.author.id,
                'username': video.author.username,
                'display_name': video.author.get_full_name() or video.author.username,
                'video_count': Video.objects.filter(author=video.author, is_published=True).count()
            }
        })
        
    except Video.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Video not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error loading video detail: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load video'
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def api_video_upload(request):
    """Upload a new video"""
    try:
        # Validate video file
        if 'video_file' not in request.FILES:
            return Response({
                'success': False,
                'error': 'Video file is required'
            }, status=400)
        
        video_file = request.FILES['video_file']
        
        # Validate file size (max 100MB)
        if video_file.size > 100 * 1024 * 1024:
            return Response({
                'success': False,
                'error': 'Video file size exceeds 100MB limit'
            }, status=400)
        
        # Validate file type
        valid_extensions = ['mp4', 'mov', 'avi', 'webm']
        file_ext = video_file.name.split('.')[-1].lower()
        if file_ext not in valid_extensions:
            return Response({
                'success': False,
                'error': f'Invalid file type. Supported: {", ".join(valid_extensions)}'
            }, status=400)
        
        # Create video
        data = request.data.copy()
        data['author'] = request.user.id
        
        serializer = VideoCreateSerializer(
            data=data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            video = serializer.save()
            
            # Create notification
            try:
                if hasattr(request.user, 'applicantprofile'):
                    Alert.objects.create(
                        applicant=request.user.applicantprofile,
                        title="Video Uploaded",
                        message=f"Your video '{video.title or 'Untitled'}' has been uploaded successfully!"
                    )
            except Exception as e:
                logger.warning(f"Notification error: {e}")
            
            return Response({
                'success': True,
                'message': 'Video uploaded successfully!',
                'video': VideoSerializer(video, context={'request': request}).data
            }, status=201)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=400)
        
    except Exception as e:
        logger.error(f"Error uploading video: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to upload video'
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_video_like(request, video_id):
    """Like or unlike a video"""
    try:
        video = get_object_or_404(Video, id=video_id, is_published=True)
        
        # Check privacy
        if video.privacy != 'public' and request.user != video.author:
            return Response({
                'success': False,
                'error': 'You cannot interact with this video'
            }, status=403)
        
        action = request.data.get('action', 'like')
        
        if action == 'like':
            video.likes.add(request.user)
            liked = True
            message = 'Video liked'
        else:
            video.likes.remove(request.user)
            liked = False
            message = 'Video unliked'
        
        return Response({
            'success': True,
            'message': message,
            'liked': liked,
            'likes_count': video.likes.count()
        })
        
    except Video.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Video not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error toggling video like: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process like'
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_video_share(request, video_id):
    """Share a video"""
    try:
        video = get_object_or_404(Video, id=video_id, is_published=True)
        video.increment_shares()
        
        return Response({
            'success': True,
            'message': 'Video shared successfully!',
            'shares_count': video.shares
        })
        
    except Video.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Video not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error sharing video: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to share video'
        }, status=500)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_video_comments(request, video_id):
    """Get or add comments to a video"""
    try:
        video = get_object_or_404(Video, id=video_id, is_published=True)
        
        # Check privacy
        if video.privacy != 'public' and request.user != video.author:
            return Response({
                'success': False,
                'error': 'You cannot comment on this video'
            }, status=403)
        
        if request.method == 'GET':
            # Get comments
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 20))
            
            comments = VideoComment.objects.filter(
                video=video,
                is_active=True,
                parent_comment__isnull=True
            ).order_by('-created_at')
            
            paginator = Paginator(comments, page_size)
            try:
                comments_page = paginator.page(page)
            except PageNotAnInteger:
                comments_page = paginator.page(1)
            except EmptyPage:
                comments_page = paginator.page(paginator.num_pages)
            
            serializer = VideoCommentSerializer(
                comments_page,
                many=True,
                context={'request': request}
            )
            
            return Response({
                'success': True,
                'comments': serializer.data,
                'pagination': {
                    'current_page': comments_page.number,
                    'total_pages': paginator.num_pages,
                    'total_comments': paginator.count,
                    'page_size': page_size
                }
            })
        
        elif request.method == 'POST':
            # Add comment
            content = request.data.get('content', '').strip()
            parent_id = request.data.get('parent_comment_id')
            
            if not content:
                return Response({
                    'success': False,
                    'error': 'Comment content is required'
                }, status=400)
            
            parent_comment = None
            if parent_id:
                try:
                    parent_comment = VideoComment.objects.get(
                        id=parent_id,
                        video=video,
                        is_active=True
                    )
                except VideoComment.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'Parent comment not found'
                    }, status=404)
            
            comment = VideoComment.objects.create(
                video=video,
                author=request.user,
                content=content,
                parent_comment=parent_comment
            )
            
            # Create notification for video author
            if video.author != request.user:
                try:
                    if hasattr(video.author, 'applicantprofile'):
                        Alert.objects.create(
                            applicant=video.author.applicantprofile,
                            title="New Comment",
                            message=f"{request.user.username} commented on your video: '{video.title or 'Untitled'}'"
                        )
                except Exception as e:
                    logger.warning(f"Notification error: {e}")
            
            serializer = VideoCommentSerializer(
                comment,
                context={'request': request}
            )
            
            return Response({
                'success': True,
                'message': 'Comment added successfully!',
                'comment': serializer.data
            }, status=201)
            
    except Video.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Video not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error in video comments: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process comments'
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_comment_like(request, comment_id):
    """Like or unlike a comment"""
    try:
        comment = get_object_or_404(VideoComment, id=comment_id, is_active=True)
        action = request.data.get('action', 'like')
        
        if action == 'like':
            comment.likes.add(request.user)
            liked = True
        else:
            comment.likes.remove(request.user)
            liked = False
        
        return Response({
            'success': True,
            'liked': liked,
            'likes_count': comment.likes.count()
        })
        
    except VideoComment.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Comment not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error toggling comment like: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to process like'
        }, status=500)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_comment(request, comment_id):
    """Delete a comment (soft delete)"""
    try:
        comment = get_object_or_404(VideoComment, id=comment_id)
        
        # Only author or video author can delete
        if comment.author != request.user and comment.video.author != request.user:
            return Response({
                'success': False,
                'error': 'You do not have permission to delete this comment'
            }, status=403)
        
        comment.is_active = False
        comment.save()
        
        return Response({
            'success': True,
            'message': 'Comment deleted successfully'
        })
        
    except VideoComment.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Comment not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error deleting comment: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to delete comment'
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_my_videos(request):
    """Get current user's videos"""
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        videos = Video.objects.filter(author=request.user).order_by('-created_at')
        
        paginator = Paginator(videos, page_size)
        try:
            videos_page = paginator.page(page)
        except PageNotAnInteger:
            videos_page = paginator.page(1)
        except EmptyPage:
            videos_page = paginator.page(paginator.num_pages)
        
        serializer = VideoSerializer(
            videos_page,
            many=True,
            context={'request': request}
        )
        
        return Response({
            'success': True,
            'videos': serializer.data,
            'pagination': {
                'current_page': videos_page.number,
                'total_pages': paginator.num_pages,
                'total_videos': paginator.count,
                'page_size': page_size
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading user videos: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to load videos'
        }, status=500)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_video(request, video_id):
    """Delete a video"""
    try:
        video = get_object_or_404(Video, id=video_id, author=request.user)
        
        # Delete video file
        if video.video_file:
            video.video_file.delete(save=False)
        if video.thumbnail:
            video.thumbnail.delete(save=False)
        
        video.delete()
        
        return Response({
            'success': True,
            'message': 'Video deleted successfully'
        })
        
    except Video.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Video not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error deleting video: {str(e)}")
        return Response({
            'success': False,
            'error': 'Failed to delete video'
        }, status=500)

def pwa_manifest(request):
    """Serve PWA manifest"""
    # Try multiple possible paths
    possible_paths = [
        os.path.join(settings.BASE_DIR, 'hiring', 'static', 'hiring', 'manifest.json'),
        os.path.join('hiring', 'static', 'hiring', 'manifest.json'),
        os.path.join(settings.STATIC_ROOT, 'hiring', 'manifest.json'),
    ]
    
    for path in possible_paths:
        print(f"Checking: {path}")  # Check console to see where it's looking
        if os.path.exists(path):
            print(f"✅ FOUND at: {path}")
            with open(path, 'r') as f:
                manifest_data = json.load(f)
            return JsonResponse(manifest_data, content_type='application/json')
    
    print("❌ Manifest not found in any location")
    return JsonResponse({'error': 'Manifest not found'}, status=404)

def pwa_sw(request):
    """Serve service worker"""
    sw_path = os.path.join(settings.BASE_DIR, 'hiring', 'static', 'hiring', 'js', 'sw.js')
    print(f"Checking SW: {sw_path}")  # Check console
    
    try:
        with open(sw_path, 'r') as f:
            sw_content = f.read()
        return HttpResponse(sw_content, content_type='application/javascript')
    except FileNotFoundError:
        print(f"❌ SW not found at: {sw_path}")
        return HttpResponse('', status=404)

def offline_page(request):
    """Offline page for PWA"""
    return render(request, 'hiring/offline.html', {
        'title': 'You are offline'
    })

@csrf_exempt
def subscribe_to_notifications(request):
    """Handle push notification subscription"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        subscription = data.get('subscription')
        
        if not subscription:
            return JsonResponse({'error': 'Subscription data required'}, status=400)
        
        # Store subscription in database
        print(f"[PWA] Push subscription received: {subscription}")
        
        return JsonResponse({
            'success': True,
            'message': 'Subscription saved successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)